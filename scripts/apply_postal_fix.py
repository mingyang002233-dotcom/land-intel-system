#!/usr/bin/env python3
"""
apply_postal_fix.py  v4.2
只補郵遞區號（信心高 / 可自動補），不動其他欄位。
不覆蓋正式主清冊，輸出新檔。
"""
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

MASTER_SRC   = '/Users/xiaomingyang/Desktop/excel土地資料維護/土地主清冊_正式版_20260522.xlsx'
REPORT_SRC   = '/Users/xiaomingyang/Desktop/excel土地資料維護/資料品質檢查報表_20260522_165501.xlsx'
MASTER_DEST  = '/Users/xiaomingyang/Desktop/excel土地資料維護/土地主清冊_正式版_20260522_郵遞區號補正版.xlsx'
LOG_DEST     = f'/Users/xiaomingyang/Desktop/excel土地資料維護/郵遞區號補正紀錄_20260522.xlsx'

# ── 1. 讀取品質報表，建立 row_number → 建議郵遞區號 對照表 ──────────────────
print('[1] 讀取品質報表...')
rp_wb = openpyxl.load_workbook(REPORT_SRC, read_only=True, data_only=True)
rp_ws = rp_wb['資料品質報表']
rp_headers = [c.value for c in next(rp_ws.iter_rows(min_row=1, max_row=1))]
rp_col = {h: i for i, h in enumerate(rp_headers)}

fix_map = {}   # row_number (int) → {'suggested': str, 'original': str, 'address': str}
skipped = 0

for row in rp_ws.iter_rows(min_row=2, values_only=True):
    issue_type = str(row[rp_col['問題類型']] or '')
    action     = str(row[rp_col['建議動作']] or '')
    conf       = str(row[rp_col['信心等級']] or '')
    row_number = row[rp_col['row_number']]
    suggested  = str(row[rp_col['建議值']] or '').strip()
    original   = str(row[rp_col['郵遞區號']] or '').strip()
    address    = str(row[rp_col['住址']] or '').strip()

    if not issue_type.startswith('B'):
        continue  # 只處理郵遞區號問題
    if action != '可自動補' or conf != '高':
        skipped += 1
        continue
    if not suggested or not re.match(r'^\d{3}$', suggested):
        skipped += 1
        continue
    if row_number not in fix_map:
        fix_map[int(row_number)] = {'suggested': suggested, 'original': original, 'address': address}

rp_wb.close()
print(f'    可自動補：{len(fix_map):,} 筆　跳過（低信心/非B類）：{skipped:,} 筆')

# ── 2. 讀取主清冊（完整讀入記憶體）──────────────────────────────────────────
print('[2] 讀取主清冊...')
m_wb = openpyxl.load_workbook(MASTER_SRC, read_only=True, data_only=True)
m_ws = m_wb.active
all_rows = list(m_ws.iter_rows(values_only=True))
m_wb.close()
headers = list(all_rows[0])
col = {h: i for i, h in enumerate(headers)}
postal_col_idx = col.get('郵遞區號')
print(f'    {len(all_rows)-1:,} 資料列 × {len(headers)} 欄')
print(f'    郵遞區號欄位索引：{postal_col_idx}')

# ── 3. 套用補正（修改 all_rows 副本）────────────────────────────────────────
print('[3] 套用郵遞區號補正...')
applied = 0
log_entries = []  # (row_number, 縣市, 地區, 地段, 地號, 所有權人, 原始郵遞區號, 補正值, 住址)

for r_idx, data_row in enumerate(all_rows[1:], start=2):
    if r_idx not in fix_map:
        continue
    fix = fix_map[r_idx]
    row_list = list(data_row)
    row_list[postal_col_idx] = fix['suggested']
    all_rows[r_idx - 1] = tuple(row_list)
    applied += 1
    log_entries.append((
        r_idx,
        data_row[col.get('縣市', 0)],
        data_row[col.get('地區', 0)],
        data_row[col.get('地段', 0)],
        data_row[col.get('地號', 0)],
        data_row[col.get('所有權人', 0)],
        fix['original'],
        fix['suggested'],
        fix['address'],
    ))

print(f'    實際補正：{applied:,} 筆')

# ── 4. 寫出補正版主清冊 ──────────────────────────────────────────────────────
print('[4] 寫出補正版主清冊...')
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
EVEN_FILL   = PatternFill('solid', fgColor='EEF3FB')
FIX_FILL    = PatternFill('solid', fgColor='E2EFDA')   # 綠底標示補正列
DATA_FONT   = Font(name='微軟正黑體', size=10)
TEXT_COLS   = {'地號', '電話', '統一編號（遮罩）', '統一編號（完整）', '郵遞區號'}

fixed_rows = set(fix_map.keys())   # 補正列 row_number set

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '土地主清冊'
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(1, c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')

postal_col_1based = postal_col_idx + 1

for r_idx, row in enumerate(all_rows[1:], start=2):
    ws.append(list(row))
    is_fixed = r_idx in fixed_rows
    row_fill = FIX_FILL if is_fixed else (EVEN_FILL if r_idx % 2 == 0 else None)
    for c_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(r_idx, c_idx)
        cell.font = DATA_FONT
        if row_fill:
            cell.fill = row_fill
        if hdr in TEXT_COLS:
            cell.number_format = '@'
    if r_idx % 30000 == 0:
        print(f'    {r_idx-1:,} / {len(all_rows)-1:,}')

COL_WIDTHS = [12, 8, 14, 6, 8, 16, 12, 10, 6, 12, 10, 12,
              12, 16, 16, 8, 30, 6, 6, 6, 12, 10, 40, 20]
for i, w in enumerate(COL_WIDTHS[:len(headers)], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

wb.save(MASTER_DEST)
size_mb = os.path.getsize(MASTER_DEST) / 1024 / 1024
print(f'    已儲存：{MASTER_DEST}  ({size_mb:.1f} MB)')

# ── 5. 補正紀錄 ──────────────────────────────────────────────────────────────
print('[5] 寫出補正紀錄...')
log_wb = openpyxl.Workbook()
log_ws = log_wb.active
log_ws.title = '郵遞區號補正紀錄'
LOG_COLS = ['row_number', '縣市', '地區', '地段', '地號', '所有權人',
            '原始郵遞區號', '補正後郵遞區號', '住址']
log_ws.append(LOG_COLS)
for c in range(1, len(LOG_COLS)+1):
    log_ws.cell(1, c).font = Font(bold=True)
for entry in log_entries:
    log_ws.append(list(entry))
log_ws.auto_filter.ref = log_ws.dimensions
log_ws.column_dimensions['A'].width = 10
log_ws.column_dimensions['B'].width = 8
log_ws.column_dimensions['C'].width = 8
log_ws.column_dimensions['D'].width = 14
log_ws.column_dimensions['E'].width = 12
log_ws.column_dimensions['F'].width = 14
log_ws.column_dimensions['G'].width = 10
log_ws.column_dimensions['H'].width = 10
log_ws.column_dimensions['I'].width = 40

# 摘要頁
log_ws2 = log_wb.create_sheet('摘要')
ts_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
summary = [
    ('執行時間', ts_str),
    ('來源主清冊', MASTER_SRC),
    ('品質報表', REPORT_SRC),
    ('輸出主清冊', MASTER_DEST),
    ('輸出補正紀錄', LOG_DEST),
    ('', ''),
    ('可自動補筆數（報表）', len(fix_map)),
    ('實際補正筆數', applied),
    ('跳過筆數（低信心/建議值異常）', skipped),
    ('', ''),
    ('補正原則', '只補郵遞區號，信心=高，建議動作=可自動補'),
    ('未修改項目', '所有權人、電話、地址、SQLite、原始正式主清冊'),
]
for row in summary:
    log_ws2.append(list(row))
log_ws2.column_dimensions['A'].width = 28
log_ws2.column_dimensions['B'].width = 60

log_wb.save(LOG_DEST)
log_size = os.path.getsize(LOG_DEST) / 1024 / 1024

print(f'\n── 完成 ──────────────────────────────────')
print(f'實際補正筆數  ：{applied:,}')
print(f'跳過筆數      ：{skipped:,}')
print(f'補正版主清冊  ：{MASTER_DEST}')
print(f'補正紀錄      ：{LOG_DEST}  ({log_size:.1f} MB)')
print(f'原始正式版    ：未修改')
print(f'SQLite        ：未修改')
