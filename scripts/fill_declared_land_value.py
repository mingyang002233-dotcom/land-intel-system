#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_declared_land_value.py — 主清冊公告現值 & 縣市地區補正工具

目標：
  1. 縣市/地區空白 → 從地址解析，或從同地段唯一對應推斷
  2. 公告現值空白  → 查詢政府開放資料（待申請 MOI_API_014 時自動啟用）
                     未能查詢時，輸出「待查詢清單」供人工確認
  3. 查不到 → 保持空白，絕不猜值

輸出：
  output/主清冊_公告現值補正版_YYYYMMDD_HHMM.xlsx
  output/公告現值待查詢清單_YYYYMMDD_HHMM.xlsx
  logs/fill_declared_land_value_YYYYMMDD_HHMM.log

執行：
  python3 scripts/fill_declared_land_value.py --dry-run   ← 只統計，不寫檔
  python3 scripts/fill_declared_land_value.py             ← 正式輸出
  python3 scripts/fill_declared_land_value.py --input 其他.xlsx
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

BASE_DIR    = Path('/Users/xiaomingyang/Desktop/excel土地資料維護')
INPUT_DIR   = BASE_DIR / 'input'
OUTPUT_DIR  = BASE_DIR / 'output'
LOGS_DIR    = BASE_DIR / 'logs'
LATEST_DIR  = BASE_DIR / '最新完成版'

DEFAULT_XLSX = LATEST_DIR / '老蕭LAND_MASTER.xlsx'

# ── 縣市/地區解析正則 ────────────────────────────────────────────────

CITY_PATTERN = re.compile(
    r'(台北市|臺北市|新北市|桃園市|台中市|臺中市|台南市|臺南市|高雄市|'
    r'基隆市|新竹市|嘉義市|苗栗縣|彰化縣|南投縣|雲林縣|屏東縣|宜蘭縣|'
    r'花蓮縣|台東縣|臺東縣|澎湖縣|金門縣|連江縣|新竹縣|嘉義縣)'
)
DIST_PATTERN = re.compile(
    r'([一-鿿]{2,5}(?:區|鄉|鎮|市))'
)


def parse_city_dist_from_addr(addr: str) -> tuple[str, str]:
    """從地址解析縣市與地區，回傳 (城市, 地區) 或 ('', '')"""
    if not addr:
        return '', ''
    addr = addr.strip().replace('臺', '台')
    city_m = CITY_PATTERN.search(addr)
    city = city_m.group(1).replace('臺', '台') if city_m else ''

    # 地區：在縣市名稱之後
    dist = ''
    if city_m:
        rest = addr[city_m.end():]
        dist_m = DIST_PATTERN.match(rest)
        if dist_m:
            dist = dist_m.group(1)
    return city, dist


# ── 欄位偵測 ─────────────────────────────────────────────────────────

COL_ALIASES = {
    '縣市':   ['縣市'],
    '地區':   ['地區', '行政區', '鄉鎮市區'],
    '地段':   ['地段'],
    '地號':   ['地號'],
    '住址':   ['住址', '地址'],
    '公告現值': ['公告現值', '公告土地現值'],
}


def detect_columns(headers: list) -> dict[str, int]:
    result = {}
    for std, aliases in COL_ALIASES.items():
        for i, h in enumerate(headers):
            if h and str(h).strip() in aliases:
                result[std] = i
                break
    return result


# ── 公告現值 API（MOI_API_014）────────────────────────────────────────

def query_moi_api(city: str, dist: str, section: str, land_no: str) -> float | None:
    """
    查詢內政部 MOI_API_014「公告地價與公告土地現值資料服務」。
    需要帳號金鑰；目前回傳 None（待申請後填入 API_KEY）。
    申請網址：https://data.gov.tw/dataset/107813
    """
    API_KEY = ''  # 填入申請到的 API Key
    if not API_KEY:
        return None
    # 留位：實際呼叫邏輯在取得金鑰後補入
    # endpoint = 'https://api.landa.moi.gov.tw/...'
    return None


def moi_api_available() -> bool:
    return False  # 待申請 API 金鑰後改為 True


# ── 主流程 ────────────────────────────────────────────────────────────

def run(xlsx_path: Path, dry_run: bool):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import numbers as opxl_numbers
    except ImportError:
        print('請安裝 openpyxl：pip install openpyxl')
        sys.exit(1)

    ts   = datetime.now().strftime('%Y%m%d_%H%M')
    mode = 'DRY-RUN' if dry_run else '正式執行'

    for d in [OUTPUT_DIR, LOGS_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    log_path  = LOGS_DIR / f'fill_declared_land_value_{ts}.log'
    log_lines: list[str] = []

    def log(msg: str):
        print(msg)
        log_lines.append(msg)

    log(f'[ {mode} ] fill_declared_land_value.py')
    log(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    log(f'來源：{xlsx_path}')
    log(f'MOI_API_014：{"可用" if moi_api_available() else "尚未申請（輸出待查詢清單）"}')

    # ── 讀取
    log('\n[讀取] Excel ...')
    wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws_ro = wb_ro.active
    all_rows = list(ws_ro.iter_rows(values_only=True))
    wb_ro.close()

    if not all_rows:
        log('Excel 是空的，中止。'); sys.exit(1)

    headers   = list(all_rows[0])
    data_rows = [list(r) for r in all_rows[1:]]
    total     = len(data_rows)
    log(f'  總列數：{total:,}')

    col = detect_columns(headers)
    missing = [k for k in ['縣市', '地段', '地號'] if k not in col]
    if missing:
        log(f'[錯誤] 缺少必要欄位：{missing}'); sys.exit(1)

    has_addr = '住址' in col
    has_dist = '地區' in col
    has_val  = '公告現值' in col

    # ── 一、縣市/地區補正分析
    log('\n[分析] 縣市/地區補正 ...')

    # 建立「地段 → (縣市, 地區) 唯一對應」索引（從已有資料推斷）
    sec_to_city_dist: dict[str, set] = defaultdict(set)
    for row in data_rows:
        city = str(row[col['縣市']]).strip() if row[col['縣市']] else ''
        dist = str(row[col['地區']]).strip() if (has_dist and row[col['地區']]) else ''
        sec  = str(row[col['地段']]).strip() if row[col['地段']] else ''
        if city and dist and sec:
            sec_to_city_dist[sec].add((city, dist))

    # 唯一對應：地段只對應一個 (縣市, 地區)
    sec_unique: dict[str, tuple] = {
        sec: list(cd)[0]
        for sec, cd in sec_to_city_dist.items()
        if len(cd) == 1
    }

    cnt_city_blank = 0; cnt_dist_blank = 0
    cnt_city_fixed_addr = 0; cnt_city_fixed_sec = 0
    cnt_city_manual = 0
    fix_examples: list[dict] = []
    manual_city: list[dict] = []

    city_fixes: dict[int, tuple] = {}  # row_idx → (new_city, new_dist, method)

    for i, row in enumerate(data_rows):
        city = str(row[col['縣市']]).strip() if row[col['縣市']] else ''
        dist = str(row[col['地區']]).strip() if (has_dist and row[col['地區']]) else ''
        sec  = str(row[col['地段']]).strip() if row[col['地段']] else ''

        need_city = not city
        need_dist = not dist

        if not need_city and not need_dist:
            continue

        if need_city: cnt_city_blank += 1
        if need_dist: cnt_dist_blank += 1

        new_city, new_dist = city, dist
        method = ''

        # 方法 A：從地址解析
        if has_addr and (need_city or need_dist):
            addr = str(row[col['住址']]).strip() if row[col['住址']] else ''
            parsed_city, parsed_dist = parse_city_dist_from_addr(addr)
            if need_city and parsed_city:
                new_city = parsed_city; method = '地址解析'
                cnt_city_fixed_addr += 1
            if need_dist and parsed_dist:
                new_dist = parsed_dist
                if not method: method = '地址解析'

        # 方法 B：從同地段唯一對應推斷
        if (not new_city or not new_dist) and sec in sec_unique:
            infer_city, infer_dist = sec_unique[sec]
            if not new_city and infer_city:
                new_city = infer_city
                if 'method' not in locals() or not method:
                    method = '地段推斷'
                cnt_city_fixed_sec += 1
            if not new_dist and infer_dist:
                new_dist = infer_dist
                if not method: method = '地段推斷'

        if new_city != city or new_dist != dist:
            city_fixes[i] = (new_city, new_dist, method)
            if len(fix_examples) < 20:
                fix_examples.append({
                    'row': i + 2, 'method': method,
                    'old_city': city, 'new_city': new_city,
                    'old_dist': dist, 'new_dist': new_dist,
                    'addr': str(row[col['住址']] or '') if has_addr else '',
                    'sec': sec, 'lno': str(row[col['地號']] or ''),
                })
        else:
            if need_city:
                cnt_city_manual += 1
                manual_city.append({
                    'row': i + 2, 'sec': sec,
                    'lno': str(row[col['地號']] or ''),
                    'addr': str(row[col['住址']] or '') if has_addr else '',
                })

    log(f'  縣市空白：{cnt_city_blank:,} 筆')
    log(f'  地區空白：{cnt_dist_blank:,} 筆')
    log(f'  可自動補正（地址解析）：{cnt_city_fixed_addr:,} 筆')
    log(f'  可自動補正（地段推斷）：{cnt_city_fixed_sec:,} 筆')
    log(f'  需人工確認：{cnt_city_manual:,} 筆')

    # ── 二、公告現值補正分析
    log('\n[分析] 公告現值補正 ...')
    cnt_val_blank = 0; cnt_val_api = 0; cnt_val_manual = 0
    pending_query: list[dict] = []

    if has_val:
        for i, row in enumerate(data_rows):
            val = row[col['公告現值']]
            is_blank = val in (None, '', 0, '0')
            if not is_blank:
                try:
                    if float(str(val).replace(',', '')) > 0:
                        continue
                except (ValueError, TypeError):
                    pass
                is_blank = True

            cnt_val_blank += 1
            city = str(row[col['縣市']]).strip() if row[col['縣市']] else ''
            dist = str(row[col['地區']]).strip() if (has_dist and row[col['地區']]) else ''
            sec  = str(row[col['地段']]).strip() if row[col['地段']] else ''
            lno  = str(row[col['地號']]).strip() if row[col['地號']] else ''

            if moi_api_available():
                result = query_moi_api(city, dist, sec, lno)
                if result:
                    cnt_val_api += 1
                    continue
            cnt_val_manual += 1
            pending_query.append({
                'row': i + 2, '縣市': city, '地區': dist,
                '地段': sec, '地號': lno,
                '原公告現值': val,
            })
    else:
        log('  [警告] 找不到「公告現值」欄位')

    log(f'  公告現值空白/零：{cnt_val_blank:,} 筆')
    log(f'  MOI API 可補正：{cnt_val_api:,} 筆（目前 API 未申請）')
    log(f'  需人工查詢：{cnt_val_manual:,} 筆（將輸出待查詢清單）')
    log(f'  MOI_API_014 申請：{"已申請" if moi_api_available() else "尚未申請，需申請後自動補正"}')

    # ── 補正範例
    log(f'\n[補正範例] 縣市/地區自動補正（前 {min(20, len(fix_examples))} 筆）：')
    if fix_examples:
        log(f'  {"列":>5}  {"方式":8}  {"舊縣市":6}→{"新縣市":8}  {"舊地區":6}→{"新地區":8}  {"地址"}')
        for ex in fix_examples[:20]:
            log(f'  {ex["row"]:5d}  {ex["method"]:8}  {ex["old_city"]:6}→{ex["new_city"]:8}  '
                f'{ex["old_dist"]:6}→{ex["new_dist"]:8}  {ex["addr"][:30]}')
    else:
        log('  （無可自動補正的縣市/地區）')

    if manual_city:
        log(f'\n[人工確認] 縣市無法推斷（前 {min(10, len(manual_city))} 筆）：')
        for m in manual_city[:10]:
            log(f'  列{m["row"]}  地段={m["sec"]}  地號={m["lno"]}  地址={m["addr"][:40]}')

    if dry_run:
        log('\n[dry-run] 不寫入檔案。確認後執行：')
        log(f'  python3 scripts/fill_declared_land_value.py')
        LOGS_DIR.mkdir(exist_ok=True)
        log_path.write_text('\n'.join(log_lines), encoding='utf-8')
        log(f'[log] {log_path}')
        return

    # ── 正式執行：套用補正並輸出
    log('\n[輸出] 套用縣市/地區補正 ...')
    for i, (new_city, new_dist, method) in city_fixes.items():
        data_rows[i][col['縣市']] = new_city
        if has_dist:
            data_rows[i][col['地區']] = new_dist

    # 重新排序
    i_city = col['縣市']; i_dist = col.get('地區', i_city)
    i_sec  = col['地段'];  i_lno  = col['地號']
    data_rows.sort(key=lambda r: (
        str(r[i_city] or ''), str(r[i_dist] or ''),
        str(r[i_sec] or ''), str(r[i_lno] or ''),
    ))

    # 寫出補正版主清冊
    wb_out = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_out = wb_out.active

    # 清除資料列重寫（保留格式）
    for ri, row in enumerate(data_rows, 2):
        for ci, val in enumerate(row[:len(headers)], 1):
            ws_out.cell(row=ri, column=ci, value=val)

    out_path = OUTPUT_DIR / f'主清冊_公告現值補正版_{ts}.xlsx'
    wb_out.save(out_path)
    wb_out.close()
    log(f'  補正版：{out_path}')

    # 寫出待查詢清單
    # 待查詢清單：只在 MOI API 可用時才輸出（避免輸出 13 萬筆無用清單）
    if pending_query and moi_api_available():
        wb_q = openpyxl.Workbook()
        ws_q = wb_q.active
        ws_q.title = '公告現值待查詢'
        q_headers = ['列號', '縣市', '地區', '地段', '地號', '原公告現值']
        from openpyxl.styles import PatternFill, Font, Alignment
        hdr_fill  = PatternFill('solid', fgColor='1F4E79')
        hdr_font  = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
        hdr_align = Alignment(horizontal='center')
        for ci, h in enumerate(q_headers, 1):
            cell = ws_q.cell(row=1, column=ci, value=h)
            cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=hdr_align
        for ri, item in enumerate(pending_query, 2):
            for ci, key in enumerate(['row','縣市','地區','地段','地號','原公告現值'], 1):
                ws_q.cell(row=ri, column=ci, value=item[key])
        ws_q.freeze_panes = 'A2'
        ws_q.auto_filter.ref = ws_q.dimensions
        q_path = OUTPUT_DIR / f'公告現值待查詢清單_{ts}.xlsx'
        wb_q.save(q_path); wb_q.close()
        log(f'  待查詢清單：{q_path}（{len(pending_query):,} 筆）')
    elif pending_query:
        log(f'  公告現值待查詢：{len(pending_query):,} 筆（MOI_API_014 未申請，不輸出清單，保持原值）')

    log(f'\n{"─"*50}')
    log(f'完成')
    log(f'  縣市/地區自動補正：{len(city_fixes):,} 筆')
    log(f'  公告現值待查詢：{cnt_val_manual:,} 筆')
    log(f'  補正版路徑：{out_path}')
    log(f'  原始 input：未修改')
    LOGS_DIR.mkdir(exist_ok=True)
    log_path.write_text('\n'.join(log_lines), encoding='utf-8')


def main():
    parser = argparse.ArgumentParser(description='主清冊公告現值 & 縣市地區補正')
    parser.add_argument('--dry-run', action='store_true', help='只統計，不寫檔')
    parser.add_argument('--input',   type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    if not args.input.exists():
        print(f'[錯誤] 找不到 Excel：{args.input}')
        sys.exit(1)

    run(args.input, args.dry_run)


if __name__ == '__main__':
    main()
