#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
realprice_alert.py — 實價登錄比對 × 獨立提醒報表 × Telegram 推播

流程：
  1. 比對 land_master.db × land_intel.db land_transactions
  2. 命中 → 輸出獨立報表 Excel（不修改主清冊）
  3. 命中 → Telegram 推播摘要
  4. 防重複：realprice_alert_log（--send-all 時逐筆防重）

執行：
  python3 scripts/realprice_alert.py --dry-run   ← 只報告，不寫入
  python3 scripts/realprice_alert.py             ← 正式：輸出報表 + 推播摘要
  python3 scripts/realprice_alert.py --send-all  ← 正式：逐筆推播全部（確認報表後再用）
  python3 scripts/realprice_alert.py --reconcile ← 回溯比對：將已反映的地號從報表移除

主清冊 Excel 不再被修改。

══ 核心業務規則（正式鎖定）══

實價提醒「是否已反映」的唯一判定標準：
  同地號在 MASTER / SQLite 已存在更新後的新地主（reg_reason='買賣', reg_date >= 實價日期）
  → 代表此實價交易已被調閱並反映於清冊，無需繼續提醒。

不作為判定依據的欄位：
  × is_sold / 已售出欄位
  × 舊地主是否被標已售
  理由：舊清冊歷史資料很多缺乏完整的已售標記，
        「新地主結構是否已存在」與「舊地主是否標已售」是兩件獨立的事。

實價提醒報表定位：
  = 買賣異動待調閱清單（僅追蹤買賣交易訊號）
  ≠ 增貸 / 抵押 / 他項設定 / 繼承（這些屬於其他模組）
"""

import argparse
import json
import os
import re
import sqlite3
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

PROJECT_ROOT  = Path(__file__).resolve().parent.parent
LM_DB         = PROJECT_ROOT / 'data' / 'database' / 'land_master.db'
LI_DB         = Path('/Users/xiaomingyang/Desktop/:Users:mingyang:land-ai:/land-intel-system/db/land_intel.db')
BASE_DIR      = Path('/Users/xiaomingyang/Desktop/excel土地資料維護')
REPORT_DIR    = BASE_DIR / 'output'
LATEST_DIR    = BASE_DIR / '最新完成版'
LATEST_REPORT = LATEST_DIR / '實價提醒報表_最新完成版.xlsx'

# 報表欄位（獨立報表，不寫主清冊）
REPORT_COLS = [
    '縣市', '地區', '地段', '地號', '所有人',
    '次序', '登記日期', '登記原因', '權利範圍', '已售出', '備註',
    '實價日期', '實價總價(萬)', '同批命中地號', '建議動作',
]

# ── .env 載入 ─────────────────────────────────────────────────────

def load_dotenv():
    for p in [PROJECT_ROOT / '.env', Path.home() / '.env']:
        if p.exists():
            for line in p.read_text(encoding='utf-8').splitlines():
                line = line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())
            break

load_dotenv()

# ── 共用工具 ──────────────────────────────────────────────────────

def norm_land_no(raw) -> str:
    if not raw:
        return ''
    s = re.sub(r'之', '-', str(raw).strip())
    s = re.sub(r'[^\d\-]', '', s)
    if not s:
        return ''
    parts = s.split('-')
    try:
        main = int(parts[0]) if parts[0] else 0
        sub  = int(parts[1]) if len(parts) > 1 and parts[1] else 0
        return f'{main:04d}-{sub:04d}'
    except ValueError:
        return s


def roc_date(td: str) -> str:
    if not td:
        return '—'
    td = str(td).strip()
    if re.match(r'^\d{3}/\d{2}/\d{2}$', td):
        return td
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', td)
    if m:
        roc = int(m.group(1)) - 1911
        return f'{roc}/{m.group(2)}/{m.group(3)}'
    return td


def fmt_price(wan):
    if wan is None:
        return '—'
    return f'{wan:,.0f}'


def fmt_unit(per_sqm):
    if per_sqm is None:
        return '—'
    return f'{per_sqm:,.0f}'

# ── DB 初始化 ─────────────────────────────────────────────────────

def init_alert_log(con: sqlite3.Connection):
    con.execute("""
        CREATE TABLE IF NOT EXISTS realprice_alert_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            dedup_key    TEXT UNIQUE,          -- city|district|section|land_no|trade_date|total_price
            alert_type   TEXT,
            hit_count    INTEGER,
            notified_at  TEXT DEFAULT (datetime('now')),
            telegram_ok  INTEGER DEFAULT 0
        )
    """)
    con.commit()

# ── 資料載入 ─────────────────────────────────────────────────────

def load_land_master() -> tuple[list[dict], dict]:
    """回傳 (all_rows, index)，index key = (city, district, norm_section, norm_land_no)"""
    con = sqlite3.connect(LM_DB, timeout=10)
    con.row_factory = sqlite3.Row
    rows = [dict(r) for r in con.execute("""
        SELECT city, district, normalized_section, section_raw,
               normalized_land_no, land_no_raw, owner_name,
               address, phone, note, is_sold,
               reg_seq, reg_date, reg_reason, total_area_ping, ownership_range,
               land_match_key, event_key
        FROM land_master
    """).fetchall()]
    con.close()
    idx = defaultdict(list)
    for r in rows:
        idx[(r['city'], r['district'], r['normalized_section'], r['normalized_land_no'])].append(r)
    return rows, idx


def load_transactions() -> list[dict]:
    con = sqlite3.connect(LI_DB, timeout=10)
    con.row_factory = sqlite3.Row
    rows = [dict(r) for r in con.execute("""
        SELECT city, district, section_name, land_number,
               trade_date, total_price_wan, unit_price_per_sqm,
               unique_key, transaction_target, target_category
        FROM land_transactions
        WHERE trade_date >= '2025-01-01'
          AND (target_category LIKE '%土地%' OR target_category IS NULL)
          AND land_number IS NOT NULL AND land_number != ''
        ORDER BY trade_date DESC
    """).fetchall()]
    con.close()
    return rows


def load_alert_log_keys(con: sqlite3.Connection) -> set:
    return {r[0] for r in con.execute('SELECT dedup_key FROM realprice_alert_log').fetchall()}

# ── 比對核心 ─────────────────────────────────────────────────────

def parse_to_date(s) -> date | None:
    """
    把各種日期格式解析成 date 物件。
    支援：114/02/07、114年02月07日、2025-02-07、20250207
    """
    if not s:
        return None
    s = str(s).strip()
    # ROC 斜線格式：114/02/07
    m = re.match(r'^(\d{2,3})/(\d{1,2})/(\d{1,2})$', s)
    if m:
        try:
            return date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # ROC 中文格式：114年02月07日
    m = re.match(r'^(\d{2,3})年(\d{1,2})月(\d{1,2})日$', s)
    if m:
        try:
            return date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # 西元格式：2025-02-07 or 2025/02/07
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # 西元無分隔：20250207
    m = re.match(r'^(\d{4})(\d{2})(\d{2})$', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def is_reflected_in_master(hit_masters: list[dict], trade_date_str: str,
                            window_days: int = 90) -> tuple[bool, str]:
    """
    判斷本次實價交易是否已由主清冊反映（= 不需再提醒調閱）。

    ══ 核心判定規則 ══
    「已反映」= 同地號在 MASTER/DB 已存在更新後的新地主結構：
      - 有新的買賣登記（reg_reason = '買賣'）
      - 且新登記日期 >= 實價成交日期（允許最長 window_days 天的登記延遲）

    ══ 明確不使用的欄位 ══
    is_sold / 已售出 欄位 → 不作為判定依據。
    理由：舊 MASTER 清冊歷史資料往往缺乏完整的已售標記，
          若以 is_sold 判斷會誤判大量「已反映但舊地主未標售」的案例。
    「舊地主是否標已售」不影響「新地主結構是否已存在」這個事實。

    回傳 (is_reflected: bool, reason_str: str)
    """
    trade_dt = parse_to_date(trade_date_str)
    if not trade_dt:
        return False, '實價日期無法解析'

    for m in hit_masters:
        if (m.get('reg_reason') or '').strip() != '買賣':
            continue
        reg_dt = parse_to_date(m.get('reg_date'))
        if not reg_dt:
            continue
        delta = (reg_dt - trade_dt).days
        # 新地主登記日期 >= 實價成交日，且差距在 window_days 天內
        # → 代表此實價交易事件已在 MASTER 中有對應的新地主登記
        if 0 <= delta <= window_days:
            return True, (f'{m["owner_name"]} 登記日={m["reg_date"]} '
                          f'實價日={trade_date_str} 差={delta}天')
    return False, ''


def build_alerts(lm_idx: dict, lt_rows: list[dict]) -> list[dict]:
    """
    回傳 alert 列表，每個 alert = 一筆實價交易 × 命中的 land_master 列。
    同交易組（city+district+section+trade_date+total_price）保留分筆（不合並總價）。
    同一組內命中的所有 land_master 列合併顯示。
    """
    def tx_key(t):
        return (t['city'], t['district'], t['section_name'] or '',
                t['trade_date'], str(t['total_price_wan']))

    # 按交易 key 聚合實價地號
    tx_groups: dict[tuple, list] = defaultdict(list)
    for t in lt_rows:
        tx_groups[tx_key(t)].append(t)

    alerts = []
    seen_dedup = set()

    for gk, tx_list in tx_groups.items():
        city, district, sec, trade_date, _ = gk
        sample_t = tx_list[0]
        t_sec = sec.strip()

        # 找出本交易組命中的所有 land_master 列（去重 event_key）
        hit_masters: list[dict] = []
        seen_ek = set()
        for t in tx_list:
            nn = norm_land_no(t['land_number'])
            if not nn:
                continue
            for k, masters in lm_idx.items():
                if k[0] != city or k[1] != district or k[3] != nn:
                    continue
                if not (t_sec in k[2] or k[2] in t_sec or t_sec == k[2]):
                    continue
                for m in masters:
                    if m['event_key'] not in seen_ek:
                        seen_ek.add(m['event_key'])
                        hit_masters.append(m)

        if not hit_masters:
            continue

        dedup_key = '|'.join([city, district, sec, trade_date,
                               str(sample_t['total_price_wan']),
                               ','.join(sorted(seen_ek))])

        if dedup_key in seen_dedup:
            continue
        seen_dedup.add(dedup_key)

        # ── 排除規則：主清冊已反映此次買賣 ──
        reflected, reflect_reason = is_reflected_in_master(hit_masters, trade_date)
        if reflected:
            alerts.append({
                'dedup_key':       dedup_key,
                'sample_t':        sample_t,
                'tx_list':         tx_list,
                'hit_masters':     hit_masters,
                'alert_type':      '已由主清冊反映',
                'tx_parcel_count': len(tx_list),
                'hit_master_count': len(hit_masters),
                'unique_land_nos': [],
                'reflected':       True,
                'reflect_reason':  reflect_reason,
            })
            continue

        # alert_type 是顯示分類標籤，不影響「是否提醒」的決策。
        # 注意：這裡的 is_sold 僅用於將提醒分類為「疑似異動」或「已售地主再命中」，
        #       不作為「此實價是否已被 MASTER 反映」的依據（那由 is_reflected_in_master 決定）。
        # 即使所有地主的 is_sold 都是 0（舊清冊未完整標售），提醒分類仍能正常運作。
        any_unsold = any(m['is_sold'] == 0 for m in hit_masters)
        alert_type = '地號疑似異動' if any_unsold else '已售地主再命中'

        # 同批命中地號（去重顯示）
        seen_lno = set()
        unique_land_nos = []
        for m in hit_masters:
            if m['land_no_raw'] not in seen_lno:
                seen_lno.add(m['land_no_raw'])
                unique_land_nos.append(m['land_no_raw'])

        alerts.append({
            'dedup_key':       dedup_key,
            'sample_t':        sample_t,
            'tx_list':         tx_list,
            'hit_masters':     hit_masters,
            'alert_type':      alert_type,
            'tx_parcel_count': len(tx_list),
            'hit_master_count': len(hit_masters),
            'unique_land_nos': unique_land_nos,
        })

    alerts.sort(key=lambda a: a['sample_t']['trade_date'], reverse=True)
    return alerts

# ── 實價提醒報表回溯比對（reconcile）───────────────────────────────────────

def reconcile_realprice_alerts(dry_run: bool = False) -> dict:
    """
    對照現有 land_master.db，回溯比對實價提醒報表_最新完成版.xlsx。

    ══ 正式規則（v5.3）══
    實價提醒報表 = 尚未調閱的待辦清單。
    已反映的地號直接從報表移除，不留「已調閱」狀態列。
    歷史資料保存於 SQLite + MASTER 清冊，報表只留真正待辦事項。

    ══ 移除條件 ══
    1. 已標記「已調閱」的舊有列（清除前次標記遺留）
    2. 同地號在 DB 已有 reg_reason='買賣' 且 reg_date >= 實價日期（差距 ≤ 90 天）

    ══ 明確不使用的欄位 ══
    is_sold / 已售出 → 不作為「已反映」的判定依據。

    典型觸發場景：
      import_land_master.py 批量匯入新地主後（未走 process_land_transcripts.py），
      實價提醒報表未自動更新 → 手動或定期執行此函數修正。

    回傳：
      {
        'checked':   int,   # 掃描的未處理提醒筆數
        'removed':   int,   # 本次移除的列數（含舊有已調閱列）
        'remaining': int,   # 移除後仍在清單的列數
        'dry_run':   bool,
      }
    """
    import openpyxl
    import sqlite3

    if not LATEST_REPORT.exists():
        print(f'[reconcile] 找不到報表：{LATEST_REPORT}')
        return {'checked': 0, 'removed': 0, 'remaining': 0, 'dry_run': dry_run}

    # ── 讀取報表，找欄位 index ──
    wb = openpyxl.load_workbook(str(LATEST_REPORT))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        return headers.index(name) + 1 if name in headers else None

    sec_col    = col('地段')
    no_col     = col('地號')
    date_col   = col('實價日期')
    action_col = col('建議動作')
    owner_col  = col('所有權人') or col('所有人')

    if not (sec_col and no_col and date_col and action_col):
        print(f'[reconcile] 報表欄位不完整，中止')
        wb.close()
        return {'checked': 0, 'removed': 0, 'remaining': 0, 'dry_run': dry_run}

    # ── 載入 land_master DB（只需要買賣記錄）──
    con = sqlite3.connect(str(LM_DB), timeout=10)
    con.row_factory = sqlite3.Row
    db_rows = [dict(r) for r in con.execute("""
        SELECT normalized_section, normalized_land_no,
               reg_date, reg_reason, owner_name
        FROM land_master
        WHERE reg_reason = '買賣'
    """).fetchall()]
    con.close()

    # 建立 (normalized_section, normalized_land_no) → [master row, ...] 快速索引
    from collections import defaultdict
    db_idx: dict[tuple, list[dict]] = defaultdict(list)
    for r in db_rows:
        db_idx[(r['normalized_section'], r['normalized_land_no'])].append(r)

    # ── 逐列比對，收集待刪列號（由大到小刪，避免位移）──
    rows_to_delete: list[int] = []
    checked = 0
    skipped = 0
    total_data_rows = ws.max_row - 1  # 不含 header

    for row_num in range(2, ws.max_row + 1):
        action = str(ws.cell(row_num, action_col).value or '')

        # 已標「已調閱」的舊有列也一併清除
        if '已調閱' in action:
            rows_to_delete.append(row_num)
            continue

        raw_sec    = str(ws.cell(row_num, sec_col).value or '').strip()
        raw_no     = str(ws.cell(row_num, no_col).value or '').strip()
        trade_date = str(ws.cell(row_num, date_col).value or '').strip()

        if not (raw_sec and raw_no and trade_date):
            continue

        n_sec = re.sub(r'\([^)]*\)', '', raw_sec).strip()
        n_no  = norm_land_no(raw_no)

        hit_masters = []
        for (db_sec, db_no), masters in db_idx.items():
            if db_no != n_no:
                continue
            if n_sec in db_sec or db_sec in n_sec or n_sec == db_sec:
                hit_masters.extend(masters)

        checked += 1

        if not hit_masters:
            skipped += 1
            continue

        reflected, reason = is_reflected_in_master(hit_masters, trade_date)
        if reflected:
            owner = str(ws.cell(row_num, owner_col).value or '') if owner_col else ''
            rows_to_delete.append(row_num)
            print(f'  [{"DRY" if dry_run else "DEL"}] row {row_num:4d}  {raw_sec} {raw_no}'
                  f'  {owner}  實價={trade_date}  → {reason[:50]}')
        else:
            skipped += 1

    removed = len(rows_to_delete)
    remaining = total_data_rows - removed

    if removed and not dry_run:
        for row_num in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_num)
        wb.save(str(LATEST_REPORT))
        print(f'[reconcile] 已移除 {removed} 列，剩餘 {remaining} 筆待調閱')
        print(f'[reconcile] 已儲存：{LATEST_REPORT}')
    elif dry_run:
        print(f'[reconcile] DRY-RUN：將移除 {removed} 列，剩餘 {remaining} 筆待調閱')
    wb.close()

    return {'checked': checked, 'removed': removed, 'remaining': remaining, 'dry_run': dry_run}


def reconcile_sold_status(dry_run: bool = False) -> dict:
    """
    回溯修正舊地主的 is_sold 狀態。

    ══ 適用場景 ══
    bulk import（import_land_master.py）匯入歷史資料後，未跑 process_land_transcripts.py
    的 diff 邏輯，導致舊地主 is_sold 仍為 0，Excel 未反灰。

    ══ 判定規則（保守，只標確定已售出者）══
    對每個地號：
      1. 找出最早的 買賣 事件日期（earliest_buysel_date）
      2. 凡 reg_reason ≠ '買賣'（即 贈與/繼承/總登記/分割繼承 等前手）
         且 reg_date < earliest_buysel_date（或 reg_date 為空）
         且 is_sold = 0
         → 標記為已售（is_sold=1, 已售出=已售）

    ══ 明確不動的記錄 ══
    - reg_reason = '買賣' 的記錄（買賣本身即為現任或歷史購入方，不自動標售）
    - reg_date >= earliest_buysel_date 的記錄（可能是買賣後的新繼承/贈與）

    ══ 影響範圍 ══
    - land_master.db: is_sold, sys_note 更新
    - 老蕭LAND_MASTER.xlsx: 已售出欄 = '已售'
    - 執行後須呼叫 reformat_and_sort_master() 更新顏色（此函數會自動呼叫）

    回傳：
      { 'updated_db': int, 'updated_excel': int, 'parcels': int, 'dry_run': bool }
    """
    import openpyxl
    from datetime import date as _date

    MASTER_XLSX = LATEST_DIR / '老蕭LAND_MASTER.xlsx'
    if not MASTER_XLSX.exists():
        print(f'[reconcile_sold] 找不到 MASTER: {MASTER_XLSX}')
        return {'updated_db': 0, 'updated_excel': 0, 'parcels': 0, 'dry_run': dry_run}

    today_str = _date.today().strftime('%Y-%m-%d')

    # ── Step 1: 從 DB 找出需要標已售的 rowid ──
    con = sqlite3.connect(str(LM_DB), timeout=10)
    con.row_factory = sqlite3.Row

    # 取出所有地號的所有記錄（rowid 明確 alias 確保 sqlite3.Row 可 key 存取）
    all_rows = [dict(r) for r in con.execute("""
        SELECT rowid AS row_id, normalized_section, normalized_land_no,
               owner_name, reg_date, reg_reason, is_sold, sys_note
        FROM land_master
        ORDER BY normalized_section, normalized_land_no, reg_date, CAST(reg_seq AS INTEGER)
    """).fetchall()]

    def _roc_to_date(s) -> '_date | None':
        if not s: return None
        import re as _re
        m = _re.match(r'(\d+)年(\d+)月(\d+)日', str(s))
        if m:
            try: return _date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
            except (ValueError, OverflowError): pass
        m = _re.match(r'(\d+)/(\d+)/(\d+)', str(s))
        if m:
            try: return _date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
            except (ValueError, OverflowError): pass
        return None

    # 建立 parcel → rows 索引
    from collections import defaultdict as _dd
    parcel_rows: dict[tuple, list] = _dd(list)
    for r in all_rows:
        parcel_rows[(r['normalized_section'], r['normalized_land_no'])].append(r)

    to_sell_rowids: list[int] = []   # DB rowids 待標售
    detail_log: list[str] = []

    for (sec, lno), rows in parcel_rows.items():
        # 找出此地號最早的買賣日期
        buysel_dates = [_roc_to_date(r['reg_date']) for r in rows
                        if (r['reg_reason'] or '').strip() == '買賣']
        buysel_dates = [d for d in buysel_dates if d]
        if not buysel_dates:
            continue  # 此地號沒有任何買賣記錄，跳過

        earliest_buysel = min(buysel_dates)

        for r in rows:
            if r['is_sold']:
                continue  # 已標售，跳過
            reason = (r['reg_reason'] or '').strip()
            if reason == '買賣':
                continue  # 買賣記錄本身不自動標售
            reg_dt = _roc_to_date(r['reg_date'])
            # NULL 日期 or 日期早於最早買賣 → 前手賣方
            if reg_dt is None or reg_dt < earliest_buysel:
                to_sell_rowids.append(r['row_id'])
                detail_log.append(
                    f'  {"DRY" if dry_run else "OK"}  rowid={r["row_id"]:6d}  '
                    f'{sec} {lno}  {r["owner_name"]}  {r["reg_reason"]}  '
                    f'reg={r["reg_date"] or "NULL"}  earliest_buysel={earliest_buysel}'
                )

    parcels_affected = len({
        (r['normalized_section'], r['normalized_land_no'])
        for r in all_rows if r['row_id'] in set(to_sell_rowids)
    })

    print(f'\n[reconcile_sold] 找到需標已售：{len(to_sell_rowids)} 筆 / {parcels_affected} 個地號')
    for line in detail_log[:50]:
        print(line)
    if len(detail_log) > 50:
        print(f'  ... 另 {len(detail_log) - 50} 筆省略 ...')

    updated_db = 0
    if not dry_run and to_sell_rowids:
        for rid in to_sell_rowids:
            con.execute("""
                UPDATE land_master
                SET is_sold = 1,
                    sys_note = COALESCE(sys_note || ' | ', '') || ?
                WHERE rowid = ?
            """, (f'reconcile_sold {today_str}', rid))
        con.commit()
        updated_db = len(to_sell_rowids)
        print(f'[reconcile_sold] SQLite 已更新 {updated_db} 筆 is_sold=1')

    con.close()

    # ── Step 2: 更新 Excel 已售出欄 ──
    sell_owners: dict[tuple[str, str], set[str]] = _dd(set)
    for r in all_rows:
        if r['row_id'] in set(to_sell_rowids):
            sell_owners[(r['normalized_section'], r['normalized_land_no'])].add(
                str(r['owner_name'] or ''))

    wb = openpyxl.load_workbook(str(MASTER_XLSX), read_only=False, data_only=True)
    ws = wb.active
    headers_xl = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def _col(name):
        try: return headers_xl.index(name) + 1
        except ValueError: return None

    sec_c    = _col('地段')
    no_c     = _col('地號')
    owner_c  = _col('所有權人')
    sold_c   = _col('已售出')

    if not (sec_c and no_c and owner_c and sold_c):
        print(f'[reconcile_sold] Excel 欄位缺失，跳過 Excel 更新')
        wb.close()
        return {'updated_db': updated_db, 'updated_excel': 0,
                'parcels': parcels_affected, 'dry_run': dry_run}

    updated_excel = 0

    for row_num in range(2, ws.max_row + 1):
        raw_sec   = str(ws.cell(row_num, sec_c).value or '').strip()
        raw_no    = str(ws.cell(row_num, no_c).value or '').strip()
        raw_owner = str(ws.cell(row_num, owner_c).value or '').strip()
        already_sold = str(ws.cell(row_num, sold_c).value or '').strip()

        if '已售' in already_sold:
            continue

        # 標準化 section（移除括號部分）
        n_sec = re.sub(r'\([^)]*\)', '', raw_sec).strip()
        n_no  = norm_land_no(raw_no)

        # 找匹配的 sell_owners
        matched = False
        for (db_sec, db_no), owners in sell_owners.items():
            if db_no != n_no:
                continue
            if n_sec in db_sec or db_sec in n_sec or n_sec == db_sec:
                if raw_owner in owners:
                    matched = True
                    break

        if matched:
            if not dry_run:
                ws.cell(row_num, sold_c).value = '已售'
            updated_excel += 1
            print(f'  {"DRY" if dry_run else "XL"} row {row_num:5d}  '
                  f'{raw_sec} {raw_no}  {raw_owner} → 已售')

    if updated_excel and not dry_run:
        wb.save(str(MASTER_XLSX))
        print(f'[reconcile_sold] Excel 已儲存：{MASTER_XLSX}')
        print(f'[reconcile_sold] Excel 標已售：{updated_excel} 列')
    wb.close()

    return {
        'updated_db':    updated_db,
        'updated_excel': updated_excel,
        'parcels':       parcels_affected,
        'dry_run':       dry_run,
    }


# ── Excel 更新 ────────────────────────────────────────────────────

def generate_report(alerts: list[dict], dry_run: bool) -> tuple[int, Path | None]:
    """
    輸出獨立實價提醒報表。不修改主清冊 Excel。
    回傳 (列數, 報表路徑)。
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    today_str = date.today().isoformat()
    ts        = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path  = REPORT_DIR / f'實價提醒報表_{ts}.xlsx'

    # 展開 alert → 各 master 列（每位地主一列）
    rows_out = []
    for a in alerts:
        t   = a['sample_t']
        lns = '、'.join(a['unique_land_nos'])
        at  = a['alert_type']
        for m in a['hit_masters']:
            rows_out.append({
                '縣市':         m['city']              or '',
                '地區':         m['district']          or '',
                '地段':         m['section_raw']       or '',
                '地號':         m['land_no_raw']       or '',
                '所有人':       m['owner_name']        or '',
                '次序':         m.get('reg_seq')       or '',
                '登記日期':     roc_date(m.get('reg_date', '')),
                '登記原因':     m.get('reg_reason')    or '',
                '權利範圍':     m.get('ownership_range') or '',
                '已售出':       '是' if m['is_sold'] == 1 else '否',
                '備註':         m.get('note')          or '',
                '實價日期':     roc_date(t['trade_date']),
                '實價總價(萬)': t['total_price_wan'],
                '同批命中地號': lns,
                '建議動作':     '請確認此地號是否已有地主異動',
            })

    total = len(rows_out)
    print(f'\n[報表] 預計輸出 {total} 列')

    if dry_run:
        print(f'[報表] dry-run：不寫入檔案')
        print(f'[報表] 預計路徑：{out_path}')
        return total, None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '實價提醒'

    # Header
    header_fill = PatternFill('solid', fgColor='FFC000')
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(REPORT_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    alert_fill   = PatternFill('solid', fgColor='FFEB9C')  # 黃：高度疑似
    sold_fill    = PatternFill('solid', fgColor='FFCCCC')  # 紅：已售地主再命中

    for row_idx, r in enumerate(rows_out, 2):
        fill = sold_fill if r['已售出'] == '是' else alert_fill
        for col_idx, col_name in enumerate(REPORT_COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(col_name))
            cell.fill = fill

    # 凍結標題列
    ws.freeze_panes = 'A2'

    # 欄寬
    col_widths = {
        '縣市': 8, '地區': 8, '地段': 12, '地號': 10, '所有人': 10,
        '次序': 6, '登記日期': 10, '登記原因': 10, '權利範圍': 12, '已售出': 6,
        '備註': 20, '實價日期': 10, '實價總價(萬)': 10,
        '同批命中地號': 24, '建議動作': 20,
    }
    for col_idx, col_name in enumerate(REPORT_COLS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
            col_widths.get(col_name, 12)

    wb.save(out_path)
    wb.close()
    print(f'[報表] 已儲存：{out_path}')

    # ── 同步到 最新完成版/（只覆蓋報表那一份）
    import shutil as _shutil
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    if LATEST_REPORT.exists():
        LATEST_REPORT.unlink()
    _shutil.copy2(out_path, LATEST_REPORT)
    print(f'[最新版] 已同步：{LATEST_REPORT}')

    # ── output/ 自動清理：實價報表只保留最近 2 份
    report_files = sorted(REPORT_DIR.glob('實價提醒報表_*.xlsx'),
                          key=lambda p: p.stat().st_mtime, reverse=True)
    for old in report_files[2:]:
        old.unlink()
        print(f'[清理] 已刪除舊版：{old.name}')
    print(f'[清理] 實價報表 output 保留 {min(len(report_files), 2)} 份')

    return total, out_path

# ── Telegram 推播 ─────────────────────────────────────────────────

def tg_send(chat_id: str, text: str, token: str):
    url  = f'https://api.telegram.org/bot{token}/sendMessage'
    data = json.dumps({'chat_id': chat_id, 'text': text}).encode()
    req  = urllib.request.Request(url, data=data,
                                  headers={'Content-Type': 'application/json'})
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


def build_tg_message(a: dict) -> str:
    t     = a['sample_t']
    ms    = a['hit_masters']
    at    = a['alert_type']
    lns   = '、'.join(a['unique_land_nos'])
    price = fmt_price(t['total_price_wan'])
    unit  = fmt_unit(t['unit_price_per_sqm'])

    # 地主清單（去重姓名，顯示已售狀態）
    seen_names = set()
    owner_lines = []
    for m in ms:
        nm = m['owner_name'] or '—'
        if nm not in seen_names:
            seen_names.add(nm)
            sold_tag = ' [已售]' if m['is_sold'] == 1 else ''
            owner_lines.append(f"  {nm}{sold_tag}")
    owners_str = '\n'.join(owner_lines[:10])
    if len(seen_names) > 10:
        owners_str += f'\n  …共 {len(seen_names)} 人'

    sec = (t.get('section_name') or '').strip()

    return (
        f"⚠️ {at}\n\n"
        f"📍 {t['city']} {t['district']} {sec}\n"
        f"📌 命中地號：{lns}\n"
        f"👤 主清冊地主：\n{owners_str}\n"
        f"📅 實價日期：{roc_date(t['trade_date'])}\n"
        f"💰 總價：{price} 萬\n"
        f"🎯 命中主清冊：{a['hit_master_count']} 筆\n\n"
        f"請確認此地號是否已有地主異動。"
    )


def build_summary_message(alerts: list[dict], report_rows: int) -> str:
    unsold = sum(1 for a in alerts if a['alert_type'] == '地號疑似異動')
    sold   = sum(1 for a in alerts if a['alert_type'] == '已售地主再命中')
    multi  = sum(1 for a in alerts if a['hit_master_count'] > 1)
    today  = date.today().strftime('%Y-%m-%d')
    return (
        f"⚠️ 實價提醒報表已產生\n\n"
        f"📅 比對日期：{today}\n"
        f"📊 本次命中交易組：{len(alerts)} 組\n"
        f"  地號疑似異動：{unsold} 組\n"
        f"  已售地主再命中：{sold} 組\n"
        f"  同批多地號命中：{multi} 組\n\n"
        f"📝 報表列數：{report_rows} 列\n"
        f"📁 檔案：實價提醒報表_*.xlsx\n\n"
        f"請打開報表查看，確認是否有地主異動需調新謄本。\n"
        f"主清冊未修改。"
    )


def send_summary(alerts: list[dict], report_rows: int, dry_run: bool) -> int:
    """只推播一則摘要，不寫 realprice_alert_log（保留逐筆推播的空間）。"""
    token   = os.environ.get('TELEGRAM_BOT_TOKEN', '')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID', '')
    if not token or not chat_id:
        print('[Telegram] 缺少 TELEGRAM_BOT_TOKEN 或 TELEGRAM_CHAT_ID，跳過推播')
        return 0

    msg = build_summary_message(alerts, report_rows)
    print(f'\n[Telegram] 摘要推播：\n{msg}')
    if dry_run:
        print('（dry-run，未發送）')
        return 0
    try:
        tg_send(chat_id, msg, token)
        print('[Telegram] 摘要已推播')
        return 1
    except Exception as e:
        print(f'[Telegram] 推播失敗：{e}')
        return 0


def send_all_alerts(alerts: list[dict], dry_run: bool) -> int:
    """逐筆推播，並寫入 realprice_alert_log 防重複。需明確 --send-all 才執行。"""
    token   = os.environ.get('TELEGRAM_BOT_TOKEN', '')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID', '')
    if not token or not chat_id:
        print('[Telegram] 缺少 TELEGRAM_BOT_TOKEN 或 TELEGRAM_CHAT_ID，跳過推播')
        return 0

    con = sqlite3.connect(LM_DB, timeout=10)
    init_alert_log(con)
    existing_keys = load_alert_log_keys(con)

    new_alerts = [a for a in alerts if a['dedup_key'] not in existing_keys]
    print(f'\n[Telegram] 逐筆推播：{len(new_alerts)} 則（已推播過 {len(alerts)-len(new_alerts)} 則）')

    sent = 0
    for a in new_alerts:
        msg = build_tg_message(a)
        if dry_run:
            print(f'\n--- dry-run ---\n{msg}\n')
            continue
        try:
            tg_send(chat_id, msg, token)
            ok = 1
        except Exception as e:
            print(f'[Telegram] 推播失敗：{e}')
            ok = 0
        con.execute("""
            INSERT OR IGNORE INTO realprice_alert_log
                (dedup_key, alert_type, hit_count, telegram_ok)
            VALUES (?,?,?,?)
        """, (a['dedup_key'], a['alert_type'], a['hit_master_count'], ok))
        con.commit()
        sent += ok

    con.close()
    print(f'[Telegram] 逐筆推播完成，成功 {sent} 則')
    return sent

# ── 主程式 ────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='實價登錄比對提醒')
    parser.add_argument('--dry-run',   action='store_true', help='只報告，不寫入')
    parser.add_argument('--send-all',  action='store_true', help='逐筆推播（確認 Excel 後再用）')
    parser.add_argument('--reconcile', action='store_true',
                        help='回溯比對：對照現有 DB 將已反映的地號直接從實價提醒報表移除')
    parser.add_argument('--reconcile-sold', action='store_true',
                        help='回溯標已售：將 bulk import 舊地主補標 is_sold=1 + Excel 反灰')
    args = parser.parse_args()
    dry_run        = args.dry_run
    send_all       = args.send_all
    reconcile      = args.reconcile
    reconcile_sold = args.reconcile_sold

    if send_all and dry_run:
        print('--dry-run 與 --send-all 不可同時使用')
        sys.exit(1)

    # ── reconcile 模式：獨立執行，不跑完整比對流程 ──
    if reconcile:
        mode = 'DRY-RUN' if dry_run else '正式執行'
        print(f'[ {mode} ] realprice_alert.py --reconcile')
        print(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
        print(f'報表：{LATEST_REPORT}')
        print(f'DB  ：{LM_DB}')
        print()
        result = reconcile_realprice_alerts(dry_run=dry_run)
        print()
        print(f'── 結果 ──')
        print(f'  掃描待調閱提醒  ：{result["checked"]}')
        print(f'  移除（已反映）  ：{result["removed"]}')
        print(f'  剩餘待調閱      ：{result["remaining"]}')
        if dry_run:
            print(f'\n確認後執行（正式移除）：')
            print(f'  python3 scripts/realprice_alert.py --reconcile')
        return

    # ── reconcile-sold 模式：補標已售 + 重新套用 Excel 格式 ──
    if reconcile_sold:
        mode = 'DRY-RUN' if dry_run else '正式執行'
        print(f'[ {mode} ] realprice_alert.py --reconcile-sold')
        print(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
        print(f'DB  ：{LM_DB}')
        print()
        result = reconcile_sold_status(dry_run=dry_run)
        print()
        print(f'── 結果 ──')
        print(f'  影響地號數      ：{result["parcels"]}')
        print(f'  SQLite 標已售   ：{result["updated_db"]}')
        print(f'  Excel 標已售    ：{result["updated_excel"]}')
        if not dry_run and result['updated_excel']:
            print(f'\n[reformat] 重新套用格式與排序…')
            try:
                import importlib.util, sys as _sys
                spec = importlib.util.spec_from_file_location(
                    'plt', str(PROJECT_ROOT / 'scripts' / 'process_land_transcripts.py'))
                plt = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(plt)
                fmt = plt.reformat_and_sort_master(dry_run=False)
                print(f'  已售灰色列      ：{fmt["sold_grey"]}')
                print(f'  近半年黃色列    ：{fmt["recent_yellow"]}')
                print(f'  排序完成        ：{fmt["sorted"]}')
                print(f'  耗時            ：{fmt["elapsed"]:.1f}s')
            except Exception as e:
                print(f'[reformat] 失敗：{e}')
                print('請手動執行：python3 scripts/process_land_transcripts.py --reformat')
        elif dry_run:
            print(f'\n確認後執行（正式寫入 + 重新格式化）：')
            print(f'  python3 scripts/realprice_alert.py --reconcile-sold')
        return

    mode = 'DRY-RUN' if dry_run else ('逐筆推播' if send_all else '正式執行（摘要推播）')
    print(f'[ {mode} ] realprice_alert.py')
    print(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    print(f'LM_DB  : {LM_DB}')
    print(f'LI_DB  : {LI_DB}')
    print(f'報表目錄: {REPORT_DIR}')

    print('\n[載入] land_master ...')
    lm_rows, lm_idx = load_land_master()
    print(f'  {len(lm_rows):,} 筆（未售：{sum(1 for r in lm_rows if r["is_sold"]==0):,}  '
          f'已售：{sum(1 for r in lm_rows if r["is_sold"]==1):,}）')

    print('[載入] land_transactions ...')
    lt_rows = load_transactions()
    print(f'  {len(lt_rows):,} 筆（2025起）')

    print('[比對] ...')
    all_alerts = build_alerts(lm_idx, lt_rows)

    # 分離：已反映 vs 需提醒
    reflected_alerts = [a for a in all_alerts if a.get('reflected')]
    alerts           = [a for a in all_alerts if not a.get('reflected')]

    unsold = sum(1 for a in alerts if a['alert_type'] == '地號疑似異動')
    sold   = sum(1 for a in alerts if a['alert_type'] == '已售地主再命中')
    multi  = sum(1 for a in alerts if a['hit_master_count'] > 1)

    print(f'  原始命中組數（未排除）：{len(all_alerts):,}')
    print(f'  排除（主清冊已反映）  ：{len(reflected_alerts):,}')
    print(f'  淨提醒組數            ：{len(alerts):,}')
    print(f'    地號疑似異動        ：{unsold}')
    print(f'    已售地主再命中      ：{sold}')
    print(f'    同批多地號命中      ：{multi}')

    if dry_run and reflected_alerts:
        print(f'\n[排除案例] 前 {min(10, len(reflected_alerts))} 筆「已由主清冊反映」：')
        for i, a in enumerate(reflected_alerts[:10], 1):
            t = a['sample_t']
            print(f'  [{i}] {t["city"]} {t["district"]} '
                  f'{(t.get("section_name") or "").strip()} '
                  f'  實價日={roc_date(t["trade_date"])}'
                  f'  → {a["reflect_reason"]}')

    # ── 輸出獨立報表（不修改主清冊）
    report_rows, report_path = generate_report(alerts, dry_run)

    # ── Telegram
    if send_all:
        tg_sent = send_all_alerts(alerts, dry_run)
    else:
        tg_sent = send_summary(alerts, report_rows, dry_run)

    print(f'\n{"─"*50}')
    print(f'完成{"（dry-run）" if dry_run else ""}')
    print(f'  命中組數   ：{len(alerts)}')
    print(f'  報表列數   ：{report_rows}{"（未寫入）" if dry_run else ""}')
    if report_path:
        print(f'  報表路徑   ：{report_path}')
    if send_all:
        print(f'  Telegram   ：逐筆 {tg_sent} 則{"（未發送）" if dry_run else ""}')
    else:
        print(f'  Telegram   ：摘要 {tg_sent} 則{"（未發送）" if dry_run else ""}')

    if dry_run:
        print('\n確認無誤後請執行（輸出報表 + 推播摘要）：')
        print('  python3 scripts/realprice_alert.py')
        print('\n確認報表後，逐筆推播：')
        print('  python3 scripts/realprice_alert.py --send-all')


if __name__ == '__main__':
    main()
