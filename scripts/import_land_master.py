#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_land_master.py — 土地主清冊「總表」→ SQLite 匯入工具

用法：
  python3 scripts/import_land_master.py
  python3 scripts/import_land_master.py --file data/land_master/land_master.xlsx
  python3 scripts/import_land_master.py --dry-run     # 解析不寫入
  python3 scripts/import_land_master.py --analyze     # event_key 差異報告（不寫入）
  python3 scripts/import_land_master.py --schema
  python3 scripts/import_land_master.py --sample

唯一鍵設計（v2）：
  event_key = SHA16(land_match_key | owner_key | reg_seq | reg_date | reg_reason | share_numer | share_denom)
  用途：識別「同地主同地號的不同登記事件」，同時排除純重複列。
"""

import argparse
import hashlib
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX  = Path('/Users/xiaomingyang/Desktop/excel土地資料維護/最新完成版/老蕭LAND_MASTER.xlsx')
DEFAULT_DB    = PROJECT_ROOT / 'data' / 'database' / 'land_master.db'

ALLOWED_SHEET   = '總表'
ALLOWED_SHEET_ALT = '土地主清冊'  # 老蕭LAND_MASTER.xlsx 使用此工作表名稱
FORBIDDEN_SHEETS = {
    '隱匿地址', '退件區', '工作表5',
    '縣市選單', '地段小段選單', '分區原因選單',
    '郵遞區號表1', '郵遞區號表2', '不重要隱匿地址',
}


# ── 欄位對應（header-name based，不依賴欄位順序）──────────────────
# 欄名 → (db_col, action)
# action: keep / skip / masked_id / full_id / phone
# aliases: 新舊 Excel 欄名不同但語意相同的，統一對應到同一 db_col
HEADER_MAP: dict[str, tuple[str | None, str]] = {
    '更新日期':        ('updated_at',      'keep'),
    '分區':           ('zone_type',       'keep'),
    '位置':           ('location_tag',    'keep'),
    '縣市':           ('city',            'keep'),
    '地區':           ('district',        'keep'),
    '地段':           ('section_raw',     'keep'),
    '小段':           ('sub_section',     'keep'),
    '地號':           ('land_no_raw',     'keep'),
    '公告現值':        ('announced_value', 'keep'),
    '次序':           ('reg_seq',         'keep'),
    '登記次序':        ('reg_seq',         'keep'),   # alias
    '登記日期':        ('reg_date',        'keep'),
    '登記原因':        ('reg_reason',      'keep'),
    '發生日期':        ('cause_date',      'keep'),
    '所有權人':        ('owner_name',      'keep'),
    '統一編號（遮罩）':  ('owner_id_masked', 'masked_id'),
    '統一編號遮罩':     ('owner_id_masked', 'masked_id'),  # alias
    '統一編號（完整）':  ('owner_id_full',   'full_id'),
    '統一編號完整':     ('owner_id_full',   'full_id'),    # alias
    '郵遞區號':        ('postal_code',     'keep'),
    '住址':           ('address',         'keep'),
    '已售出':         ('is_sold',         'keep'),
    '分母':           ('share_denom',     'keep'),
    '分子':           ('share_numer',     'keep'),
    '持分':           (None,              'skip'),   # 計算欄，不匯入
    '持分坪數':        (None,              'skip'),   # 計算欄，不匯入
    '土地總坪數':       ('total_area_ping', 'keep'),
    '權利範圍':        ('ownership_range', 'keep'),
    '備註':           ('note',            'keep'),
    '進價':           ('purchase_price',  'keep'),
    '電話':           (None,              'phone'),
    # 舊 Excel 特有欄（忽略）
    '實價命中':        (None,              'skip'),
    '實價日期':        (None,              'skip'),
    '實價總價(萬)':     (None,              'skip'),
    '同批命中地號':      (None,              'skip'),
    '建議動作':        (None,              'skip'),
    # 系統判定欄位（AB–AF），由程式自動寫入，import 時略過
    '系統處理狀態':     (None,              'skip'),
    '系統處理備註':     (None,              'skip'),
    '系統來源':        (None,              'skip'),
    '系統更新時間':     (None,              'skip'),
    '系統批次ID':      (None,              'skip'),
}

# 必要欄位（欄名），缺任何一個就停止
REQUIRED_HEADERS = ['縣市', '地區', '地段', '地號', '所有權人']

REQUIRED_DB_COLS = ['city', 'district', 'section_raw']


def build_col_index_map(header_row: list) -> dict[str, int]:
    """
    從標題列建立「欄名 → 欄索引」對應。
    回傳 {col_name: index}，並驗證必要欄位是否存在。
    """
    col_map = {}
    for i, h in enumerate(header_row):
        name = str(h).strip() if h is not None else ''
        if name:
            col_map[name] = i

    missing = [h for h in REQUIRED_HEADERS if h not in col_map]
    if missing:
        raise ValueError(f'Excel 缺少必要欄位：{missing}，停止匯入。')
    return col_map


# ── Normalize ────────────────────────────────────────────────────

def normalize_section(raw: str | None) -> str:
    """普義段(0835) / 普義段（0835） → 普義段"""
    if not raw:
        return ''
    s = str(raw).strip()
    s = re.sub(r'[\(（][^)\）]*[\)）]', '', s)
    return re.sub(r'\s+', '', s)


def normalize_land_no(raw: str | None) -> str:
    """
    100      → 0100-0000
    100-1    → 0100-0001
    100之1   → 0100-0001
    26-3     → 0026-0003
    0026-3   → 0026-0003
    """
    if not raw:
        return ''
    s = str(raw).strip()
    s = re.sub(r'之', '-', s)
    s = re.sub(r'[^\d\-]', '', s)
    if not s:
        return ''
    parts = s.split('-')
    try:
        main = int(parts[0]) if parts[0] else 0
        sub  = int(parts[1]) if len(parts) > 1 and parts[1] else 0
        return f'{main:04d}-{sub:04d}'
    except ValueError:
        return s


def make_land_match_key(city: str, district: str,
                        norm_sec: str, norm_no: str) -> str:
    return '_'.join(s.strip() for s in [city, district, norm_sec, norm_no] if s.strip())


def make_owner_key(owner_id_full: str | None, owner_name: str | None,
                   owner_id_masked: str | None) -> str:
    if owner_id_full and owner_id_full.strip():
        src = owner_id_full.strip().upper()
    else:
        name   = (owner_name   or '').strip()
        masked = (owner_id_masked or '').strip()
        src    = f'{name}|{masked}'
    if not src or src == '|':
        return ''
    return hashlib.sha256(src.encode()).hexdigest()[:16]


def calc_actual_area(total, numer, denom) -> float | None:
    try:
        t, n, d = float(total), float(numer), float(denom)
        return round(t * n / d, 4) if d else None
    except (TypeError, ValueError):
        return None


# ── SQLite DDL ───────────────────────────────────────────────────

DDL_MAIN = """
CREATE TABLE IF NOT EXISTS land_master (
    id                      INTEGER PRIMARY KEY AUTOINCREMENT,
    -- 人工維護欄位
    updated_at              TEXT,
    zone_type               TEXT,
    location_tag            TEXT,
    city                    TEXT,
    district                TEXT,
    section_raw             TEXT,
    sub_section             TEXT,
    land_no_raw             TEXT,
    announced_value         REAL,
    reg_seq                 TEXT,
    reg_date                TEXT,
    cause_date              TEXT,
    reg_reason              TEXT,
    owner_name              TEXT,
    owner_id_masked         TEXT,
    owner_id_full           TEXT,
    postal_code             TEXT,
    address                 TEXT,
    is_sold                 INTEGER DEFAULT 0,
    share_denom             REAL,
    share_numer             REAL,
    total_area_ping         REAL,
    ownership_range         TEXT,
    note                    TEXT,
    purchase_price          REAL,
    phone                   TEXT,
    -- Python 自動生成
    normalized_section      TEXT,
    normalized_land_no      TEXT,
    land_match_key          TEXT,
    owner_key               TEXT,
    event_key               TEXT,   -- 事件級唯一鍵（lmk+owner+seq+date+reason+持分）
    row_hash                TEXT,   -- 資料內容指紋，相同則 SKIP，不同則 UPDATE
    actual_owned_area       REAL,
    realprice_match_status  TEXT DEFAULT 'pending',
    last_realprice_check_at TEXT,
    telegram_last_note_at   TEXT,
    -- 系統欄位
    source_row              INTEGER,
    created_at              TEXT DEFAULT (datetime('now')),
    imported_at             TEXT DEFAULT (datetime('now')),
    sys_status              TEXT,
    sys_note                TEXT,
    sys_source              TEXT,
    sys_updated_at          TEXT,
    sys_batch_id            TEXT
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_lm_event_key
    ON land_master(event_key) WHERE event_key != '';
CREATE INDEX IF NOT EXISTS idx_lm_land_match_key ON land_master(land_match_key);
CREATE INDEX IF NOT EXISTS idx_lm_owner_key      ON land_master(owner_key);
CREATE INDEX IF NOT EXISTS idx_lm_location_tag   ON land_master(location_tag);
CREATE INDEX IF NOT EXISTS idx_lm_city           ON land_master(city);
CREATE INDEX IF NOT EXISTS idx_lm_district       ON land_master(district);
"""

DDL_LOOKUP = """
CREATE TABLE IF NOT EXISTS land_section_lookup (
    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
    city               TEXT,
    district           TEXT,
    section_raw        TEXT,
    normalized_section TEXT,
    created_at         TEXT DEFAULT (datetime('now')),
    updated_at         TEXT DEFAULT (datetime('now')),
    UNIQUE(city, district, section_raw)
);
CREATE INDEX IF NOT EXISTS idx_lsl_city     ON land_section_lookup(city);
CREATE INDEX IF NOT EXISTS idx_lsl_district ON land_section_lookup(district);
"""

INSERT_SQL = """
INSERT INTO land_master (
    updated_at, zone_type, location_tag, city, district,
    section_raw, sub_section, land_no_raw,
    announced_value, reg_seq, reg_date, cause_date, reg_reason,
    owner_name, owner_id_masked, owner_id_full,
    postal_code, address, is_sold,
    share_denom, share_numer, total_area_ping,
    ownership_range, note, purchase_price, phone,
    normalized_section, normalized_land_no,
    land_match_key, owner_key, event_key, row_hash, actual_owned_area,
    source_row, created_at, imported_at
) VALUES (
    :updated_at, :zone_type, :location_tag, :city, :district,
    :section_raw, :sub_section, :land_no_raw,
    :announced_value, :reg_seq, :reg_date, :cause_date, :reg_reason,
    :owner_name, :owner_id_masked, :owner_id_full,
    :postal_code, :address, :is_sold,
    :share_denom, :share_numer, :total_area_ping,
    :ownership_range, :note, :purchase_price, :phone,
    :normalized_section, :normalized_land_no,
    :land_match_key, :owner_key, :event_key, :row_hash, :actual_owned_area,
    :source_row, datetime('now'), datetime('now')
)
"""

UPDATE_SQL = """
UPDATE land_master SET
    updated_at         = :updated_at,
    zone_type          = :zone_type,
    location_tag       = :location_tag,
    announced_value    = :announced_value,
    cause_date         = :cause_date,
    owner_name         = :owner_name,
    owner_id_masked    = :owner_id_masked,
    owner_id_full      = COALESCE(:owner_id_full, owner_id_full),
    postal_code        = :postal_code,
    address            = :address,
    is_sold            = :is_sold,
    total_area_ping    = :total_area_ping,
    ownership_range    = :ownership_range,
    note               = :note,
    purchase_price     = :purchase_price,
    phone              = :phone,
    normalized_section = :normalized_section,
    normalized_land_no = :normalized_land_no,
    row_hash           = :row_hash,
    actual_owned_area  = :actual_owned_area,
    source_row         = :source_row,
    imported_at        = datetime('now')
WHERE event_key = :event_key
"""


# ── 資料清理 ─────────────────────────────────────────────────────

def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ('', 'nan', 'None', 'NaN') else s


def _to_real(v) -> float | None:
    s = _clean(v)
    if not s:
        return None
    s = re.sub(r'[,，\s]', '', s)
    try:
        return float(s)
    except ValueError:
        return None


def parse_row(cells: list, row_idx: int,
              col_map: dict[str, int] | None = None) -> dict:
    """
    將一列 cell 值轉成 DB record dict。
    col_map: build_col_index_map() 的結果，{欄名: 欄索引}。
    若為 None 則退回舊 index 模式（僅供向下相容，不建議使用）。
    """
    rec: dict = {}
    phones: list[str] = []

    if col_map is None:
        raise ValueError('parse_row 需要 col_map，請用 build_col_index_map() 建立。')

    # 已對應到同一 db_col 的欄名只取第一個出現的（alias 去重）
    seen_db_col: set[str] = set()
    for col_name, (db_col, action) in HEADER_MAP.items():
        col_idx = col_map.get(col_name)
        if col_idx is None:
            continue   # 此 Excel 無此欄，略過

        raw = cells[col_idx] if col_idx < len(cells) else None
        val = _clean(raw)

        if action == 'skip':
            continue
        elif action == 'phone':
            if val:
                phones.append(val)
        elif action in ('masked_id', 'full_id', 'keep'):
            if db_col not in seen_db_col:
                rec[db_col] = val
                seen_db_col.add(db_col)

    rec['phone'] = '、'.join(phones) if phones else None

    # 數值欄位轉型
    rec['announced_value'] = _to_real(rec.get('announced_value'))
    rec['share_denom']     = _to_real(rec.get('share_denom'))
    rec['share_numer']     = _to_real(rec.get('share_numer'))
    rec['total_area_ping'] = _to_real(rec.get('total_area_ping'))
    rec['purchase_price']  = _to_real(rec.get('purchase_price'))

    # 已售出 → 0/1
    sold_raw = rec.get('is_sold')
    rec['is_sold'] = 1 if sold_raw else 0

    # Normalize
    sec_raw  = rec.get('section_raw') or ''
    no_raw   = rec.get('land_no_raw') or ''
    city     = rec.get('city') or ''
    district = rec.get('district') or ''

    norm_sec = normalize_section(sec_raw)
    norm_no  = normalize_land_no(no_raw)

    rec['normalized_section'] = norm_sec
    rec['normalized_land_no'] = norm_no
    rec['land_match_key']     = make_land_match_key(city, district, norm_sec, norm_no)
    rec['owner_key']          = make_owner_key(
        rec.get('owner_id_full'), rec.get('owner_name'), rec.get('owner_id_masked'))
    rec['actual_owned_area']  = calc_actual_area(
        rec.get('total_area_ping'), rec.get('share_numer'), rec.get('share_denom'))
    rec['event_key']          = make_event_key(rec)
    rec['row_hash']           = make_row_hash(rec)
    rec['source_row'] = row_idx

    return rec


def make_row_hash(rec: dict) -> str:
    """
    資料內容指紋：對所有可能被人工修改的欄位計算 SHA16。
    event_key 相同但 row_hash 不同 → 需要 UPDATE。
    row_hash 相同 → 資料未變，SKIP。
    """
    fields = [
        'updated_at', 'zone_type', 'location_tag', 'city', 'district',
        'section_raw', 'sub_section', 'land_no_raw',
        'announced_value', 'reg_seq', 'reg_date', 'cause_date', 'reg_reason',
        'owner_name', 'owner_id_masked', 'owner_id_full',
        'postal_code', 'address', 'is_sold',
        'share_denom', 'share_numer', 'total_area_ping',
        'ownership_range', 'note', 'purchase_price', 'phone',
    ]
    src = '|'.join(str(rec.get(f) or '') for f in fields)
    return hashlib.sha256(src.encode()).hexdigest()[:16]


def make_event_key(rec: dict) -> str:
    """
    事件級唯一鍵：同地主同地號的不同登記事件各有唯一 key。
    key 組成：land_match_key | owner_key | owner_name | reg_seq | reg_date | reg_reason | share_numer | share_denom | source_row
    - land_no_raw 為 NULL 時，land_match_key 缺地號，加入 source_row 確保每列唯一，避免 SQLite NULL UNIQUE 失效。
    - 空值統一用空字串。
    """
    lmk = rec.get('land_match_key') or ''
    okey = rec.get('owner_key') or ''
    has_land_no = bool((rec.get('land_no_raw') or '').strip())
    parts = [
        lmk,
        okey,
        rec.get('owner_name') or '',
        rec.get('reg_seq')    or '',
        rec.get('reg_date')   or '',
        rec.get('reg_reason') or '',
        str(rec.get('share_numer')  or ''),
        str(rec.get('share_denom') or ''),
    ]
    if not has_land_no:
        # 無地號：加入 source_row 使每列有唯一 key，防止重複插入
        parts.append(str(rec.get('source_row') or ''))
    src = '|'.join(parts)
    if not lmk and not okey:   # 連縣市地段人名都空 → 整列無意義
        return ''
    return hashlib.sha256(src.encode()).hexdigest()[:20]


# ── event_key 差異分析（dry-run 專用）────────────────────────────

def analyze_event_key(records: list[dict]):
    """
    比較舊 key（land_match_key + owner_key）vs 新 event_key，
    輸出差異報告，不寫入任何 DB。
    """
    from collections import defaultdict

    total = len(records)

    # 舊 key：land_match_key + owner_key
    old_key_set: set[str] = set()
    old_key_groups: dict[str, list] = defaultdict(list)
    for r in records:
        k = f"{r['land_match_key']}||{r['owner_key']}"
        old_key_set.add(k)
        old_key_groups[k].append(r)

    # 新 event_key
    event_key_set:  set[str] = set()
    event_key_dupes: set[str] = set()   # event_key 為空或重複
    seen_event: dict[str, dict] = {}
    pure_dupes = 0

    for r in records:
        ek = r['event_key']
        if not ek:
            pure_dupes += 1
            continue
        if ek in seen_event:
            pure_dupes += 1
        else:
            seen_event[ek] = r
            event_key_set.add(ek)

    uniq_old   = len(old_key_set)
    uniq_event = len(event_key_set)
    extra_events = uniq_event - uniq_old   # 新 key 比舊 key 多出的事件數

    print()
    print('━' * 60)
    print('event_key 差異分析報告')
    print('━' * 60)
    print(f'  Excel 總列數（非空）       ：{total:>8,}')
    print(f'  舊 key（lmk+owner）唯一數  ：{uniq_old:>8,}  ← 目前 UPSERT 只保留這些')
    print(f'  新 event_key 唯一數        ：{uniq_event:>8,}  ← 改用後可保留的事件數')
    print(f'  純重複列（完全相同 → 排除）：{pure_dupes:>8,}')
    print(f'  額外保留的歷史事件         ：{extra_events:>8,}  （原本被 UPSERT 覆蓋）')
    print('━' * 60)

    # 找出舊 key 衝突但 event_key 不同的案例（前 10）
    print('\n前 10 筆「同 lmk+owner，但 event_key 不同」案例：')
    print('（這些是被舊 UPSERT 覆蓋、改用 event_key 後可保留的歷史事件）\n')

    shown = 0
    for old_k, group in sorted(old_key_groups.items(),
                                key=lambda x: -len(x[1])):
        if len(group) <= 1:
            continue
        # 去掉純重複，只看 event_key 不同的
        uniq_ek = {}
        for r in group:
            ek = r['event_key']
            if ek and ek not in uniq_ek:
                uniq_ek[ek] = r
        if len(uniq_ek) <= 1:
            continue

        first = list(uniq_ek.values())[0]
        print(f"  [{shown+1}] {first['land_match_key']}")
        print(f"      owner: {first['owner_name']!r}  共 {len(group)} 列 → {len(uniq_ek)} 個不同事件")
        for i, (ek, r) in enumerate(list(uniq_ek.items())[:5]):
            print(f"      事件{i+1}: seq={r['reg_seq']!r:8}  "
                  f"date={r['reg_date']!r:20}  "
                  f"reason={r['reg_reason']!r:12}  "
                  f"持分={r['share_numer']}/{r['share_denom']}")
        if len(uniq_ek) > 5:
            print(f"      … 還有 {len(uniq_ek)-5} 個事件")
        print()
        shown += 1
        if shown >= 10:
            break

    print('━' * 60)
    print(f'✅ 確認無誤後，執行正式匯入（不帶 --analyze）即可重建 DB。')
    print('━' * 60)


# ── 主流程 ───────────────────────────────────────────────────────

def run_import(xlsx_path: Path, db_path: Path, dry_run: bool = False,
               rebuild: bool = False) -> dict:
    print(f"讀取 Excel：{xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # 安全檢查：允許 總表 或 土地主清冊
    _sheet = ALLOWED_SHEET if ALLOWED_SHEET in wb.sheetnames else (
             ALLOWED_SHEET_ALT if ALLOWED_SHEET_ALT in wb.sheetnames else None)
    if _sheet is None:
        print(f'❌ 找不到「{ALLOWED_SHEET}」或「{ALLOWED_SHEET_ALT}」sheet，停止。')
        print(f'   現有 sheets：{wb.sheetnames}')
        wb.close()
        sys.exit(1)

    ws = wb[_sheet]
    rows = ws.iter_rows(values_only=True)
    header_row = list(next(rows))   # 第 1 列 = 欄位名稱

    # 建立欄名對應（必要欄位缺失則停止）
    try:
        col_map = build_col_index_map(header_row)
    except ValueError as e:
        print(f'❌ {e}')
        wb.close()
        sys.exit(1)

    mapped = {h: col_map[h] for h in col_map if h in HEADER_MAP}
    print(f'  欄名對應：共 {len(header_row)} 欄，成功對應 {len(mapped)} 欄')

    records      = []
    errors       = []
    lookup_set: set[tuple] = set()

    for row_idx, row in enumerate(rows, start=2):
        if all(v is None for v in row):
            continue   # 跳過全空列
        try:
            rec = parse_row(list(row), row_idx, col_map=col_map)
            records.append(rec)
            # lookup 資料
            lk = (rec.get('city') or '', rec.get('district') or '',
                  rec.get('section_raw') or '', rec.get('normalized_section') or '')
            if lk[2]:   # section_raw 非空才收錄
                lookup_set.add(lk)
        except Exception as e:
            errors.append((row_idx, str(e)))

    wb.close()
    print(f'  解析完成：{len(records):,} 筆，錯誤：{len(errors)} 筆')
    if errors:
        for ri, msg in errors[:10]:
            print(f'    Row {ri}: {msg}')

    # 缺必要 DB 欄位值檢查（以第一筆有效資料為準）
    if records:
        sample = records[0]
        missing = [c for c in REQUIRED_DB_COLS if not sample.get(c)]
        if missing:
            print(f'⚠️  必要 DB 欄位值缺失（第 2 列）：{missing}')
            print('   請確認 Excel 包含對應欄位。')

    if dry_run:
        print('\n[dry-run] 未寫入資料庫，前 3 筆預覽：')
        for r in records[:3]:
            print(f"  lmk={r['land_match_key']!r}  "
                  f"owner={r['owner_name']!r}  "
                  f"event_key={r['event_key']!r}  "
                  f"area={r['actual_owned_area']}")
        return dict(total=len(records), inserted=0, updated=0, skipped=0,
                    errors=len(errors), lookup=len(lookup_set), dry_run=True)

    # ── 寫入 SQLite ──────────────────────────────────────────────
    db_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Phase 1：DDL（用獨立 connection，executescript 後即關閉）
    ddl_con = sqlite3.connect(db_path)
    ddl_con.execute('PRAGMA journal_mode=WAL')
    ddl_con.execute('PRAGMA synchronous=NORMAL')
    if rebuild:
        ddl_con.execute('BEGIN')
        ddl_con.execute('DELETE FROM land_master')
        ddl_con.execute('DELETE FROM land_section_lookup')
        ddl_con.execute('COMMIT')
        print('  ♻️  rebuild 模式：已清空 land_master + land_section_lookup')
    ddl_con.executescript(DDL_MAIN)
    ddl_con.executescript(DDL_LOOKUP)
    # 既有 DB 若缺 row_hash 欄位，自動補上（向下相容）
    cols = {row[1] for row in ddl_con.execute('PRAGMA table_info(land_master)')}
    if 'row_hash' not in cols:
        ddl_con.execute('BEGIN')
        ddl_con.execute('ALTER TABLE land_master ADD COLUMN row_hash TEXT')
        ddl_con.execute('COMMIT')
        print('  ℹ️  已新增 row_hash 欄位（既有 DB migration）')
    ddl_con.close()   # 關閉 DDL connection，清除 executescript 殘留狀態

    # ── Phase 2：讀取 + 寫入（全新 connection，isolation_level=None 手動管理事務）
    con = sqlite3.connect(db_path, isolation_level=None)
    con.execute('PRAGMA journal_mode=WAL')
    con.execute('PRAGMA synchronous=NORMAL')
    con.execute('PRAGMA cache_size=-65536')   # 64MB cache

    # ① 一次性讀取既有 event_key + row_hash → dict，避免逐筆 SELECT
    existing: dict[str, str] = {
        row[0]: row[1]
        for row in con.execute(
            "SELECT event_key, COALESCE(row_hash,'') FROM land_master WHERE event_key != ''")
    }
    print(f'  DB 既有 event_key：{len(existing):,} 筆')

    # ② 分流：INSERT / UPDATE（hash 變） / SKIP（hash 同） / 無 key INSERT
    to_insert:      list[dict] = []
    to_update:      list[dict] = []
    to_skip:        int        = 0
    to_insert_nokey: list[dict] = []

    for rec in records:
        ek = rec['event_key']
        if ek:
            if ek not in existing:
                to_insert.append(rec)
            elif existing[ek] != rec['row_hash']:
                to_update.append(rec)
            else:
                to_skip += 1
        else:
            to_insert_nokey.append(rec)

    # Excel 內部去重：同一 event_key 只保留最後一筆（source_row 最大）
    seen_in_batch: dict[str, dict] = {}
    for rec in to_insert:
        seen_in_batch[rec['event_key']] = rec
    to_insert = list(seen_in_batch.values())

    print(f'  新增：{len(to_insert):,}  更新：{len(to_update):,}  '
          f'SKIP：{to_skip:,}  無 key：{len(to_insert_nokey):,}')

    # ③ 整批寫入（isolation_level=None 模式下明確 BEGIN/COMMIT）
    BATCH = 5000
    try:
        con.execute('BEGIN')
        for i in range(0, len(to_insert), BATCH):
            con.executemany(INSERT_SQL, to_insert[i:i+BATCH])
            print(f'    INSERT {min(i+BATCH, len(to_insert)):,}/{len(to_insert):,}', end='\r')
        if to_insert:
            print()
        for i in range(0, len(to_update), BATCH):
            con.executemany(UPDATE_SQL, to_update[i:i+BATCH])
            print(f'    UPDATE {min(i+BATCH, len(to_update)):,}/{len(to_update):,}', end='\r')
        if to_update:
            print()
        if to_insert_nokey:
            for i in range(0, len(to_insert_nokey), BATCH):
                con.executemany(INSERT_SQL, to_insert_nokey[i:i+BATCH])
        # lookup UPSERT
        con.executemany("""
            INSERT INTO land_section_lookup
                (city, district, section_raw, normalized_section, updated_at)
            VALUES (?, ?, ?, ?, datetime('now'))
            ON CONFLICT(city, district, section_raw) DO UPDATE SET
                normalized_section = excluded.normalized_section,
                updated_at = datetime('now')
        """, list(lookup_set))
        con.execute('COMMIT')
        print('  ✅ COMMIT 完成')
    except Exception as e:
        con.execute('ROLLBACK')
        con.close()
        raise RuntimeError(f'寫入失敗，已 ROLLBACK：{e}') from e

    con.close()
    return dict(total=len(records),
                inserted=len(to_insert) + len(to_insert_nokey),
                updated=len(to_update),
                skipped=to_skip,
                errors=len(errors), lookup=len(lookup_set), dry_run=False)


# ── 驗收工具 ─────────────────────────────────────────────────────

def show_schema(db_path: Path):
    if not db_path.exists():
        print(f'DB 不存在：{db_path}'); return
    con = sqlite3.connect(db_path)
    for tbl in ('land_master', 'land_section_lookup'):
        row = con.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (tbl,)
        ).fetchone()
        if row:
            print(f'\n[{tbl}]\n{row[0]}\n')
    con.close()


def show_sample_queries(db_path: Path):
    if not db_path.exists():
        print(f'DB 不存在：{db_path}'); return
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    q1 = """
        SELECT city, district, normalized_section, normalized_land_no,
               owner_name, actual_owned_area, location_tag
        FROM land_master
        WHERE location_tag IS NOT NULL AND location_tag != ''
        LIMIT 5
    """
    print('[Sample 1] 有位置標記的地主，前 5 筆')
    for r in con.execute(q1).fetchall():
        print(f"  {dict(r)}")

    q2 = """
        SELECT city, COUNT(*) cnt, ROUND(SUM(actual_owned_area),1) total_ping
        FROM land_master
        WHERE is_sold = 1
        GROUP BY city ORDER BY cnt DESC LIMIT 5
    """
    print('\n[Sample 2] 已售出統計（by 縣市）')
    for r in con.execute(q2).fetchall():
        print(f"  {dict(r)}")

    q3 = """
        SELECT city, district, normalized_section, COUNT(*) cnt
        FROM land_section_lookup
        GROUP BY city, district ORDER BY cnt DESC LIMIT 5
    """
    print('\n[Sample 3] 地段 lookup 筆數最多的縣市行政區')
    for r in con.execute(q3).fetchall():
        print(f"  {dict(r)}")

    total     = con.execute('SELECT COUNT(*) FROM land_master').fetchone()[0]
    lk_total  = con.execute('SELECT COUNT(*) FROM land_section_lookup').fetchone()[0]
    print(f'\n  land_master 總筆數：{total:,}')
    print(f'  land_section_lookup 筆數：{lk_total:,}')
    con.close()


# ── CLI ──────────────────────────────────────────────────────────

def print_summary(res: dict, db_path: Path):
    print()
    print('━' * 52)
    print('匯入結果')
    print('━' * 52)
    print(f"  解析總筆數：{res['total']:>8,}")
    print(f"  新增：      {res['inserted']:>8,}")
    print(f"  更新：      {res['updated']:>8,}")
    print(f"  SKIP：      {res.get('skipped',0):>8,}  （資料未變）")
    print(f"  錯誤：      {res['errors']:>8,}")
    print(f"  lookup 筆數：{res['lookup']:>7,}")
    if not res['dry_run']:
        print(f"  資料庫：    {db_path}")
    print('━' * 52)


def main():
    ap = argparse.ArgumentParser(description='土地主清冊 總表 → SQLite 匯入工具')
    ap.add_argument('--file',    default=str(DEFAULT_XLSX))
    ap.add_argument('--db',      default=str(DEFAULT_DB))
    ap.add_argument('--dry-run', action='store_true', help='只解析，不寫入')
    ap.add_argument('--rebuild', action='store_true', help='清空後全量重建（非增量）')
    ap.add_argument('--analyze', action='store_true', help='輸出 event_key 差異分析報告（不寫入）')
    ap.add_argument('--schema',  action='store_true', help='顯示 table schema')
    ap.add_argument('--sample',  action='store_true', help='顯示 sample query')
    args = ap.parse_args()

    xlsx = Path(args.file)
    db   = Path(args.db)

    if args.schema or args.sample:
        if args.schema: show_schema(db)
        if args.sample: show_sample_queries(db)
        return

    if not xlsx.exists():
        print(f'❌ 找不到 Excel：{xlsx}')
        print(f'   請將「總表.xlsx」放置於：{xlsx}')
        sys.exit(1)

    if args.analyze:
        # 只讀解析 + 分析，完全不寫入
        res = run_import(xlsx, db, dry_run=True)
        # re-parse to get records with event_key for analysis
        # 直接重跑解析（analyze 模式下不在意時間）
        import openpyxl as _ox
        wb = _ox.load_workbook(xlsx, read_only=True, data_only=True)
        _s = ALLOWED_SHEET if ALLOWED_SHEET in wb.sheetnames else ALLOWED_SHEET_ALT
        ws = wb[_s]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        _cm = build_col_index_map(hdr)
        recs = []
        for row_idx, row in enumerate(it, start=2):
            if all(v is None for v in row): continue
            try:
                recs.append(parse_row(list(row), row_idx, col_map=_cm))
            except Exception:
                pass
        wb.close()
        analyze_event_key(recs)
        return

    res = run_import(xlsx, db, dry_run=args.dry_run, rebuild=args.rebuild)
    print_summary(res, db)

    if not args.dry_run:
        show_schema(db)
        show_sample_queries(db)


if __name__ == '__main__':
    main()
