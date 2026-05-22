#!/usr/bin/env python3

import argparse
import csv
import sqlite3
from pathlib import Path
from datetime import date
import re
import json
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = PROJECT_ROOT / 'config' / 'realprice_config.json'
DB_PATH = PROJECT_ROOT / 'db' / 'land_intel.db'
OUTPUT_DIR = PROJECT_ROOT / 'output'


def load_config():
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def subtract_months(date_obj, months):
    year = date_obj.year
    month = date_obj.month - months
    while month <= 0:
        month += 12
        year -= 1
    return date(year, month, date_obj.day)


def parse_natural_query(query, config):
    params = {}
    city_match = re.search(r'(\w+市)', query)
    if city_match:
        params['city'] = city_match.group(1)
    district_match = re.search(r'(\w+區)', query)
    if district_match:
        params['district'] = district_match.group(1).strip()
    section_match = re.search(r'([\u4e00-\u9fff0-9]{1,10}?段)([\u4e00-\u9fff0-9]{0,10}小段)?', query)
    if section_match:
        params['section_name'] = section_match.group(1).strip()
        if section_match.group(2):
            params['subsection'] = section_match.group(2).strip()
    today = date.today()
    if '今年' in query:
        params['start_date'] = date(today.year, 1, 1).isoformat()
    elif '近半年' in query or '半年' in query:
        params['start_date'] = subtract_months(today, 6).isoformat()
    elif '三個月' in query or '三月' in query:
        params['start_date'] = subtract_months(today, 3).isoformat()
    else:
        custom = re.search(r'(\d+)\s*個月', query)
        if custom:
            months = int(custom.group(1))
            params['start_date'] = subtract_months(today, months).isoformat()
    return params, query


def build_query(params):
    sql = [
        'SELECT city, district, location_raw, trade_date, area_ping, unit_price_per_ping_wan, total_price_wan, land_use_zone, note, transaction_target'
        ' FROM land_transactions WHERE 1=1'
    ]
    values = []
    if params.get('city'):
        sql.append('AND city = ?')
        values.append(params['city'])
    if params.get('district'):
        sql.append('AND district = ?')
        values.append(params['district'])
    if params.get('section_name'):
        sql.append('AND (section_name LIKE ? OR location_raw LIKE ?)')
        values.extend([f"%{params['section_name']}%", f"%{params['section_name']}%"])
    if params.get('subsection'):
        sql.append('AND location_raw LIKE ?')
        values.append(f"%{params['subsection']}%")
    if params.get('start_date'):
        sql.append('AND trade_date >= ?')
        values.append(params['start_date'])
    sql.append('ORDER BY trade_date DESC')
    return ' '.join(sql), values


def is_recommend(note, land_use_zone, excluded_note_terms):
    # Check note for excluded terms
    if note and any(term in note.strip() for term in excluded_note_terms):
        return False
    
    # Check land use zone for excluded types
    excluded_land_uses = ['人行步道', '道路保留地', '公共設施保留地']
    if land_use_zone and any(term in land_use_zone for term in excluded_land_uses):
        return False
    
    return True


def format_row(row, excluded_note_terms):
    city, district, location_raw, trade_date, area_ping, unit_price_per_ping_wan, total_price_wan, land_use_zone, note, transaction_target = row
    return {
        'city': city,
        'district': district,
        'location_raw': location_raw,
        'trade_date': trade_date,
        'area_ping': area_ping,
        'unit_price_per_ping_wan': unit_price_per_ping_wan,
        'total_price_wan': total_price_wan,
        'land_use_zone': land_use_zone,
        'note': note,
        'transaction_target': transaction_target,
        'is_recommend': is_recommend(note, land_use_zone, excluded_note_terms)
    }


def parse_location(location_raw):
    """Parse location_raw to extract section and land_number"""
    if not location_raw:
        return '', ''
    
    # Try to match patterns like "內海墘段725地號"
    match = re.search(r'(.+段)(.+地號)', location_raw)
    if match:
        section = match.group(1).strip()
        land_number = match.group(2).replace('地號', '').strip()
        return section, land_number
    
    # Fallback: try to find segment
    segment_match = re.search(r'(.+段)', location_raw)
    if segment_match:
        section = segment_match.group(1).strip()
        remaining = location_raw.replace(section, '').strip()
        return section, remaining
    
    return '', location_raw


def parse_section_and_subsection(location_raw):
    if not location_raw:
        return '', ''
    match = re.search(r'([\u4e00-\u9fff0-9]{1,10}段)([\u4e00-\u9fff0-9]{1,10}小段)?', location_raw)
    if match:
        section = match.group(1).strip()
        subsection = match.group(2).strip() if match.group(2) else ''
        return section, subsection
    return '', ''


def prepare_export_data(results):
    """Prepare data for export with parsed fields"""
    export_data = []
    for row in results:
        section, subsection = parse_section_and_subsection(row['location_raw'])
        if not section:
            section, _ = parse_location(row['location_raw'])
        _, land_number = parse_location(row['location_raw'])
        export_data.append({
            '交易日期': row['trade_date'],
            '縣市': row['city'],
            '行政區': row['district'],
            '地段': section,
            '小段': subsection,
            '地號': land_number,
            '坪數': row['area_ping'],
            '單價(萬/坪)': row['unit_price_per_ping_wan'],
            '總價(萬)': row['total_price_wan'],
            '使用分區': row['land_use_zone'] or '',
            '備註': row['note'] or '',
            '是否值得調謄本': '是' if row['is_recommend'] else '否'
        })
    return export_data


def export_csv(data, filename):
    """Export data to CSV"""
    if not data:
        return
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    filepath = OUTPUT_DIR / filename
    
    fieldnames = [
        '交易日期', '縣市', '行政區', '地段', '小段', '地號',
        '坪數', '單價(萬/坪)', '總價(萬)', '使用分區', '備註', '是否值得調謄本'
    ]
    with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    
    return filepath


def export_excel(data, filename):
    """Export data to Excel using openpyxl"""
    if not data:
        return
    
    OUTPUT_DIR.mkdir(exist_ok=True)
    filepath = OUTPUT_DIR / filename
    
    wb = Workbook()
    ws = wb.active
    ws.title = "土地成交資料"
    
    # Write headers
    headers = list(data[0].keys())
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data[header]
            # Handle None values
            if value is None:
                value = ''
            ws.cell(row=row_num, column=col_num, value=value)
    
    wb.save(filepath)
    return filepath


def generate_filename(query, extension):
    """Generate filename based on query"""
    # Parse city and district from query
    city_match = re.search(r'(\w+市)', query)
    district_match = re.search(r'(\w+區)', query)
    
    city = city_match.group(1) if city_match else 'unknown'
    district = district_match.group(1) if district_match else 'unknown'
    
    # Convert to English lowercase
    city_en = {
        '台北市': 'taipei',
        '新北市': 'newtaipei', 
        '桃園市': 'taoyuan',
        '台中市': 'taichung',
        '新竹市': 'hsinchu'
    }.get(city, city.lower().replace('市', ''))
    
    district_en = district.lower().replace('區', '')
    
    today = date.today().strftime('%Y%m%d')
    return f"{city_en}_{district_en}_{today}.{extension}"


def main():
    parser = argparse.ArgumentParser(description='匯出土地成交資料報表')
    parser.add_argument('query', nargs='+', help='查詢條件，例如：桃園市 大園區 近半年')
    parser.add_argument('--db', default=str(DB_PATH), help='SQLite 資料庫路徑')
    args = parser.parse_args()

    config = load_config()
    params, query_text = parse_natural_query(' '.join(args.query), config)
    
    # Execute query
    conn = sqlite3.connect(args.db)
    sql, values = build_query(params)
    cursor = conn.execute(sql, values)
    rows = cursor.fetchall()
    conn.close()
    
    # Format results
    results = [format_row(row, config['excluded_note_terms']) for row in rows]
    
    if not results:
        print('查詢無資料')
        return
    
    # Prepare export data
    export_data = prepare_export_data(results)
    
    # Generate filenames
    csv_filename = generate_filename(query_text, 'csv')
    xlsx_filename = generate_filename(query_text, 'xlsx')
    
    # Export CSV
    csv_path = export_csv(export_data, csv_filename)
    print(f'CSV 檔案已匯出：{csv_path}')
    
    # Export Excel
    xlsx_path = export_excel(export_data, xlsx_filename)
    print(f'Excel 檔案已匯出：{xlsx_path}')
    
    print(f'共匯出 {len(export_data)} 筆資料')


if __name__ == '__main__':
    main()