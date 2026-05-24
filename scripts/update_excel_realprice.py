#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_excel_realprice.py — Excel 實價登錄比對更新

流程：
  1. 讀取主清冊 Excel
  2. 以縣市 + 地區 + 地段 + 地號比對實價登錄 SQLite
  3. 命中列新增 5 個輔助欄位，輸出新檔（原檔不動）

執行：
  python3 scripts/update_excel_realprice.py --dry-run        ← 只統計，不寫檔
  python3 scripts/update_excel_realprice.py                  ← 正式輸出
  python3 scripts/update_excel_realprice.py --input 其他.xlsx ← 指定來源
"""

import argparse
import re
import shutil
import sqlite3
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
LI_DB        = Path('/Users/xiaomingyang/Desktop/:Users:mingyang:land-ai:/land-intel-system/db/land_intel.db')

BASE_DIR    = Path('/Users/xiaomingyang/Desktop/excel土地資料維護')
INPUT_DIR   = BASE_DIR / 'input'
OUTPUT_DIR  = BASE_DIR / 'output'
BACKUP_DIR  = BASE_DIR / 'backup'
LOGS_DIR    = BASE_DIR / 'logs'
ARCHIVE_DIR = BASE_DIR / 'archive'
LATEST_DIR  = BASE_DIR / '最新完成版'
LATEST_FILE = LATEST_DIR / '老蕭LAND_MASTER.xlsx'

DEFAULT_XLSX = LATEST_FILE

# 輔助欄位（4 欄）：地號已正規化，不再額外輸出標準地號
NEW_COLS = ['實價命中', '實價日期', '實價總價(萬)', '建議動作']

# 曾用過但已廢棄的舊欄位名稱（若 Excel 內已存在，略過不寫）
OBSOLETE_COLS = {
    '標準地號', '同批命中地號', '實價提醒狀態', '實價成交日期',
    '最後比對日', '實價交易土地筆數',
}

# Excel 欄位名稱 mapping（允許不同版本的欄位名稱）
COL_ALIASES = {
    '縣市':   ['縣市'],
    '地區':   ['地區', '行政區', '鄉鎮市區'],
    '地段':   ['地段'],
    '地號':   ['地號'],
    '所有人':   ['所有人', '所有權人'],
    '次序':     ['次序', '登記次序'],
    '已售出':   ['已售出'],
    '備註':     ['備註'],
    '登記日期': ['登記日期'],
}

# ── 全形→半形工具 ────────────────────────────────────────────────

def full_to_half(s: str) -> str:
    """全形數字/英文/符號轉半形"""
    result = []
    for c in s:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif c == '　':
            result.append(' ')
        else:
            result.append(c)
    return ''.join(result)

# ── 資料統一 normalize 函式（只用於比對 key，不修改 Excel 原始內容）────

def norm_city(raw) -> str:
    """縣市統一：補「市」、臺→台"""
    if not raw:
        return ''
    s = str(raw).strip()
    s = s.replace('臺', '台')
    # 補「市」（若只有縣市名稱主幹）
    CITIES = ['桃園', '台北', '新北', '台中', '台南', '高雄', '基隆',
              '新竹', '嘉義', '苗栗', '彰化', '南投', '雲林', '屏東',
              '宜蘭', '花蓮', '台東', '澎湖', '金門', '連江']
    for c in CITIES:
        if s == c:
            return c + '市'
    COUNTIES = ['苗栗', '彰化', '南投', '雲林', '屏東', '宜蘭',
                '花蓮', '台東', '澎湖', '金門', '連江', '新竹', '嘉義']
    for c in COUNTIES:
        if s == c:
            return c + '縣'
    return s


def norm_district(raw) -> str:
    """行政區統一：補「區」"""
    if not raw:
        return ''
    s = str(raw).strip()
    s = full_to_half(s)
    DISTRICTS = [
        '大園', '蘆竹', '中壢', '桃園', '八德', '龜山', '楊梅', '大溪',
        '平鎮', '龍潭', '新屋', '觀音', '復興',
        '南屯', '西屯', '北屯', '中區', '東區', '西區', '南區', '北區',
        '大里', '太平', '霧峰', '烏日', '大肚', '龍井', '沙鹿', '梧棲',
        '清水', '大甲', '外埔', '大安', '神岡', '潭子', '后里', '石岡',
        '東勢', '和平', '新社',
        '中山', '大同', '中正', '萬華', '信義', '松山', '大安', '文山',
        '南港', '內湖', '士林', '北投',
    ]
    for d in DISTRICTS:
        if s == d:
            return d + '區'
    return s


def norm_section(raw) -> str:
    """地段統一：去括號代碼、去空白、全形→半形、保留「段」"""
    if not raw:
        return ''
    s = str(raw).strip()
    s = full_to_half(s)
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'（.*?）', '', s)
    s = re.sub(r'\s+', '', s)
    return s.strip()


def norm_land_no(raw) -> str:
    """
    地號統一為 XXXX-YYYY（主號4碼-子號4碼），供比對 key 使用。
    - 842.003  → 0842-0003  （小數點 = 主號.子號）
    - 08420003 → 0842-0003  （8位數純數字）
    - 96-3     → 0096-0003  （已有破折號）
    - 842      → 0842-0000  （純整數）
    - 3之51    → 0003-0051  （中文「之」）
    """
    if not raw:
        return ''
    s = str(raw).strip()
    s = full_to_half(s)
    s = re.sub(r'地號', '', s)
    s = re.sub(r'之', '-', s)
    s = re.sub(r'\s+', '', s)
    s = re.sub(r'[^\d\-\.]', '', s)
    if not s:
        return ''
    try:
        if '.' in s and '-' not in s:
            parts = s.split('.', 1)
            main = int(parts[0]) if parts[0] else 0
            sub  = int(parts[1]) if parts[1] else 0
            return f'{main:04d}-{sub:04d}'
        if '-' in s:
            parts = s.split('-', 1)
            main = int(parts[0]) if parts[0] else 0
            sub  = int(parts[1]) if parts[1] else 0
            return f'{main:04d}-{sub:04d}'
        digits = re.sub(r'\D', '', s)
        if len(digits) == 8:
            return f'{int(digits[:4]):04d}-{int(digits[4:]):04d}'
        return f'{int(digits):04d}-0000'
    except (ValueError, IndexError):
        return ''


def roc_date(td) -> str:
    if not td:
        return ''
    td = str(td).strip()
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', td)
    if m:
        roc = int(m.group(1)) - 1911
        return f'{roc}/{m.group(2)}/{m.group(3)}'
    return td


def parse_date(val) -> 'date | None':
    """解析各種日期格式，回傳 datetime.date 或 None。"""
    from datetime import date as _date, datetime as _datetime
    if not val:
        return None
    if isinstance(val, _datetime):
        return val.date()
    if isinstance(val, _date):
        return val
    s = str(val).strip()
    # 114年05月08日
    m = re.match(r'^(\d{2,3})年(\d{1,2})月(\d{1,2})日$', s)
    if m:
        try: return _date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
        except ValueError: pass
    # 114/05/08 或 114-05-08
    m = re.match(r'^(\d{2,3})[\/\-](\d{1,2})[\/\-](\d{1,2})$', s)
    if m:
        try: return _date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
        except ValueError: pass
    # 113.12.31
    m = re.match(r'^(\d{2,3})\.(\d{1,2})\.(\d{1,2})$', s)
    if m:
        try: return _date(int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3)))
        except ValueError: pass
    # 西元格式
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y%m%d'):
        try: return _datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

# ── 欄位辨識 ─────────────────────────────────────────────────────

def detect_columns(headers: list) -> dict[str, int]:
    """回傳 {標準欄位名: 0-based index}，找不到的欄位不含在 dict 內"""
    result = {}
    for std_name, aliases in COL_ALIASES.items():
        for i, h in enumerate(headers):
            if h and str(h).strip() in aliases:
                result[std_name] = i
                break
    return result


def detect_new_col_positions(headers: list) -> dict[str, int]:
    """回傳已存在的輔助欄位的 0-based index（供更新用）"""
    positions = {}
    for i, h in enumerate(headers):
        if h and str(h).strip() in NEW_COLS:
            positions[str(h).strip()] = i
    return positions

# ── 實價資料載入 ──────────────────────────────────────────────────

def load_transactions() -> dict:
    """
    回傳 index: (city, district, norm_section, norm_land_no) → list[dict]
    同一交易可能含多筆地號，用 unique_key 聚合
    """
    con = sqlite3.connect(LI_DB, timeout=10)
    con.row_factory = sqlite3.Row
    rows = [dict(r) for r in con.execute("""
        SELECT city, district, section_name, land_number,
               trade_date, total_price_wan, unique_key
        FROM land_transactions
        WHERE trade_date >= '2025-01-01'
          AND (target_category LIKE '%土地%' OR target_category IS NULL)
          AND land_number IS NOT NULL AND land_number != ''
        ORDER BY trade_date DESC
    """).fetchall()]
    con.close()

    idx = defaultdict(list)
    for r in rows:
        c   = norm_city(r['city'] or '')
        d   = norm_district(r['district'] or '')
        sec = norm_section(r['section_name'] or '')
        lno = norm_land_no(r['land_number'])
        if sec and lno:
            idx[(c, d, sec, lno)].append(r)
    return idx


def find_hits(city, district, section_raw, land_no_raw, tx_idx) -> list[dict]:
    """比對單筆主清冊列，回傳命中的實價交易列表（使用 normalize 後的 key）"""
    c   = norm_city(city)
    d   = norm_district(district)
    sec = norm_section(section_raw)
    lno = norm_land_no(land_no_raw)
    if not sec or not lno:
        return []

    hits = []
    for (tc, td, ts, tn), txs in tx_idx.items():
        if tc != c or td != d:
            continue
        if tn != lno:
            continue
        # 地段：包含比對（允許小段名稱差異）
        if sec not in ts and ts not in sec and sec != ts:
            continue
        hits.extend(txs)
    return hits


def build_hit_summary(hits: list[dict], is_sold: bool) -> dict:
    """把命中的多筆交易聚合成一筆摘要"""
    if not hits:
        return {}

    # 按 unique_key 聚合（同批交易的多地號）
    by_batch: dict[str, list] = defaultdict(list)
    for h in hits:
        by_batch[h['unique_key']].append(h)

    # 取最新一筆交易（按 trade_date）
    latest_batch_key = max(by_batch, key=lambda k: by_batch[k][0]['trade_date'])
    latest_batch = by_batch[latest_batch_key]
    sample = latest_batch[0]

    all_land_nos = sorted({norm_land_no(h['land_number']) for h in latest_batch})
    batch_label  = '、'.join(all_land_nos) if len(all_land_nos) > 1 else ''

    if is_sold:
        action = '已售地主再命中，確認是否需更新資料'
    elif len(all_land_nos) > 1:
        action = f'同批多地號（{len(all_land_nos)} 筆），請確認是否有地主異動'
    else:
        action = '請確認此地號是否已有地主異動'

    price = sample['total_price_wan']
    price_str = f'{price:,.0f}' if price else ''

    return {
        '實價命中':    '是',
        '實價日期':    roc_date(sample['trade_date']),
        '實價總價(萬)': price_str,
        '建議動作':    action,
    }

# ── log 寫出 ─────────────────────────────────────────────────────

def _write_log(log_path: Path, lines: list[str]):
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_path.write_text('\n'.join(lines), encoding='utf-8')

# ── 排序 key ──────────────────────────────────────────────────────

def sort_key(row: tuple, city_idx: int, dist_idx: int, sec_idx: int, lno_idx: int) -> tuple:
    city = str(row[city_idx] or '')
    dist = str(row[dist_idx] or '')
    sec  = norm_section(str(row[sec_idx]  or ''))
    lno  = norm_land_no(str(row[lno_idx]  or ''))
    return (city, dist, sec, lno)

# ── Excel 格式套用 ────────────────────────────────────────────────

# 欄寬設定
COL_WIDTHS = {
    '更新日期': 11, '分區': 8, '位置': 14, '縣市': 6, '地區': 8,
    '地段': 14, '地號': 12, '公告現值': 10, '次序': 6,
    '登記日期': 11, '登記原因': 10, '發生日期': 11,
    '所有權人': 10, '統一編號（遮罩）': 14, '統一編號（完整）': 14,
    '郵遞區號': 8, '住址': 32, '已售出': 6,
    '分母': 6, '分子': 6, '土地總坪數': 10, '權利範圍': 12,
    '備註': 28, '電話': 18,
    # 實價欄（地號已正規化，不再有標準地號欄）
    '實價命中': 8, '實價日期': 11, '實價總價(萬)': 12, '建議動作': 22,
    '小段': 10,
}

# 日期欄（置中）
DATE_COLS  = {'更新日期', '登記日期', '發生日期', '實價日期'}
# 數字欄（千分位）
NUM_COLS   = {'公告現值', '實價總價(萬)', '土地總坪數'}
# 文字格式欄（強制文字，避免數字吃掉前導零）
TEXT_COLS  = {'電話', '統一編號（遮罩）', '統一編號（完整）', '郵遞區號'}


def read_col_dims(path: Path) -> 'dict[str, tuple]':
    """
    讀取現有 Excel 的欄寬與隱藏狀態，回傳 {欄名: (width, hidden)}。
    供 apply_format 保留使用者手動調整。
    """
    if not path or not path.exists():
        return {}
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        from openpyxl.utils import get_column_letter as _gcl
        dims = {}
        for ci, h in enumerate(headers, 1):
            if not h:
                continue
            letter = _gcl(ci)
            cd = ws.column_dimensions.get(letter)
            if cd:
                dims[str(h)] = (cd.width, cd.hidden)
        wb.close()
        return dims
    except Exception:
        return {}


def apply_format(wb, headers: list, is_report: bool = False,
                 preserved_dims: 'dict[str, tuple] | None' = None):
    """
    套用格式到 workbook（in-place）。
    preserved_dims: {欄名: (width, hidden)} — 已存在欄保留使用者調整，新欄套預設值。
    is_report=True 時對「實價命中」列套用紅底，「建議動作」欄套用黃底。
    主清冊 148k 列只做標題列與欄寬，不做逐列填色（效能考量）。
    """
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter

    ws = wb.active
    if preserved_dims is None:
        preserved_dims = {}

    # ── 標題列：深藍底白字粗體置中
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill  = hdr_fill
        cell.font  = hdr_font
        cell.alignment = hdr_align

    # ── 欄寬 + 隱藏：現有欄保留使用者設定，新欄套預設值
    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        cd = ws.column_dimensions[col_letter]
        if h in preserved_dims:
            saved_width, saved_hidden = preserved_dims[h]
            if saved_width:
                cd.width = saved_width
            cd.hidden = saved_hidden
        else:
            cd.width = COL_WIDTHS.get(h, 10)

    # ── 資料欄預設字體（column-level 無效，只設列高預設）
    # 對資料列逐列設字體過慢（148k），改為只在報表（1566列）做逐列格式
    body_font   = Font(name='微軟正黑體', size=10)
    center_align = Alignment(horizontal='center')
    text_fmt    = numbers.FORMAT_TEXT

    if is_report:
        hit_col_idx  = next((i+1 for i,h in enumerate(headers) if h == '實價命中'),  None)
        act_col_idx  = next((i+1 for i,h in enumerate(headers) if h == '建議動作'), None)
        red_fill  = PatternFill('solid', fgColor='C00000')
        red_font  = Font(name='微軟正黑體', size=10, bold=True, color='FFFFFF')
        yel_fill  = PatternFill('solid', fgColor='FFEB9C')
        yel_font  = Font(name='微軟正黑體', size=10, color='000000')

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                h = headers[cell.column - 1] if cell.column <= len(headers) else ''
                cell.font = body_font
                if h in DATE_COLS:
                    cell.alignment = center_align
                if h in NUM_COLS and cell.value is not None:
                    try:
                        cell.value        = float(str(cell.value).replace(',', ''))
                        cell.number_format = '#,##0'
                    except (ValueError, TypeError):
                        pass
                if h == '持分坪數' and cell.value is not None:
                    try:
                        cell.value        = float(str(cell.value).replace(',', ''))
                        cell.number_format = '#,##0.0000'
                        cell.alignment    = Alignment(horizontal='right')
                    except (ValueError, TypeError):
                        pass
                if h in TEXT_COLS:
                    cell.number_format = text_fmt
            # 實價命中欄紅底白字
            if hit_col_idx:
                c = ws.cell(row=row[0].row, column=hit_col_idx)
                c.fill = red_fill; c.font = red_font
            # 建議動作欄黃底
            if act_col_idx:
                c = ws.cell(row=row[0].row, column=act_col_idx)
                c.fill = yel_fill; c.font = yel_font

    else:
        # 主清冊：只做欄格式（number_format），不逐列填色
        for ci, h in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            if h in NUM_COLS:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        cell.number_format = '#,##0'
            if h == '持分坪數':
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        cell.number_format = '#,##0.0000'
                        cell.alignment = Alignment(horizontal='right')
            if h in TEXT_COLS:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        cell.number_format = text_fmt
            if h in DATE_COLS:
                for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')

    # ── 凍結 A2 + 自動篩選
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

# ── 主流程 ────────────────────────────────────────────────────────

def run(xlsx_path: Path, dry_run: bool):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
    except ImportError:
        print('請安裝 openpyxl：pip install openpyxl')
        sys.exit(1)

    ts    = datetime.now().strftime('%Y%m%d_%H%M')
    mode  = 'DRY-RUN' if dry_run else '正式執行'

    # ── 確保資料夾存在
    for d in [INPUT_DIR, OUTPUT_DIR, BACKUP_DIR, LOGS_DIR, ARCHIVE_DIR, LATEST_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    # ── log 初始化
    log_path = LOGS_DIR / f'excel_realprice_update_{ts}.log'
    log_lines: list[str] = []

    def log(msg: str):
        print(msg)
        log_lines.append(msg)

    log(f'[ {mode} ] update_excel_realprice.py')
    log(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    log(f'來源：{xlsx_path}')
    log(f'LI_DB：{LI_DB}')
    log(f'OUTPUT_DIR：{OUTPUT_DIR}')
    log(f'BACKUP_DIR：{BACKUP_DIR}')
    log(f'LOGS_DIR  ：{LOGS_DIR}')

    # ── 載入實價資料
    log('\n[載入] 實價登錄資料 ...')
    tx_idx = load_transactions()
    log(f'  {sum(len(v) for v in tx_idx.values()):,} 筆（2025起，已建索引）')

    # ── 讀取 Excel（read_only 只讀取值，寫入時重開）
    log('\n[讀取] Excel ...')
    wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws_ro = wb_ro.active

    all_rows = list(ws_ro.iter_rows(values_only=True))
    wb_ro.close()

    if not all_rows:
        log('Excel 是空的，中止。')
        sys.exit(1)

    headers = list(all_rows[0])
    data_rows = all_rows[1:]
    total_rows = len(data_rows)
    log(f'  總列數（不含標題）：{total_rows:,}')

    # ── 偵測欄位
    col_map = detect_columns(headers)
    missing = [k for k in ['縣市', '地區', '地段', '地號'] if k not in col_map]
    if missing:
        log(f'\n[錯誤] 找不到必要欄位：{missing}')
        log(f'  目前辨識到的標題：{[h for h in headers if h]}')
        sys.exit(1)

    found_optional = {k: col_map[k] for k in col_map if k not in ['縣市', '地區', '地段', '地號']}
    log(f'  必要欄位：OK（縣市@{col_map["縣市"]+1} 地區@{col_map["地區"]+1} 地段@{col_map["地段"]+1} 地號@{col_map["地號"]+1}）')
    log(f'  選用欄位：{", ".join(f"{k}@{v+1}" for k, v in found_optional.items()) or "（無）"}')

    # 找出已存在輔助欄位位置
    existing_new_cols = detect_new_col_positions(headers)
    log(f'  已有輔助欄位：{list(existing_new_cols.keys()) or "（無，將新增至尾端）"}')

    # ── 比對（normalize 前：用舊版 4-zero padding 比較，計算基準命中數）
    log('\n[比對] ...')

    def _norm_land_no_old(raw) -> str:
        """舊版格式（補零 0001-0000），用於統計 normalize 前命中數"""
        if not raw:
            return ''
        s = full_to_half(str(raw).strip())
        s = re.sub(r'地號', '', s)
        s = re.sub(r'之', '-', s)
        s = re.sub(r'\s+', '', s)
        s = re.sub(r'[^\d\-]', '', s)
        if not s:
            return ''
        parts = s.split('-')
        try:
            main = int(parts[0]) if parts[0] else 0
            sub  = int(parts[1]) if len(parts) > 1 and parts[1] else 0
            return f'{main:04d}-{sub:04d}'
        except ValueError:
            return s

    hit_count        = 0
    multi_batch      = 0
    unresolvable     = 0
    skipped_no_key   = 0
    newly_hit        = 0   # normalize 後新增命中（原本沒有）
    filtered_by_date = 0   # 登記日期 >= 實價日期，已過濾
    within_90_count  = 0   # 差距 90 天內，人工確認

    hit_results: dict[int, dict] = {}

    # 建立舊版索引（只用於比較）
    tx_idx_old: dict = defaultdict(list)
    con_tmp = sqlite3.connect(LI_DB, timeout=10)
    con_tmp.row_factory = sqlite3.Row
    rows_tmp = [dict(r) for r in con_tmp.execute("""
        SELECT city, district, section_name, land_number,
               trade_date, total_price_wan, unique_key
        FROM land_transactions
        WHERE trade_date >= '2025-01-01'
          AND (target_category LIKE '%土地%' OR target_category IS NULL)
          AND land_number IS NOT NULL AND land_number != ''
    """).fetchall()]
    con_tmp.close()
    for r in rows_tmp:
        sec = norm_section(r['section_name'] or '')
        lno = _norm_land_no_old(r['land_number'])
        if sec and lno:
            tx_idx_old[(r['city'], r['district'], sec, lno)].append(r)

    for i, row in enumerate(data_rows):
        city     = row[col_map['縣市']]    if '縣市'   in col_map else None
        district = row[col_map['地區']]    if '地區'   in col_map else None
        section  = row[col_map['地段']]    if '地段'   in col_map else None
        land_no  = row[col_map['地號']]    if '地號'   in col_map else None
        is_sold  = row[col_map['已售出']]  if '已售出' in col_map else None

        if not city or not district or not section:
            skipped_no_key += 1
            continue

        if not land_no or not norm_land_no(land_no):
            unresolvable += 1
            continue

        sold_flag = str(is_sold).strip() in ('是', '1', 'True', 'Y', 'y') if is_sold else False

        # 新版 normalize 比對
        hits = find_hits(str(city).strip(), str(district).strip(),
                         str(section).strip(), str(land_no).strip(), tx_idx)

        if not hits:
            continue

        # 判斷是否為 normalize 後新增命中（舊版不命中）
        city_raw  = str(city).strip()
        dist_raw  = str(district).strip()
        sec_raw   = norm_section(str(section).strip())
        lno_old   = _norm_land_no_old(str(land_no).strip())
        old_hit = any(
            k[0] == city_raw and k[1] == dist_raw and k[3] == lno_old and
            (sec_raw in k[2] or k[2] in sec_raw or sec_raw == k[2])
            for k in tx_idx_old
        )
        if not old_hit:
            newly_hit += 1

        # ── 日期過濾：比較主清冊登記日期 vs 實價最新日期
        reg_raw = row[col_map['登記日期']] if '登記日期' in col_map else None
        reg_date = parse_date(reg_raw)
        trade_dates = [parse_date(h['trade_date']) for h in hits if parse_date(h['trade_date'])]
        latest_trade = max(trade_dates) if trade_dates else None

        if latest_trade and reg_date and reg_date >= latest_trade:
            # 主清冊登記日期 >= 實價日期 → 不列入報表
            filtered_by_date += 1
            continue

        summary = build_hit_summary(hits, sold_flag)

        # 90 天內：建議動作改為人工確認
        if latest_trade and reg_date and (latest_trade - reg_date).days <= 90:
            summary['建議動作'] = '近期已更新，請人工確認'
            within_90_count += 1

        hit_results[i] = summary
        hit_count += 1
        if '同批多地號' in summary.get('建議動作', ''):
            multi_batch += 1

    log(f'  可辨識地號列數（normalize 後）：{total_rows - unresolvable - skipped_no_key:,}')
    log(f'  命中筆數（過濾前）             ：{hit_count + filtered_by_date:,}')
    log(f'  過濾：登記日期 >= 實價日期     ：{filtered_by_date:,} 筆')
    log(f'  命中筆數（過濾後）             ：{hit_count:,}')
    log(f'  其中 90 天內人工確認           ：{within_90_count:,} 筆')
    log(f'  normalize 後新增命中           ：{newly_hit:,}')
    log(f'  同批多地號筆數                 ：{multi_batch:,}')
    log(f'  地號無法解析                   ：{unresolvable:,}')
    log(f'  缺必要欄位                     ：{skipped_no_key:,}')

    if dry_run:
        orig_col_count   = len([h for h in headers if h])
        report_col_count = orig_col_count + len(NEW_COLS)
        city_i = col_map['縣市']; dist_i = col_map['地區']
        sec_i  = col_map['地段']; lno_i  = col_map['地號']

        # 預覽報表欄位順序（dry-run 用）
        orig_hdrs_dry = [h for h in headers if h][:orig_col_count]
        insert_after_dry = lno_i + 1
        preview_hdrs = (orig_hdrs_dry[:insert_after_dry]
                        + NEW_COLS
                        + orig_hdrs_dry[insert_after_dry:])

        log(f'\n[主清冊] 欄位數：{orig_col_count}（不新增實價欄位，保持乾淨）')
        log(f'[報表]   欄位數：{report_col_count}（主清冊 {orig_col_count} 欄 + 實價 {len(NEW_COLS)} 欄）')
        log(f'[報表]   命中列：{hit_count:,}（只輸出命中列，不含全部 {total_rows:,} 列）')
        log(f'\n[欄位順序] 報表前 15 欄：')
        for idx, h in enumerate(preview_hdrs[:15], 1):
            marker = ' ← 實價欄' if h in NEW_COLS else ''
            log(f'  {idx:2d}. {h}{marker}')

        # 排序後前 20 筆預覽（主清冊）
        sorted_preview = sorted(data_rows[:500],
                                key=lambda r: sort_key(r, city_i, dist_i, sec_i, lno_i))
        log(f'\n[排序預覽] 主清冊排序後前 20 筆（縣市→地區→地段→標準地號）：')
        for i, row in enumerate(sorted_preview[:20], 1):
            lno_std = norm_land_no(str(row[lno_i] or ''))
            log(f'  {i:2d}. {row[city_i]} {row[dist_i]} {row[sec_i]}  地號={row[lno_i]} → {lno_std}')

        log(f'\n[格式] 套用：微軟正黑體 10、深藍標題列、凍結A2、自動篩選')
        log(f'[格式] 日期欄置中：{sorted(DATE_COLS)}')
        log(f'[格式] 千分位欄：{sorted(NUM_COLS)}')
        log(f'[格式] 文字格式欄：{sorted(TEXT_COLS)}')
        log(f'[格式] 實價報表：實價命中→紅底白字、建議動作→黃底')

        log('\n[dry-run] 不寫入檔案。確認無誤後執行正式輸出：')
        log(f'  python3 scripts/update_excel_realprice.py')
        _write_log(log_path, log_lines)
        return

    if hit_count == 0:
        log('\n無命中，不輸出檔案。')
        _write_log(log_path, log_lines)
        return

    # ── 備份原始 Excel（正式執行才備份）
    backup_path = BACKUP_DIR / f'{xlsx_path.stem}_backup_{ts}{xlsx_path.suffix}'
    shutil.copy2(xlsx_path, backup_path)
    log(f'\n[備份] {backup_path}')

    # ── 欄位資訊
    orig_col_count = len([h for h in headers if h])
    orig_headers_list = [h for h in headers if h][:orig_col_count]

    # 排序 key 用欄位索引（0-based）
    city_i = col_map['縣市']; dist_i = col_map['地區']
    sec_i  = col_map['地段']; lno_i  = col_map['地號']

    # 5 個實價欄插在「地號」欄之後
    insert_after = lno_i + 1
    all_report_headers = (orig_headers_list[:insert_after]
                          + NEW_COLS
                          + orig_headers_list[insert_after:])

    # ── 產生實價提醒報表（排序 + 格式）
    log('\n[輸出] 產生實價提醒報表 ...')

    # 命中列 + summary，依排序 key 排序
    hit_pairs = [(data_rows[i], summary) for i, summary in hit_results.items()]
    hit_pairs.sort(key=lambda x: sort_key(x[0], city_i, dist_i, sec_i, lno_i))

    wb_rpt = openpyxl.Workbook()
    ws_rpt = wb_rpt.active
    ws_rpt.title = '實價提醒'

    # 標題列
    for ci, h in enumerate(all_report_headers, 1):
        ws_rpt.cell(row=1, column=ci, value=h)

    # 資料列（5 個實價欄插在地號後）
    for out_row, (orig_row, summary) in enumerate(hit_pairs, 2):
        # 地號前（含地號）
        for ci in range(insert_after):
            ws_rpt.cell(row=out_row, column=ci + 1, value=orig_row[ci])
        # 插入 5 個實價欄
        for offset, col_name in enumerate(NEW_COLS):
            ws_rpt.cell(row=out_row, column=insert_after + 1 + offset,
                        value=summary.get(col_name, ''))
        # 地號之後的原始欄位
        for ci in range(insert_after, orig_col_count):
            ws_rpt.cell(row=out_row, column=ci + 1 + len(NEW_COLS), value=orig_row[ci])

    # 格式（保留使用者手動調整的欄寬/隱藏）
    latest_report_path = LATEST_DIR / '實價提醒報表_最新完成版.xlsx'
    preserved_rpt = read_col_dims(latest_report_path)
    apply_format(wb_rpt, all_report_headers, is_report=True, preserved_dims=preserved_rpt)

    out_path = OUTPUT_DIR / f'實價提醒報表_{ts}.xlsx'
    wb_rpt.save(out_path)
    wb_rpt.close()
    written = len(hit_pairs)
    log(f'[輸出] 已儲存（含排序+格式）：{out_path}')
    log(f'[輸出] 報表欄位數：{len(all_report_headers)}（主清冊 {orig_col_count} + 實價 {len(NEW_COLS)}）')

    # ── 主清冊乾淨版 → 排序+格式後存到最新完成版
    log('\n[輸出] 產生排序版主清冊 ...')
    orig_headers_clean = [h for h in headers if h][:orig_col_count]

    # 讀取全部資料列並排序
    all_data = sorted(data_rows, key=lambda r: sort_key(r, city_i, dist_i, sec_i, lno_i))

    wb_mst = openpyxl.Workbook()
    ws_mst = wb_mst.active
    ws_mst.title = '土地主清冊'

    for ci, h in enumerate(orig_headers_clean, 1):
        ws_mst.cell(row=1, column=ci, value=h)

    for ri, row in enumerate(all_data, 2):
        for ci in range(orig_col_count):
            ws_mst.cell(row=ri, column=ci + 1, value=row[ci])

    # 格式（保留使用者手動調整的欄寬/隱藏）
    preserved_mst = read_col_dims(LATEST_FILE)
    apply_format(wb_mst, orig_headers_clean, is_report=False, preserved_dims=preserved_mst)

    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    if LATEST_FILE.exists():
        LATEST_FILE.unlink()
    wb_mst.save(LATEST_FILE)
    wb_mst.close()
    log(f'[最新版] 主清冊（排序+格式）：{LATEST_FILE}')

    # ── 實價報表 → 最新完成版
    latest_report = LATEST_DIR / '實價提醒報表_最新完成版.xlsx'
    if latest_report.exists():
        latest_report.unlink()
    shutil.copy2(out_path, latest_report)
    log(f'[最新版] 實價報表         ：{latest_report}')

    # ── output/ 清理：實價提醒報表只保留最近 2 份
    report_files = sorted(OUTPUT_DIR.glob('實價提醒報表_*.xlsx'),
                          key=lambda p: p.stat().st_mtime, reverse=True)
    for old in report_files[2:]:
        old.unlink()
        log(f'[清理] 已刪除：{old.name}')
    log(f'[清理] 實價報表 output 保留 {min(len(report_files), 2)} 份')

    log(f'\n{"─"*50}')
    log(f'完成')
    log(f'  命中列數       ：{written:,}')
    log(f'  報表欄位數     ：{len(all_report_headers)}（主清冊 {orig_col_count} + 實價 {len(NEW_COLS)}）')
    log(f'  報表路徑       ：{out_path}')
    log(f'  最新完成版主清冊：{LATEST_FILE}')
    log(f'  最新完成版報表 ：{latest_report}')
    log(f'  備份路徑       ：{backup_path}')
    log(f'  原始 input     ：未修改')
    _write_log(log_path, log_lines)


def main():
    parser = argparse.ArgumentParser(description='Excel 實價登錄比對更新')
    parser.add_argument('--dry-run', action='store_true', help='只統計，不寫檔')
    parser.add_argument('--input',   type=Path, default=DEFAULT_XLSX, help='來源 Excel 路徑')
    args = parser.parse_args()

    if not args.input.exists():
        print(f'[錯誤] 找不到 Excel：{args.input}')
        sys.exit(1)

    run(args.input, args.dry_run)


if __name__ == '__main__':
    main()
