#!/usr/bin/env python3
"""
check_land_master_quality.py  v4.2
主清冊資料品質檢查報表
輸入：土地主清冊_正式版_20260522.xlsx
輸出：資料品質檢查報表_YYYYMMDD_HHMMSS.xlsx
原則：只讀不寫，不修改主清冊、不動 SQLite
"""
import re
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

SRC = '/Users/xiaomingyang/Desktop/excel土地資料維護/土地主清冊_正式版_20260522.xlsx'
REPORT_DIR = '/Users/xiaomingyang/Desktop/excel土地資料維護'

# ── 台灣郵遞區號對照表（縣市+鄉鎮市區 → 郵遞區號）──────────────────────────
# 只收錄常見的；match 時先找 3+3（縣市＋鄉鎮），找不到再退回 縣市
TW_POSTAL = {
    # 台北市
    '台北市中正區': '100', '台北市大同區': '103', '台北市中山區': '104',
    '台北市松山區': '105', '台北市大安區': '106', '台北市萬華區': '108',
    '台北市信義區': '110', '台北市士林區': '111', '台北市北投區': '112',
    '台北市內湖區': '114', '台北市南港區': '115', '台北市文山區': '116',
    '臺北市中正區': '100', '臺北市大同區': '103', '臺北市中山區': '104',
    '臺北市松山區': '105', '臺北市大安區': '106', '臺北市萬華區': '108',
    '臺北市信義區': '110', '臺北市士林區': '111', '臺北市北投區': '112',
    '臺北市內湖區': '114', '臺北市南港區': '115', '臺北市文山區': '116',
    # 新北市
    '新北市板橋區': '220', '新北市三重區': '241', '新北市中和區': '235',
    '新北市永和區': '234', '新北市新莊區': '242', '新北市新店區': '231',
    '新北市樹林區': '238', '新北市鶯歌區': '239', '新北市三峽區': '237',
    '新北市淡水區': '251', '新北市汐止區': '221', '新北市瑞芳區': '224',
    '新北市土城區': '236', '新北市蘆洲區': '247', '新北市五股區': '248',
    '新北市泰山區': '243', '新北市林口區': '244', '新北市深坑區': '222',
    '新北市石碇區': '223', '新北市坪林區': '232', '新北市三芝區': '252',
    '新北市石門區': '253', '新北市八里區': '249', '新北市平溪區': '226',
    '新北市雙溪區': '227', '新北市貢寮區': '228', '新北市金山區': '208',
    '新北市萬里區': '207', '新北市烏來區': '233',
    # 桃園市
    '桃園市桃園區': '330', '桃園市中壢區': '320', '桃園市大溪區': '335',
    '桃園市楊梅區': '326', '桃園市蘆竹區': '338', '桃園市大園區': '337',
    '桃園市龜山區': '333', '桃園市八德區': '334', '桃園市龍潭區': '325',
    '桃園市平鎮區': '324', '桃園市新屋區': '327', '桃園市觀音區': '328',
    '桃園市復興區': '336',
    # 新竹市
    '新竹市東區': '300', '新竹市北區': '300', '新竹市香山區': '300',
    # 新竹縣
    '新竹縣竹北市': '302', '新竹縣湖口鄉': '303', '新竹縣新豐鄉': '304',
    '新竹縣新埔鎮': '305', '新竹縣關西鎮': '306', '新竹縣芎林鄉': '307',
    '新竹縣寶山鄉': '308', '新竹縣竹東鎮': '310', '新竹縣五峰鄉': '311',
    '新竹縣橫山鄉': '312', '新竹縣尖石鄉': '313', '新竹縣北埔鄉': '314',
    '新竹縣峨眉鄉': '315',
    # 苗栗縣
    '苗栗縣苗栗市': '360', '苗栗縣頭份市': '351', '苗栗縣竹南鎮': '350',
    '苗栗縣後龍鎮': '356', '苗栗縣通霄鎮': '357', '苗栗縣苑裡鎮': '358',
    '苗栗縣造橋鄉': '361', '苗栗縣頭屋鄉': '362', '苗栗縣公館鄉': '363',
    '苗栗縣大湖鄉': '364', '苗栗縣泰安鄉': '365', '苗栗縣銅鑼鄉': '366',
    '苗栗縣三義鄉': '367', '苗栗縣西湖鄉': '368', '苗栗縣三灣鄉': '352',
    '苗栗縣南庄鄉': '353', '苗栗縣獅潭鄉': '354', '苗栗縣卓蘭鎮': '369',
    # 台中市
    '台中市中區': '400', '台中市東區': '401', '台中市南區': '402',
    '台中市西區': '403', '台中市北區': '404', '台中市北屯區': '406',
    '台中市西屯區': '407', '台中市南屯區': '408', '台中市太平區': '411',
    '台中市大里區': '412', '台中市霧峰區': '413', '台中市烏日區': '414',
    '台中市豐原區': '420', '台中市后里區': '421', '台中市石岡區': '422',
    '台中市東勢區': '423', '台中市和平區': '424', '台中市新社區': '426',
    '台中市潭子區': '427', '台中市大雅區': '428', '台中市神岡區': '429',
    '台中市大肚區': '432', '台中市沙鹿區': '433', '台中市龍井區': '434',
    '台中市梧棲區': '435', '台中市清水區': '436', '台中市大甲區': '437',
    '台中市外埔區': '438', '台中市大安區': '439',
    '臺中市中區': '400', '臺中市東區': '401', '臺中市南區': '402',
    '臺中市西區': '403', '臺中市北區': '404', '臺中市北屯區': '406',
    '臺中市西屯區': '407', '臺中市南屯區': '408',
    # 彰化縣
    '彰化縣彰化市': '500', '彰化縣芬園鄉': '502', '彰化縣花壇鄉': '503',
    '彰化縣秀水鄉': '504', '彰化縣鹿港鎮': '505', '彰化縣福興鄉': '506',
    '彰化縣線西鄉': '507', '彰化縣和美鎮': '508', '彰化縣伸港鄉': '509',
    '彰化縣員林市': '510', '彰化縣社頭鄉': '511', '彰化縣永靖鄉': '512',
    '彰化縣埔心鄉': '513', '彰化縣溪湖鎮': '514', '彰化縣大村鄉': '515',
    '彰化縣埔鹽鄉': '516', '彰化縣田中鎮': '520', '彰化縣北斗鎮': '521',
    '彰化縣田尾鄉': '522', '彰化縣埤頭鄉': '523', '彰化縣溪州鄉': '524',
    '彰化縣竹塘鄉': '525', '彰化縣二林鎮': '526', '彰化縣大城鄉': '527',
    '彰化縣芳苑鄉': '528', '彰化縣二水鄉': '530',
    # 南投縣
    '南投縣南投市': '540', '南投縣中寮鄉': '541', '南投縣草屯鎮': '542',
    '南投縣國姓鄉': '544', '南投縣埔里鎮': '545', '南投縣仁愛鄉': '546',
    '南投縣名間鄉': '551', '南投縣集集鎮': '552', '南投縣水里鄉': '553',
    '南投縣魚池鄉': '555', '南投縣信義鄉': '556', '南投縣竹山鎮': '557',
    '南投縣鹿谷鄉': '558',
    # 雲林縣
    '雲林縣斗南鎮': '630', '雲林縣大埤鄉': '631', '雲林縣虎尾鎮': '632',
    '雲林縣土庫鎮': '633', '雲林縣褒忠鄉': '634', '雲林縣東勢鄉': '635',
    '雲林縣台西鄉': '636', '雲林縣崙背鄉': '637', '雲林縣麥寮鄉': '638',
    '雲林縣斗六市': '640', '雲林縣林內鄉': '643', '雲林縣古坑鄉': '646',
    '雲林縣莿桐鄉': '647', '雲林縣西螺鎮': '648', '雲林縣二崙鄉': '649',
    '雲林縣北港鎮': '651', '雲林縣水林鄉': '652', '雲林縣口湖鄉': '653',
    '雲林縣四湖鄉': '654', '雲林縣元長鄉': '655',
    # 嘉義市
    '嘉義市東區': '600', '嘉義市西區': '600',
    # 嘉義縣
    '嘉義縣番路鄉': '602', '嘉義縣梅山鄉': '603', '嘉義縣竹崎鄉': '604',
    '嘉義縣阿里山鄉': '605', '嘉義縣中埔鄉': '606', '嘉義縣大埔鄉': '607',
    '嘉義縣水上鄉': '608', '嘉義縣鹿草鄉': '611', '嘉義縣太保市': '612',
    '嘉義縣朴子市': '613', '嘉義縣東石鄉': '614', '嘉義縣六腳鄉': '615',
    '嘉義縣新港鄉': '616', '嘉義縣民雄鄉': '621', '嘉義縣大林鎮': '622',
    '嘉義縣溪口鄉': '623', '嘉義縣義竹鄉': '624', '嘉義縣布袋鎮': '625',
    # 台南市
    '台南市中西區': '700', '台南市東區': '701', '台南市南區': '702',
    '台南市北區': '704', '台南市安平區': '708', '台南市安南區': '709',
    '台南市永康區': '710', '台南市歸仁區': '711', '台南市新化區': '712',
    '台南市左鎮區': '713', '台南市玉井區': '714', '台南市楠西區': '715',
    '台南市南化區': '716', '台南市仁德區': '717', '台南市關廟區': '718',
    '台南市龍崎區': '719', '台南市官田區': '720', '台南市麻豆區': '721',
    '台南市佳里區': '722', '台南市西港區': '723', '台南市七股區': '724',
    '台南市將軍區': '725', '台南市學甲區': '726', '台南市北門區': '727',
    '台南市新營區': '730', '台南市後壁區': '731', '台南市白河區': '732',
    '台南市東山區': '733', '台南市六甲區': '734', '台南市下營區': '735',
    '台南市柳營區': '736', '台南市鹽水區': '737', '台南市善化區': '741',
    '台南市大內區': '742', '台南市山上區': '743', '台南市新市區': '744',
    '台南市安定區': '745',
    '臺南市中西區': '700', '臺南市東區': '701', '臺南市南區': '702',
    '臺南市北區': '704', '臺南市安平區': '708', '臺南市安南區': '709',
    # 高雄市
    '高雄市楠梓區': '811', '高雄市左營區': '813', '高雄市鼓山區': '804',
    '高雄市三民區': '807', '高雄市鹽埕區': '803', '高雄市前金區': '801',
    '高雄市苓雅區': '802', '高雄市新興區': '800', '高雄市前鎮區': '806',
    '高雄市旗津區': '805', '高雄市小港區': '812', '高雄市鳳山區': '830',
    '高雄市林園區': '832', '高雄市大寮區': '831', '高雄市大樹區': '840',
    '高雄市大社區': '815', '高雄市仁武區': '814', '高雄市鳥松區': '833',
    '高雄市岡山區': '820', '高雄市橋頭區': '825', '高雄市燕巢區': '824',
    '高雄市田寮區': '823', '高雄市阿蓮區': '822', '高雄市路竹區': '821',
    '高雄市湖內區': '829', '高雄市茄萣區': '852', '高雄市永安區': '828',
    '高雄市彌陀區': '827', '高雄市梓官區': '826', '高雄市旗山區': '842',
    '高雄市美濃區': '843', '高雄市六龜區': '844', '高雄市甲仙區': '847',
    '高雄市杉林區': '848', '高雄市內門區': '845', '高雄市茂林區': '851',
    '高雄市桃源區': '846', '高雄市那瑪夏區': '849',
    # 屏東縣
    '屏東縣屏東市': '900', '屏東縣三地門鄉': '901', '屏東縣霧台鄉': '902',
    '屏東縣瑪家鄉': '903', '屏東縣九如鄉': '904', '屏東縣里港鄉': '905',
    '屏東縣高樹鄉': '906', '屏東縣鹽埔鄉': '907', '屏東縣長治鄉': '908',
    '屏東縣麟洛鄉': '909', '屏東縣竹田鄉': '911', '屏東縣內埔鄉': '912',
    '屏東縣萬丹鄉': '913', '屏東縣潮州鎮': '920', '屏東縣泰武鄉': '921',
    '屏東縣來義鄉': '922', '屏東縣萬巒鄉': '923', '屏東縣崁頂鄉': '924',
    '屏東縣新埤鄉': '925', '屏東縣南州鄉': '926', '屏東縣林邊鄉': '927',
    '屏東縣東港鎮': '928', '屏東縣琉球鄉': '929', '屏東縣車城鄉': '944',
    '屏東縣滿州鄉': '947', '屏東縣枋寮鄉': '940', '屏東縣新園鄉': '932',
    '屏東縣枋山鄉': '941', '屏東縣春日鄉': '942', '屏東縣獅子鄉': '943',
    '屏東縣牡丹鄉': '945', '屏東縣恆春鎮': '946',
    # 台東縣
    '台東縣台東市': '950', '台東縣綠島鄉': '951', '台東縣蘭嶼鄉': '952',
    '台東縣延平鄉': '953', '台東縣卑南鄉': '954', '台東縣鹿野鄉': '955',
    '台東縣關山鎮': '956', '台東縣海端鄉': '957', '台東縣池上鄉': '958',
    '台東縣東河鄉': '959', '台東縣成功鎮': '961', '台東縣長濱鄉': '962',
    '台東縣太麻里鄉': '963', '台東縣金峰鄉': '964', '台東縣大武鄉': '965',
    '台東縣達仁鄉': '966',
    # 花蓮縣
    '花蓮縣花蓮市': '970', '花蓮縣新城鄉': '971', '花蓮縣秀林鄉': '972',
    '花蓮縣吉安鄉': '973', '花蓮縣壽豐鄉': '974', '花蓮縣鳳林鎮': '975',
    '花蓮縣光復鄉': '976', '花蓮縣豐濱鄉': '977', '花蓮縣瑞穗鄉': '978',
    '花蓮縣萬榮鄉': '979', '花蓮縣玉里鎮': '981', '花蓮縣卓溪鄉': '982',
    '花蓮縣富里鄉': '983',
    # 宜蘭縣
    '宜蘭縣宜蘭市': '260', '宜蘭縣頭城鎮': '261', '宜蘭縣礁溪鄉': '262',
    '宜蘭縣壯圍鄉': '263', '宜蘭縣員山鄉': '264', '宜蘭縣羅東鎮': '265',
    '宜蘭縣三星鄉': '266', '宜蘭縣大同鄉': '267', '宜蘭縣五結鄉': '268',
    '宜蘭縣冬山鄉': '269', '宜蘭縣蘇澳鎮': '270', '宜蘭縣南澳鄉': '272',
    # 基隆市
    '基隆市仁愛區': '200', '基隆市信義區': '201', '基隆市中正區': '202',
    '基隆市中山區': '203', '基隆市安樂區': '204', '基隆市暖暖區': '205',
    '基隆市七堵區': '206',
}

def lookup_postal(address: str) -> tuple[str | None, str]:
    """從地址推算郵遞區號。回傳 (建議郵遞區號 or None, 信心)"""
    if not address or not isinstance(address, str):
        return None, 'N/A'
    addr = address.strip()
    # 嘗試配對 縣市(2-3字) + 鄉鎮市區(2-4字)
    for key, postal in TW_POSTAL.items():
        if addr.startswith(key) or key in addr[:10]:
            return postal, '高'
    # 縮短到縣市層級
    for key, postal in TW_POSTAL.items():
        city = key[:3]  # 取前3字（桃園市、台北市…）
        if addr.startswith(city):
            return postal, '低'
    return None, 'N/A'

# 姓名異常 pattern
RE_NUM_ONLY   = re.compile(r'^\d+$')
RE_PHONE_PAT  = re.compile(r'^[\d\-\(\)\s]{6,}$')
RE_SYMBOL_ONLY= re.compile(r'^[-\-—–·\s/\\]+$')
NULL_NAMES    = {'', '-', '—', '無', '不詳', '查無', 'N/A', 'n/a', '未知', '不明', '?', '？'}
RE_STATS_ID   = re.compile(r'^\d{8}$')   # 8位統編格式
RE_ADDRESS_PAT= re.compile(r'[市縣].*[路街巷]')

# 電話 pattern
RE_VALID_PHONE = re.compile(r'(\d[\d\-\(\)\s]{4,})')

def check_owner(v) -> tuple[str, str, str]:
    """(問題類型, 原始值, 建議值)"""
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return 'A1_姓名空白', str(v) if v is not None else '', ''
    s = str(v).strip()
    if s in NULL_NAMES or RE_SYMBOL_ONLY.match(s):
        return 'A2_姓名為符號或占位符', s, ''
    if RE_NUM_ONLY.match(s):
        return 'A3_姓名為純數字', s, ''
    if RE_PHONE_PAT.match(s):
        return 'A4_姓名疑似電話格式', s, ''
    if RE_STATS_ID.match(s):
        return 'A5_姓名疑似統一編號', s, ''
    if RE_ADDRESS_PAT.search(s):
        return 'A6_姓名疑似地址', s, ''
    if len(s) <= 1:
        return 'A7_姓名長度過短', s, ''
    if len(s) >= 15:
        return 'A8_姓名長度過長', s, ''
    # 含多餘空格
    if s != str(v).strip() or '  ' in s:
        return 'A9_姓名含多餘空格', str(v), s
    return '', '', ''

def check_postal(postal_raw, address) -> tuple[str, str, str, str, str]:
    """(問題類型, 原始值, 建議值, 信心, 建議動作)"""
    postal_str = str(int(postal_raw)).zfill(3) if isinstance(postal_raw, (int, float)) and postal_raw else (str(postal_raw).strip() if postal_raw else '')
    suggested, confidence = lookup_postal(address)

    if not postal_str or postal_str in ('', 'None', 'nan'):
        if suggested:
            action = '可自動補' if confidence == '高' else '需人工確認'
            return 'B1_郵遞區號空白', '', suggested, confidence, action
        return 'B1_郵遞區號空白', '', '', 'N/A', '需人工確認'

    if not re.match(r'^\d{3}$', postal_str):
        return 'B2_郵遞區號格式異常', postal_str, suggested or '', confidence, '需人工確認'

    if suggested and confidence == '高' and postal_str != suggested:
        return 'B3_郵遞區號與地址不符', postal_str, suggested, '高', '建議更正'

    return '', postal_str, '', '', ''

def check_address(v) -> tuple[str, str, str]:
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return 'C1_地址空白', str(v) if v is not None else '', ''
    s = str(v).strip()
    if '圖' == s or '隱匿' in s:
        return 'C2_地址隱匿', s, ''
    if len(s) < 6:
        return 'C3_地址過短', s, ''
    if not re.search(r'[市縣]', s):
        return 'C4_地址缺縣市', s, ''
    return '', '', ''

def check_phone(v) -> tuple[str, str, str]:
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return 'E1_電話空白', str(v) if v is not None else '', ''
    s = str(v).strip()
    phones = RE_VALID_PHONE.findall(s)
    if not phones:
        return 'E2_電話格式異常', s, ''
    return '', '', ''

# ── 主程式 ────────────────────────────────────────────────────────────────────
print(f'[1] 讀取主清冊...')
src_wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
src_ws = src_wb.active
all_rows = list(src_ws.iter_rows(values_only=True))
src_wb.close()
headers = list(all_rows[0])
col = {h: i for i, h in enumerate(headers)}
print(f'    {len(all_rows)-1:,} 資料列 × {len(headers)} 欄')

# 計數器
counters = {'A_姓名': 0, 'B_郵遞區號': 0, 'C_地址': 0, 'E_電話': 0,
            'B_可自動補': 0, 'B_需人工確認': 0}
issues = []

print('[2] 逐列檢查...')
REPORT_COLS = [
    'row_number', '縣市', '地區', '地段', '地號', '所有權人',
    '統一編號（遮罩）', '統一編號（完整）', '郵遞區號', '住址', '電話',
    '問題類型', '原始值', '建議值', '信心等級', '建議動作',
]

for r_idx, row in enumerate(all_rows[1:], start=2):
    owner   = row[col.get('所有權人', -1)]   if col.get('所有權人')   is not None else None
    postal  = row[col.get('郵遞區號', -1)]   if col.get('郵遞區號')   is not None else None
    address = row[col.get('住址', -1)]        if col.get('住址')        is not None else None
    phone   = row[col.get('電話', -1)]        if col.get('電話')        is not None else None
    city    = row[col.get('縣市', -1)]        if col.get('縣市')        is not None else None
    district= row[col.get('地區', -1)]        if col.get('地區')        is not None else None
    section = row[col.get('地段', -1)]        if col.get('地段')        is not None else None
    land_no = row[col.get('地號', -1)]        if col.get('地號')        is not None else None
    masked  = row[col.get('統一編號（遮罩）', -1)] if col.get('統一編號（遮罩）') is not None else None
    full_id = row[col.get('統一編號（完整）', -1)] if col.get('統一編號（完整）') is not None else None

    base = {
        'row_number': r_idx,
        '縣市': city, '地區': district, '地段': section, '地號': land_no,
        '所有權人': owner, '統一編號（遮罩）': masked, '統一編號（完整）': full_id,
        '郵遞區號': postal, '住址': address, '電話': phone,
    }

    # A. 姓名
    issue_type, orig, suggest = check_owner(owner)
    if issue_type:
        counters['A_姓名'] += 1
        issues.append({**base, '問題類型': issue_type, '原始值': orig,
                        '建議值': suggest, '信心等級': 'N/A', '建議動作': '等待新謄本/電傳'})

    # B. 郵遞區號
    issue_type, orig, suggest, conf, action = check_postal(postal, address)
    if issue_type:
        counters['B_郵遞區號'] += 1
        if action == '可自動補':
            counters['B_可自動補'] += 1
        elif '確認' in action or '更正' in action:
            counters['B_需人工確認'] += 1
        issues.append({**base, '問題類型': issue_type, '原始值': orig,
                        '建議值': suggest, '信心等級': conf, '建議動作': action})

    # C. 地址
    issue_type, orig, suggest = check_address(address)
    if issue_type:
        counters['C_地址'] += 1
        issues.append({**base, '問題類型': issue_type, '原始值': orig,
                        '建議值': suggest, '信心等級': 'N/A', '建議動作': '需人工確認'})

    # E. 電話（空白只計數不展開，格式異常才列入報表）
    issue_type, orig, suggest = check_phone(phone)
    if issue_type:
        counters['E_電話'] += 1
        if issue_type != 'E1_電話空白':   # 空白太多，只統計不展開
            issues.append({**base, '問題類型': issue_type, '原始值': orig,
                            '建議值': suggest, '信心等級': 'N/A', '建議動作': '需人工確認'})

    if r_idx % 30000 == 0:
        print(f'    {r_idx-1:,} / {len(all_rows)-1:,}')

print(f'[3] 建立報表（{len(issues):,} 筆問題）...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '資料品質報表'

# 標題
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
ws.append(REPORT_COLS)
for c in range(1, len(REPORT_COLS)+1):
    cell = ws.cell(1, c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')

# 問題類型顏色
TYPE_COLORS = {
    'A': 'FFF2CC',  # 姓名 - 黃
    'B': 'DDEEFF',  # 郵遞區號 - 藍
    'C': 'FFE0CC',  # 地址 - 橘
    'E': 'E8F5E9',  # 電話 - 綠
}

for r_idx, issue in enumerate(issues, start=2):
    row_data = [issue.get(c, '') for c in REPORT_COLS]
    ws.append(row_data)
    issue_type = str(issue.get('問題類型', ''))
    prefix = issue_type[0] if issue_type else ''
    color = TYPE_COLORS.get(prefix)
    if color:
        fill = PatternFill('solid', fgColor=color)
        for c in range(1, len(REPORT_COLS)+1):
            ws.cell(r_idx, c).fill = fill
    # 字型
    for c in range(1, len(REPORT_COLS)+1):
        ws.cell(r_idx, c).font = Font(name='微軟正黑體', size=10)

# 欄寬
widths = [8, 6, 6, 12, 10, 12, 14, 14, 8, 35, 18, 18, 20, 12, 8, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# ── 摘要頁 ──────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('摘要')
ws2.append(['類別', '問題數'])
summary_rows = [
    ('A. 所有權人問題', counters['A_姓名']),
    ('B. 郵遞區號問題', counters['B_郵遞區號']),
    ('  ├ 可自動補', counters['B_可自動補']),
    ('  └ 需人工確認', counters['B_需人工確認']),
    ('C. 地址問題', counters['C_地址']),
    ('E. 電話問題（含空白統計）', counters['E_電話']),
    ('  └ 格式異常（列入報表）', sum(1 for i in issues if i.get('問題類型','').startswith('E2'))),
    ('合計（同一筆可多問題）', len(issues)),
]
for row in summary_rows:
    ws2.append(row)
for c in range(1, 3):
    ws2.cell(1, c).font = Font(bold=True)
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 10

ts = datetime.now().strftime('%Y%m%d_%H%M%S')
out_path = os.path.join(REPORT_DIR, f'資料品質檢查報表_{ts}.xlsx')
wb.save(out_path)

size_mb = os.path.getsize(out_path) / 1024 / 1024
print(f'\n完成！')
print(f'輸出：{out_path}')
print(f'大小：{size_mb:.1f} MB')
print(f'\n── 檢查結果摘要 ──────────────────────')
print(f'A. 所有權人問題    ：{counters["A_姓名"]:,} 筆')
print(f'B. 郵遞區號問題    ：{counters["B_郵遞區號"]:,} 筆')
print(f'   ├ 可自動補      ：{counters["B_可自動補"]:,} 筆')
print(f'   └ 需人工確認    ：{counters["B_需人工確認"]:,} 筆')
print(f'C. 地址問題        ：{counters["C_地址"]:,} 筆')
e2_cnt = sum(1 for i in issues if i.get('問題類型','').startswith('E2'))
print(f'E. 電話空白        ：{counters["E_電話"]:,} 筆（僅統計，未展開）')
print(f'   格式異常列入報表：{e2_cnt:,} 筆')
print(f'────────────────────────────────────')
print(f'合計（同一筆可多問題）：{len(issues):,} 筆')
print(f'主清冊未修改。SQLite 未修改。')
