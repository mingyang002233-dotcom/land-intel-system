#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
clean_excel_master.py — 主清冊資料清洗（輸出清洗版，不動原始 input）

流程：
  1. 從 input/ 讀取原始 Excel
  2. 對縣市、行政區、地段、地號做格式統一
  3. 輸出 output/土地主清冊_清洗版_YYYYMMDD_HHMM.xlsx
  4. 輸出 logs/excel_master_clean_YYYYMMDD_HHMM.log

執行：
  python3 scripts/clean_excel_master.py --dry-run   ← 只統計，不寫檔
  python3 scripts/clean_excel_master.py             ← 正式輸出
  python3 scripts/clean_excel_master.py --input 其他.xlsx
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

BASE_DIR    = Path('/Users/xiaomingyang/Desktop/excel土地資料維護')
INPUT_DIR   = BASE_DIR / 'input'
OUTPUT_DIR  = BASE_DIR / 'output'
LOGS_DIR    = BASE_DIR / 'logs'

DEFAULT_XLSX = INPUT_DIR / '土地主清冊_正式版_20260522_郵遞區號補正版.xlsx'

# 要清洗的欄位名稱（允許別名）
COL_ALIASES = {
    '縣市': ['縣市'],
    '地區': ['地區', '行政區', '鄉鎮市區'],
    '地段': ['地段'],
    '地號': ['地號'],
}

# ── 全形→半形 ────────────────────────────────────────────────────

def full_to_half(s: str) -> str:
    result = []
    for c in s:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif c == '　':
            result.append(' ')
        else:
            result.append(c)
    return ''.join(result)

# ── 清洗規則 ─────────────────────────────────────────────────────

def clean_city(raw) -> str:
    if not raw:
        return raw
    s = str(raw).strip()
    s = s.replace('臺', '台')
    CITY_NAMES = ['桃園', '台北', '新北', '台中', '台南', '高雄', '基隆',
                  '新竹', '嘉義']
    COUNTY_NAMES = ['苗栗', '彰化', '南投', '雲林', '屏東', '宜蘭',
                    '花蓮', '台東', '澎湖', '金門', '連江', '新竹', '嘉義']
    for c in CITY_NAMES:
        if s == c:
            return c + '市'
    for c in COUNTY_NAMES:
        if s == c:
            return c + '縣'
    return s


def clean_district(raw) -> str:
    if not raw:
        return raw
    s = str(raw).strip()
    s = full_to_half(s)
    DISTRICT_NAMES = [
        '大園', '蘆竹', '中壢', '桃園', '八德', '龜山', '楊梅', '大溪',
        '平鎮', '龍潭', '新屋', '觀音', '復興',
        '南屯', '西屯', '北屯', '中區', '東區', '西區', '南區', '北區',
        '大里', '太平', '霧峰', '烏日', '大肚', '龍井', '沙鹿', '梧棲',
        '清水', '大甲', '外埔', '大安', '神岡', '潭子', '后里', '石岡',
        '東勢', '和平', '新社',
        '中山', '大同', '中正', '萬華', '信義', '松山', '文山',
        '南港', '內湖', '士林', '北投',
        '三重', '板橋', '新店', '中和', '永和', '土城', '樹林', '三峽',
        '鶯歌', '淡水', '汐止', '新莊', '蘆洲', '五股', '泰山', '林口',
    ]
    for d in DISTRICT_NAMES:
        if s == d:
            return d + '區'
    return s


def clean_section(raw) -> str:
    if not raw:
        return raw
    s = str(raw).strip()
    s = full_to_half(s)
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'（.*?）', '', s)
    s = re.sub(r'\s+', '', s)
    return s.strip() if s.strip() else raw


def clean_land_no(raw) -> str:
    """
    統一地號格式為 XXXX-YYYY（主號4碼-子號4碼）。
    - 842.003  → 0842-0003  （小數點 = 主號.子號）
    - 08420003 → 0842-0003  （8位數純數字）
    - 96-3     → 0096-0003  （已有破折號）
    - 842      → 0842-0000  （純整數）
    - 3之51    → 0003-0051  （中文「之」）
    原始格式無法解析時保留原值。
    """
    if not raw:
        return raw
    original = str(raw).strip()
    s = full_to_half(original)
    s = re.sub(r'地號', '', s)
    s = re.sub(r'之', '-', s)
    s = re.sub(r'\s+', '', s)

    # 剩下應為 digits、-、. 的組合
    s = re.sub(r'[^\d\-\.]', '', s)
    if not s:
        return original

    try:
        # 情況 A：含小數點（842.003 / 3.51）
        if '.' in s and '-' not in s:
            parts = s.split('.', 1)
            main = int(parts[0]) if parts[0] else 0
            sub  = int(parts[1]) if parts[1] else 0
            return f'{main:04d}-{sub:04d}'

        # 情況 B：含破折號（96-3 / 0842-0003）
        if '-' in s:
            parts = s.split('-', 1)
            main = int(parts[0]) if parts[0] else 0
            sub  = int(parts[1]) if parts[1] else 0
            return f'{main:04d}-{sub:04d}'

        # 情況 C：8位純數字（08420003）
        digits = re.sub(r'\D', '', s)
        if len(digits) == 8:
            main = int(digits[:4])
            sub  = int(digits[4:])
            return f'{main:04d}-{sub:04d}'

        # 情況 D：純整數（842）
        main = int(digits)
        return f'{main:04d}-0000'

    except (ValueError, IndexError):
        return original


CLEAN_FUNCS = {
    '縣市': clean_city,
    '地區': clean_district,
    '地段': clean_section,
    '地號': clean_land_no,
}

# ── 欄位辨識 ─────────────────────────────────────────────────────

def detect_columns(headers: list) -> dict[str, int]:
    result = {}
    for std_name, aliases in COL_ALIASES.items():
        for i, h in enumerate(headers):
            if h and str(h).strip() in aliases:
                result[std_name] = i
                break
    return result

# ── log 寫出 ─────────────────────────────────────────────────────

def write_log(log_path: Path, lines: list[str]):
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_path.write_text('\n'.join(lines), encoding='utf-8')

# ── 主流程 ────────────────────────────────────────────────────────

def run(xlsx_path: Path, dry_run: bool):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
    except ImportError:
        print('請安裝 openpyxl：pip install openpyxl')
        sys.exit(1)

    ts   = datetime.now().strftime('%Y%m%d_%H%M')
    mode = 'DRY-RUN' if dry_run else '正式執行'

    for d in [INPUT_DIR, OUTPUT_DIR, LOGS_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    log_path  = LOGS_DIR / f'excel_master_clean_{ts}.log'
    log_lines: list[str] = []

    def log(msg: str):
        print(msg)
        log_lines.append(msg)

    log(f'[ {mode} ] clean_excel_master.py')
    log(f'時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    log(f'來源：{xlsx_path}')
    log(f'OUTPUT_DIR：{OUTPUT_DIR}')

    # ── 讀取
    log('\n[讀取] Excel ...')
    wb_ro = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws_ro = wb_ro.active
    all_rows = list(ws_ro.iter_rows(values_only=True))
    wb_ro.close()

    if not all_rows:
        log('Excel 是空的，中止。')
        sys.exit(1)

    headers   = list(all_rows[0])
    data_rows = all_rows[1:]
    total     = len(data_rows)
    log(f'  總列數（不含標題）：{total:,}')

    col_map = detect_columns(headers)
    missing = [k for k in COL_ALIASES if k not in col_map]
    if missing:
        log(f'[錯誤] 找不到欄位：{missing}')
        sys.exit(1)
    log(f'  欄位位置：' + '  '.join(f'{k}@{col_map[k]+1}' for k in col_map))

    # ── 統計清洗差異
    log('\n[分析] 掃描清洗差異 ...')
    counts = {k: 0 for k in col_map}
    total_modified = 0
    examples: list[dict] = []

    modified_rows: list[tuple[int, list]] = []  # (row_idx_0based, new_row_values)

    for i, row in enumerate(data_rows):
        row = list(row)
        changed = False
        row_changes: list[str] = []

        for field, col_idx in col_map.items():
            original = row[col_idx]
            if original is None:
                continue
            cleaned = CLEAN_FUNCS[field](original)
            if str(cleaned) != str(original):
                counts[field] += 1
                changed = True
                row_changes.append(f'{field}：「{original}」→「{cleaned}」')
                row[col_idx] = cleaned

        if changed:
            total_modified += 1
            modified_rows.append((i, row))
            if len(examples) < 30:
                examples.append({
                    'row': i + 2,  # excel 行號（1-based，含標題）
                    'changes': row_changes,
                })

    log(f'  總列數         ：{total:,}')
    log(f'  被修正列數     ：{total_modified:,}')
    log(f'  縣市修正       ：{counts["縣市"]:,} 筆')
    log(f'  行政區修正     ：{counts["地區"]:,} 筆')
    log(f'  地段修正       ：{counts["地段"]:,} 筆')
    log(f'  地號修正       ：{counts["地號"]:,} 筆')

    log(f'\n[範例] 前 {min(30, len(examples))} 筆修正內容：')
    for ex in examples:
        log(f'  第 {ex["row"]} 列：' + '  /  '.join(ex['changes']))

    if dry_run:
        log('\n[dry-run] 不寫入檔案。原始 input 未修改。')
        log(f'確認無誤後執行正式輸出：')
        log(f'  python3 scripts/clean_excel_master.py')
        write_log(log_path, log_lines)
        log(f'[log] {log_path}')
        return

    if total_modified == 0:
        log('\n無需修正，不輸出檔案。')
        write_log(log_path, log_lines)
        return

    # ── 寫出清洗版（完整重建，保留原始格式）
    log('\n[輸出] 載入 Excel（可寫模式）...')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 建立 modified_rows 快速查詢
    modified_idx: dict[int, list] = {i: row for i, row in modified_rows}

    clean_fill = PatternFill('solid', fgColor='E2EFDA')  # 淡綠：已清洗

    for i, row in enumerate(data_rows):
        if i not in modified_idx:
            continue
        excel_row = i + 2  # 含標題列偏移
        new_vals  = modified_idx[i]
        for field, col_idx in col_map.items():
            original = data_rows[i][col_idx]
            cleaned  = new_vals[col_idx]
            if str(cleaned) != str(original if original is not None else ''):
                cell = ws.cell(row=excel_row, column=col_idx + 1, value=cleaned)
                cell.fill = clean_fill

    out_path = OUTPUT_DIR / f'土地主清冊_清洗版_{ts}.xlsx'
    wb.save(out_path)
    wb.close()

    log(f'[輸出] 已儲存：{out_path}')
    log(f'\n{"─"*50}')
    log(f'完成')
    log(f'  修正列數   ：{total_modified:,}')
    log(f'  輸出路徑   ：{out_path}')
    log(f'  原始 input ：未修改')
    write_log(log_path, log_lines)
    log(f'  log 路徑   ：{log_path}')


def main():
    parser = argparse.ArgumentParser(description='主清冊資料清洗')
    parser.add_argument('--dry-run', action='store_true', help='只統計，不寫檔')
    parser.add_argument('--input',   type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    if not args.input.exists():
        print(f'[錯誤] 找不到 Excel：{args.input}')
        sys.exit(1)

    run(args.input, args.dry_run)


if __name__ == '__main__':
    main()
