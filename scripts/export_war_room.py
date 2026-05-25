"""
export_war_room.py
SQLite land_master → 戰情室 Excel（套 MASTER 模板）

執行模式：
  python export_war_room.py          → dry-run（預設，不寫任何檔案）
  python export_war_room.py --apply  → 正式輸出 Excel

輸出路徑：
  終極版/戰情室輸出/land_master_warroom_YYYYMMDD_HHMMSS.xlsx

上色優先順序（高→低）：
  1. 黃底  #FFFF00  effective_transfer_date 半年內
  2. 藍底  #CCE5FF  realprice_match_status = 'hit'
  3. 灰底  #C0C0C0  is_sold = 1

effective_transfer_date = max( parseable(reg_date), parseable(cause_date) )
  支援格式：
    民國 088年04月01日 / 088/04/01
    民國 113/01/04
    西元 2024-01-04
"""

import re
import shutil
import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ── 路徑 ─────────────────────────────────────────────────────────
DB_PATH     = Path("data/database/land_master.db")
MASTER_TMPL = Path("/Users/xiaomingyang/Desktop/excel土地資料維護/最新完成版/老蕭LAND_MASTER.xlsx")
OUT_DIR     = Path("/Users/xiaomingyang/Desktop/excel土地資料維護/終極版/戰情室輸出")

APPLY_MODE = '--apply' in sys.argv
TIMESTAMP  = datetime.now().strftime('%Y%m%d_%H%M%S')
OUT_XLSX   = OUT_DIR / f"land_master_warroom_{TIMESTAMP}.xlsx"

OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── 欄位對應（模板欄順序 → SQLite 欄名）────────────────────────
COL_MAP = [
    ('更新日期',         'updated_at'),
    ('分區',             'zone_type'),
    ('位置',             'location_tag'),
    ('縣市',             'city'),
    ('地區',             'district'),
    ('地段',             'section_raw'),
    ('小段',             'sub_section'),
    ('地號',             'land_no_raw'),
    ('公告現值',         'announced_value'),
    ('次序',             'reg_seq'),
    ('登記日期',         'reg_date'),
    ('登記原因',         'reg_reason'),
    ('發生日期',         'cause_date'),
    ('所有權人',         'owner_name'),
    ('統一編號（完整）', 'owner_id_full'),
    ('郵遞區號',         'postal_code'),
    ('住址',             'address'),
    ('已售出',           'is_sold'),
    ('分母',             'share_denom'),
    ('分子',             'share_numer'),
    ('持分',             'purchase_price'),
    ('持分坪數',         'actual_owned_area'),
    ('土地總坪數',       'total_area_ping'),
    ('備註',             'note'),
    ('電話',             'phone'),
]
DB_COLS = [db for _, db in COL_MAP]

# ── 顏色 ─────────────────────────────────────────────────────────
FILL_YELLOW = PatternFill('solid', fgColor='FFFF00')
FILL_BLUE   = PatternFill('solid', fgColor='CCE5FF')
FILL_GRAY   = PatternFill('solid', fgColor='C0C0C0')

# ── ROC 日期解析 ─────────────────────────────────────────────────
_ROC_FULL  = re.compile(r'^(\d{2,3})年(\d{1,2})月(\d{1,2})日$')
_ROC_SLASH = re.compile(r'^(\d{2,3})[/\-](\d{1,2})[/\-](\d{1,2})$')
_WESTERN   = re.compile(r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$')

def parse_date(s) -> date | None:
    s = '' if s is None else str(s).strip()
    if not s or s.lower() in ('nan', 'none', ''):
        return None
    m = _ROC_FULL.match(s)
    if m:
        y, mo, d = int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))
        try: return date(y, mo, d)
        except ValueError: return None
    m = _ROC_SLASH.match(s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 1000:
            y += 1911
        try: return date(y, mo, d)
        except ValueError: return None
    m = _WESTERN.match(s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try: return date(y, mo, d)
        except ValueError: return None
    return None

def effective_date(reg_date_s, cause_date_s) -> date | None:
    d1 = parse_date(reg_date_s)
    d2 = parse_date(cause_date_s)
    if d1 and d2: return max(d1, d2)
    return d1 or d2

# ── 讀取 SQLite ─────────────────────────────────────────────────
print("讀取 land_master…")
conn = sqlite3.connect(str(DB_PATH))
extra_cols = [c for c in ['realprice_match_status'] if c not in DB_COLS]
df = pd.read_sql(
    f"SELECT {', '.join(DB_COLS + extra_cols)} FROM land_master ORDER BY id",
    conn
)
conn.close()
print(f"  {len(df):,} 筆")

# ── 上色分類 ─────────────────────────────────────────────────────
today     = date.today()
cutoff    = today - timedelta(days=183)   # 約半年

n_yellow  = 0
n_blue    = 0
n_gray    = 0
n_multi   = 0
n_no_date = 0

color_tags = []   # 每列：'yellow' | 'blue' | 'gray' | None

for _, row in df.iterrows():
    eff = effective_date(row['reg_date'], row['cause_date'])
    is_recent    = (eff is not None) and (eff >= cutoff)
    rp_val       = row['realprice_match_status']
    is_realprice = str(rp_val) .strip().lower() == 'hit' if rp_val is not None else False
    sold_val     = row['is_sold']
    is_sold      = str(sold_val).strip() in ('1', '1.0', 'True', 'true', '是') if sold_val is not None else False
    no_date      = (eff is None)

    if no_date and not is_realprice and not is_sold:
        n_no_date += 1

    conditions = sum([is_recent, is_realprice, is_sold])
    if conditions > 1:
        n_multi += 1

    if is_recent:
        color_tags.append('yellow')
        n_yellow += 1
    elif is_realprice:
        color_tags.append('blue')
        n_blue += 1
    elif is_sold:
        color_tags.append('gray')
        n_gray += 1
    else:
        color_tags.append(None)

# ── Dry-run 報告 ─────────────────────────────────────────────────
print()
print("=== DRY-RUN 戰情室上色統計 ===")
print(f"  總列數                     : {len(df):,}")
print(f"  黃底（近期移轉，半年內）   : {n_yellow:,}")
print(f"  藍底（實價命中）           : {n_blue:,}")
print(f"  灰底（已售出 is_sold=1）   : {n_gray:,}")
print(f"  同時符合多條件（有衝突）   : {n_multi:,}")
print(f"  無法判斷日期（無法染黃）   : {n_no_date:,}")
print(f"  無底色                     : {len(df) - n_yellow - n_blue - n_gray:,}")
print()
print(f"  模板來源 : {MASTER_TMPL}")
print(f"  輸出路徑 : {OUT_XLSX}")
print()

if not APPLY_MODE:
    print("=== DRY-RUN 完成，未寫入任何檔案 ===")
    print("正式輸出請加 --apply 參數。")
    sys.exit(0)

# ════════════════════════════════════════════════════════════════
# --apply 模式
# ════════════════════════════════════════════════════════════════

# ── 複製模板 ─────────────────────────────────────────────────────
print(f"複製模板 → {OUT_XLSX.name}")
shutil.copy2(MASTER_TMPL, OUT_XLSX)

# ── 開啟並寫入資料 ───────────────────────────────────────────────
print("開啟 workbook…")
wb = load_workbook(OUT_XLSX)
ws = wb.active

# 清除模板殘留資料列（若模板已是空白，max_row=1，此步驟跳過）
if ws.max_row >= 2:
    ws.delete_rows(2, ws.max_row - 1)

print(f"寫入 {len(df):,} 列資料…")
for row_idx, (df_row, tag) in enumerate(zip(df.itertuples(index=False), color_tags), start=2):
    for col_idx, (_, db_col) in enumerate(COL_MAP, start=1):
        val = getattr(df_row, db_col, None)
        # pandas NaN → None
        if val is not None and not isinstance(val, str):
            try:
                import math
                if math.isnan(float(val)):
                    val = None
            except (TypeError, ValueError):
                pass
        ws.cell(row=row_idx, column=col_idx, value=val)

    if tag == 'yellow':
        fill = FILL_YELLOW
    elif tag == 'blue':
        fill = FILL_BLUE
    elif tag == 'gray':
        fill = FILL_GRAY
    else:
        continue

    for col_idx in range(1, len(COL_MAP) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

# ── 修正 AutoFilter 範圍 ─────────────────────────────────────────
last_col = get_column_letter(len(COL_MAP))
ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

# ── 儲存 ─────────────────────────────────────────────────────────
print("儲存…")
wb.save(OUT_XLSX)
wb.close()

print()
print("=== 輸出完成 ===")
print(f"  路徑    : {OUT_XLSX}")
print(f"  資料列  : {len(df):,}")
print(f"  黃底    : {n_yellow:,}")
print(f"  藍底    : {n_blue:,}")
print(f"  灰底    : {n_gray:,}")
print(f"  SQLite  : 未動")
print(f"  MASTER  : 未動")
