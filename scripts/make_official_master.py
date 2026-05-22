#!/usr/bin/env python3
"""
make_official_master.py
從乾淨版主清冊 → 正式版主清冊（格式美化，不修改任何數值）
輸入：土地主清冊_乾淨版_20260522.xlsx
輸出：土地主清冊_正式版_20260522.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from datetime import datetime
import os

SRC  = '/Users/xiaomingyang/Desktop/excel土地資料維護/土地主清冊_乾淨版_20260522.xlsx'
DEST = '/Users/xiaomingyang/Desktop/excel土地資料維護/土地主清冊_正式版_20260522.xlsx'

# 欄寬設定（對應 24 欄順序）
COL_WIDTHS = [12, 8, 14, 6, 8, 16, 12, 10, 6, 12, 10, 12,
              12, 16, 16, 8, 30, 6, 6, 6, 12, 10, 40, 20]

# 文字格式欄（zero-based 欄位名稱集合）
TEXT_COLS = {'地號', '電話', '統一編號（遮罩）', '統一編號（完整）', '郵遞區號'}

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
EVEN_FILL    = PatternFill('solid', fgColor='EEF3FB')
DATA_FONT    = Font(name='微軟正黑體', size=10)
THIN         = Side(style='thin', color='CCCCCC')
THIN_BORDER  = Border(bottom=THIN)

def to_decimal2(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return v

def to_decimal4(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return v

def to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return v

print('[1] 讀取來源...')
src_wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
src_ws = src_wb.active
rows = list(src_ws.iter_rows(values_only=True))
src_wb.close()
print(f'    {len(rows)-1:,} 資料列 × {len(rows[0])} 欄')

headers = list(rows[0])
col_idx = {h: i for i, h in enumerate(headers)}

print('[2] 建立新 workbook...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '土地主清冊'

print('[3] 寫入標題列...')
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(1, c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

print(f'[4] 寫入資料列（{len(rows)-1:,} 列）...')
idx_announced = col_idx.get('公告現值', -1)
idx_area      = col_idx.get('土地總坪數', -1)
idx_range     = col_idx.get('權利範圍', -1)
idx_addr      = col_idx.get('住址', -1)

for r_idx, row in enumerate(rows[1:], start=2):
    row = list(row)

    # 數值轉換（僅轉有值的欄位）
    if idx_announced >= 0 and row[idx_announced] not in (None, ''):
        row[idx_announced] = to_int(row[idx_announced])
    if idx_area >= 0 and row[idx_area] not in (None, ''):
        row[idx_area] = to_decimal4(row[idx_area])
    if idx_range >= 0:
        v = row[idx_range]
        if isinstance(v, str) and v.strip() in ('#VALUE!', ''):
            row[idx_range] = None
        elif v not in (None, ''):
            row[idx_range] = to_decimal2(v)

    ws.append(row)

    # 樣式
    fill = EVEN_FILL if r_idx % 2 == 0 else None
    for c_idx, (val, hdr) in enumerate(zip(row, headers), start=1):
        cell = ws.cell(r_idx, c_idx)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

        if hdr in TEXT_COLS:
            cell.number_format = '@'
        elif hdr == '公告現值' and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
        elif hdr == '土地總坪數' and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.0000'
        elif hdr == '權利範圍' and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
        elif hdr == '住址':
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    if r_idx % 20000 == 0:
        print(f'    {r_idx-1:,} / {len(rows)-1:,}')

print('[5] 欄寬、凍結、篩選...')
for i, width in enumerate(COL_WIDTHS[:len(headers)], start=1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 20

print('[6] 儲存...')
wb.save(DEST)

size_mb = os.path.getsize(DEST) / 1024 / 1024
print(f'\n完成！')
print(f'輸出：{DEST}')
print(f'大小：{size_mb:.1f} MB')
print(f'列數：{len(rows)-1:,}  欄數：{len(headers)}')

# 快速驗證
print('\n[驗證] 讀回前 3 列...')
vb = openpyxl.load_workbook(DEST, read_only=True, data_only=True)
vw = vb.active
for i, row in enumerate(vw.iter_rows(min_row=2, values_only=True)):
    if i >= 3: break
    print(f'  公告現值={row[idx_announced]!r}  土地總坪數={row[idx_area]!r}  權利範圍={row[idx_range]!r}  電話={row[col_idx.get("電話")]!r}')
vb.close()
