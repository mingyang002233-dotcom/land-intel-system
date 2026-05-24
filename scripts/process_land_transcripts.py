#!/usr/bin/env python3
"""
process_land_transcripts.py  v5.2
電傳 YAML → 解析 → 更新 SQLite land_master.db → 同步主清冊 Excel

用法：
  python3 scripts/process_land_transcripts.py             # 正式執行
  python3 scripts/process_land_transcripts.py --dry-run   # 只解析，不寫入，不移動檔案

資料夾：
  電傳解析/     ← download_watcher 自動搬入（inbox），也可手動放入 .yaml / .yaml.txt
  電傳已完成/   ← 成功後移入
  電傳錯誤/     ← 解析失敗移入
  電傳待解析/   ← 已廢棄（deprecated），請改用電傳解析/

YAML 格式（最小範例）：
  縣市: 新北市
  地區: 林口區
  地段: 力行段
  地號: "595"
  所有人:
    - 姓名: 王大明
      統一編號遮罩: H122*****1
      統一編號完整: H122345671
      登記次序: "0001"
      登記日期: "114/05/28"
      登記原因: 買賣
      原因發生日期: "114/05/20"
      分子: "1"
      分母: "1"
      地址: 新北市林口區文化一路100號
  公告現值: 12000
  土地面積坪: 638.384
"""

import argparse
import hashlib
import os
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH      = PROJECT_ROOT / 'data' / 'database' / 'land_master.db'

INBOX_DIR    = PROJECT_ROOT / '電傳解析'
DONE_DIR     = Path('/Users/xiaomingyang/Desktop/excel土地資料維護/電傳已完成')
ERROR_DIR    = Path('/Users/xiaomingyang/Desktop/excel土地資料維護/電傳錯誤')
EXCEL_DIR    = Path('/Users/xiaomingyang/Desktop/excel土地資料維護')
EXCEL_MASTER = EXCEL_DIR / '最新完成版' / '老蕭LAND_MASTER.xlsx'

# Excel 欄位索引（0-based），與主清冊 27 欄定版一致
# AB–AF（col 27–31）為系統判定欄位，只存在於 TEST 階段直至正式驗收
_XC = {
    '更新日期': 0, '分區': 1, '位置': 2, '縣市': 3, '地區': 4,
    '地段': 5, '小段': 6, '地號': 7, '公告現值': 8, '次序': 9, '登記日期': 10,
    '登記原因': 11, '發生日期': 12, '所有權人': 13,
    '統一編號（遮罩）': 14, '統一編號（完整）': 15, '郵遞區號': 16,
    '住址': 17, '已售出': 18, '分母': 19, '分子': 20,
    '持分': 21, '持分坪數': 22, '土地總坪數': 23, '權利範圍': 24, '備註': 25, '電話': 26,
    # ── 系統判定欄位（AB–AF）──
    '系統處理狀態': 27, '系統處理備註': 28, '系統來源': 29,
    '系統更新時間': 30, '系統批次ID': 31,
}
# 系統欄位只在 EXCEL_HAS_SYS_COLS=True 時寫入（TEST 驗收後正式加欄頭再啟用）
EXCEL_HAS_SYS_COLS = True

# ── 共有人名單差異比對 ──────────────────────────────────────────────────────────

def _id_value(rec: dict) -> str:
    """取 ID 純值（不含欄位前綴），用於跨欄比對。優先：完整 > 遮罩。"""
    full = str(rec.get('owner_id_full') or '').strip().upper()
    if full and full not in ('', 'NONE'):
        return full
    masked = str(rec.get('owner_id_masked') or '').strip().upper()
    if masked and masked not in ('', 'NONE'):
        return masked
    return ''


def _owner_id(rec: dict) -> str:
    """
    產生比對 key。
    規則：兩邊只要 ID 純值相同即配對（不分完整/遮罩欄），ID 都空才退回姓名。
    """
    val = _id_value(rec)
    if val:
        return f'ID:{val}'
    name = str(rec.get('owner_name') or '').strip()
    return f'NAME:{name}'


def _share_str(rec: dict) -> str:
    n = str(rec.get('share_numer') or '').strip()
    d = str(rec.get('share_denom') or '').strip()
    if n and d:
        return f'{n}/{d}'
    return ''


def diff_owners(new_recs: list[dict], old_recs: list[dict]) -> dict:
    """
    共有人名單差異比對（名單 diff，不是整筆地號 sold）。
    new_recs : 電傳中的所有人列表
    old_recs : 同地號現有地主列表

    回傳：
      disappeared  : list[old_rec]    舊有新無 → 標已售出
      added        : list[new_rec]    新有舊無 → 插入新列
      unchanged    : list[(old, new)] 兩邊皆有，持分相同
      share_changed: list[(old, new)] 兩邊皆有，持分不同 → 備註持分異動
    """
    new_by_id = {_owner_id(r): r for r in new_recs}
    old_by_id = {_owner_id(r): r for r in old_recs}

    new_ids = set(new_by_id)
    old_ids = set(old_by_id)

    disappeared   = [old_by_id[i] for i in old_ids - new_ids]
    added         = [new_by_id[i] for i in new_ids - old_ids]
    unchanged     = []
    share_changed = []

    for i in old_ids & new_ids:
        old_r = old_by_id[i]
        new_r = new_by_id[i]
        if _share_str(old_r) and _share_str(new_r) and _share_str(old_r) != _share_str(new_r):
            share_changed.append((old_r, new_r))
        else:
            unchanged.append((old_r, new_r))

    return {
        'disappeared':   disappeared,
        'added':         added,
        'unchanged':     unchanged,
        'share_changed': share_changed,
    }


# ── 與 import_land_master.py 保持完全一致的 key 函數 ─────────────────────────

def _compute_sold_mark_permission(records: list[dict], cur) -> dict:
    """
    判斷每個地號是否允許「標記消失地主已售出」及「寫入 Excel 最新狀態」。

    規則：
      - DB 無此地號（全新）        → allow_sold_mark=True  （新地號，直接通過）
      - yaml_latest > db_latest   → allow_sold_mark=True  （新電傳，允許更新）
      - yaml_latest == db_latest  → allow_sold_mark=False （同日不同事件，只補 SQLite）
      - yaml_latest < db_latest   → allow_sold_mark=False （舊電傳，只補 SQLite 歷史）
      - yaml_latest 無法解析      → allow_sold_mark=True  （無日期，寧放不擋）

    回傳 dict[land_match_key] = {
        'allow_sold_mark': bool,
        'reason': str,        -- 供 log / 顯示用
        'yaml_date': str,
        'db_date':   str,
        'section_raw': str,
        'land_no_raw': str,
    }
    """
    from collections import defaultdict
    by_land: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        by_land[rec['land_match_key']].append(rec)

    result = {}
    for lmk, recs in by_land.items():
        sample = recs[0]

        # 電傳中此地號最新登記日期
        yaml_dates = [_parse_roc_date(r.get('reg_date') or '') for r in recs]
        yaml_dates = [d for d in yaml_dates if d]
        yaml_best  = max(yaml_dates) if yaml_dates else None

        # DB 現有 is_sold=0 記錄的最新登記日期
        rows = cur.execute(
            "SELECT reg_date FROM land_master "
            "WHERE land_match_key=? AND is_sold=0 LIMIT 200",
            (lmk,)
        ).fetchall()
        db_count  = len(rows)
        db_dates  = [_parse_roc_date(r[0]) for r in rows if r[0]]
        db_dates  = [d for d in db_dates if d]
        db_best   = max(db_dates) if db_dates else None

        if db_count == 0:
            allow      = True
            reason     = '新地號（DB無資料）'
            sys_status = '正常'
            sys_note   = '新地號首次匯入'
        elif yaml_best is None:
            allow      = True
            reason     = '電傳無登記日期，寧放不擋'
            sys_status = '待人工確認'
            sys_note   = '電傳缺少登記日期，系統無法比對新舊，請人工確認'
        elif yaml_best > db_best:
            allow      = True
            reason     = f'新電傳 {yaml_best} > DB最新 {db_best}'
            sys_status = '正常'
            sys_note   = ''
        elif yaml_best == db_best:
            allow      = False
            reason     = f'同日事件 {yaml_best}，不標已售'
            sys_status = '待人工確認'
            sys_note   = f'同日事件（{yaml_best}），未標已售，請人工確認是否完整名單'
        else:
            allow      = False
            reason     = f'舊電傳 {yaml_best} < DB最新 {db_best}，只補SQLite歷史'
            sys_status = '舊電傳補歷史'
            sys_note   = f'系統已有較新登記事件(DB最新:{db_best})，本次僅補歷史，未標已售'

        result[lmk] = {
            'allow_sold_mark': allow,
            'reason':          reason,
            'sys_status':      sys_status,
            'sys_note':        sys_note,
            'yaml_date':       yaml_best or '（無日期）',
            'db_date':         db_best   or '（無資料）',
            'section_raw':     sample.get('section_raw', ''),
            'land_no_raw':     sample.get('land_no_raw', ''),
        }
    return result


def normalize_section(raw):
    if not raw: return ''
    s = str(raw).strip()
    s = re.sub(r'[\(（][^)\）]*[\)）]', '', s)
    return re.sub(r'\s+', '', s)


def normalize_land_no(raw):
    if not raw: return ''
    s = str(raw).strip()
    s = re.sub(r'之', '-', s)
    s = re.sub(r'[^\d\-]', '', s)
    if not s: return ''
    parts = s.split('-')
    try:
        main = int(parts[0]) if parts[0] else 0
        sub  = int(parts[1]) if len(parts) > 1 and parts[1] else 0
        return f'{main:04d}-{sub:04d}'
    except ValueError:
        return s


def make_land_match_key(city, district, norm_sec, norm_no):
    return '_'.join(s.strip() for s in [city, district, norm_sec, norm_no] if s.strip())


def make_owner_key(owner_id_full, owner_name, owner_id_masked):
    if owner_id_full and str(owner_id_full).strip():
        src = str(owner_id_full).strip().upper()
    else:
        name   = (owner_name   or '').strip()
        masked = (owner_id_masked or '').strip()
        src    = f'{name}|{masked}'
    if not src or src == '|': return ''
    return hashlib.sha256(src.encode()).hexdigest()[:16]


def make_event_key(rec):
    parts = [
        rec.get('land_match_key') or '',
        rec.get('owner_key')      or '',
        str(rec.get('reg_seq')    or ''),
        str(rec.get('reg_date')   or ''),
        str(rec.get('reg_reason') or ''),
        str(rec.get('share_numer') or ''),
        str(rec.get('share_denom') or ''),
    ]
    if not any(parts[:2]): return ''
    raw = '|'.join(parts)
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def calc_actual_area(total, numer, denom):
    try:
        t, n, d = float(total), float(numer), float(denom)
        return round(t * n / d, 4) if d else None
    except (TypeError, ValueError):
        return None


# ── YAML 解析 ────────────────────────────────────────────────────────────────

class ParseError(Exception):
    pass


# ── 地政謄本原始格式解析用工具 ────────────────────────────────────────────────

_CITY_PAT_RAW = re.compile(
    r'(台北市|臺北市|新北市|桃園市|台中市|臺中市|台南市|臺南市|高雄市|'
    r'基隆市|新竹市|嘉義市|苗栗縣|彰化縣|南投縣|雲林縣|屏東縣|宜蘭縣|'
    r'花蓮縣|台東縣|臺東縣|澎湖縣|金門縣|連江縣|新竹縣|嘉義縣)'
)
_DIST_PAT_RAW = re.compile(r'([^\s]+?(?:區|鄉|鎮|市))')
_LAND_NO_PAT  = re.compile(r'(\d{4}-\d{4})')


def _parse_land_id_str(s: str) -> tuple:
    """'桃園市桃園區國際段0174-0006地號' → (city, district, section, sub_section, land_no)"""
    s = str(s).strip().replace('臺', '台')
    s = re.sub(r'地號.*$', '', s).strip()

    city_m = _CITY_PAT_RAW.search(s)
    if not city_m:
        return '', '', '', '', ''
    city = city_m.group(1)
    rest = s[city_m.end():].strip()

    dist_m = _DIST_PAT_RAW.match(rest)
    if not dist_m:
        return city, '', '', '', ''
    district = dist_m.group(1)
    rest = rest[dist_m.end():].strip()

    lno_m = _LAND_NO_PAT.search(rest)
    if not lno_m:
        return city, district, rest.strip(), '', ''
    land_no = lno_m.group(1)
    sec_part = rest[:lno_m.start()].strip()

    # 若地段名含小段，例如「國際段一小段」→ section=國際段, sub=一小段
    sub_m = re.search(r'^(.+段)(.+小段)$', sec_part)
    if sub_m:
        section, sub_section = sub_m.group(1), sub_m.group(2)
    else:
        section, sub_section = sec_part, ''
    return city, district, section, sub_section, land_no


def _parse_roc_date(s) -> str:
    """民國115年03月27日 → 115/03/27"""
    if not s:
        return ''
    s = str(s).strip()
    m = re.match(r'(?:民國)?(\d+)年(\d+)月(\d+)日', s)
    if m:
        return f'{m.group(1)}/{int(m.group(2)):02d}/{int(m.group(3)):02d}'
    return s


def _parse_fraction(s) -> tuple:
    """'100分之28' → ('28', '100')  (numer, denom)"""
    if not s:
        return '', ''
    m = re.match(r'(\d+)分之(\d+)', str(s).strip())
    if m:
        return m.group(2), m.group(1)
    return '', ''


def _parse_numeric_value(s) -> 'float | None':
    """'13,100 元/平方公尺' → 13100.0"""
    if not s:
        return None
    try:
        return float(re.search(r'\d+(?:\.\d+)?', str(s).replace(',', '')).group())
    except (AttributeError, ValueError):
        return None


def _sqm_to_ping(s) -> 'float | None':
    """'420.01平方公尺' → 127.0605 坪"""
    val = _parse_numeric_value(s)
    return round(val / 3.30579, 4) if val else None


def parse_raw_registry(data: dict) -> list[dict]:
    """
    解析地政謄本原始 YAML 格式。
    結構：土地標示部（dict）+ 土地所有權部（list）
    """
    biao     = data.get('土地標示部')
    own_list = data.get('土地所有權部')

    if not isinstance(biao, dict):
        raise ParseError('地政謄本格式：缺少 土地標示部（dict）')
    if not isinstance(own_list, list) or len(own_list) == 0:
        raise ParseError('地政謄本格式：缺少 土地所有權部（list），或為空')

    land_id_str = str(biao.get('地號') or '').strip()
    city, district, section, sub_section, land_no = _parse_land_id_str(land_id_str)

    if not city or not district or not section or not land_no:
        raise ParseError(
            f'地政謄本格式：無法從地號字串解析地籍資訊：「{land_id_str}」'
            f'（縣市={city!r}, 地區={district!r}, 地段={section!r}, 地號={land_no!r}）'
        )

    ann_val        = _parse_numeric_value(biao.get('公告土地現值'))
    total_area_ping = _sqm_to_ping(biao.get('面積'))

    norm_sec = normalize_section(section)
    norm_no  = normalize_land_no(land_no)
    lmk      = make_land_match_key(city, district, norm_sec, norm_no)

    records = []
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for idx, owner_data in enumerate(own_list):
        if not isinstance(owner_data, dict):
            raise ParseError(f'土地所有權部[{idx}] 不是 dict：{owner_data!r}')

        name       = str(owner_data.get('所有權人') or '').strip()
        masked     = str(owner_data.get('統一編號') or '').strip()
        reg_seq    = str(owner_data.get('登記次序') or '').strip()
        reg_date   = _parse_roc_date(owner_data.get('登記日期'))
        reg_reason = str(owner_data.get('登記原因') or '').strip()
        cause_date = _parse_roc_date(owner_data.get('原因發生日期'))
        warrant_no = str(owner_data.get('權狀字號') or '').strip()

        addr_raw   = str(owner_data.get('地址') or '').strip()
        address    = '' if addr_raw in ('(圖)', '（圖）') else addr_raw

        share_str  = str(owner_data.get('權利範圍') or '').strip()
        numer, denom = _parse_fraction(share_str)

        if not name:
            raise ParseError(f'土地所有權部[{idx}] 缺少所有權人姓名')

        # 唯一 key：登記次序優先，退而權狀字號
        owner_key = make_owner_key(None, name, masked or None)
        actual    = calc_actual_area(total_area_ping, numer, denom)

        rec = {
            'city':               city,
            'district':           district,
            'section_raw':        section,
            'sub_section':        sub_section,
            'land_no_raw':        land_no,
            'announced_value':    ann_val,
            'total_area_ping':    total_area_ping,
            'reg_seq':            reg_seq or None,
            'reg_date':           reg_date or None,
            'cause_date':         cause_date or None,
            'reg_reason':         reg_reason or None,
            'owner_name':         name,
            'owner_id_masked':    masked or None,
            'owner_id_full':      None,
            'postal_code':        None,
            'address':            address or None,
            'phone':              None,
            'share_numer':        numer or None,
            'share_denom':        denom or None,
            'note':               warrant_no or None,
            'normalized_section': norm_sec,
            'normalized_land_no': norm_no,
            'land_match_key':     lmk,
            'owner_key':          owner_key,
            'actual_owned_area':  actual,
            'updated_at':         now_str,
            'imported_at':        now_str,
        }
        rec['event_key'] = make_event_key(rec)
        records.append(rec)

    return records


def parse_transcript(path: Path) -> list[dict]:
    """
    解析一份電傳 YAML，回傳 list of record dict（每位所有人一筆）。
    若格式錯誤 raise ParseError。
    """
    try:
        text = path.read_text(encoding='utf-8')
        data = yaml.safe_load(text)
    except Exception as e:
        raise ParseError(f'YAML 讀取失敗：{e}')

    if not isinstance(data, dict):
        raise ParseError('YAML 根層級必須是 dict')

    # 偵測格式：地政謄本原始格式（土地標示部 / 土地所有權部）
    if '土地標示部' in data or '土地所有權部' in data:
        return parse_raw_registry(data)

    # 標準整理格式（縣市、地區、地段、地號、所有人）
    city      = str(data.get('縣市') or '').strip()
    district  = str(data.get('地區') or '').strip()
    sec_raw   = str(data.get('地段') or '').strip()
    no_raw    = str(data.get('地號') or '').strip()
    ann_val   = data.get('公告現值')
    total_area= data.get('土地面積坪')

    if not city or not district or not sec_raw or not no_raw:
        raise ParseError(f'缺少必要欄位（縣市/地區/地段/地號）：{dict(縣市=city,地區=district,地段=sec_raw,地號=no_raw)}')

    owners = data.get('所有人') or []
    if not isinstance(owners, list) or len(owners) == 0:
        raise ParseError('所有人 欄位必須為非空 list')

    norm_sec = normalize_section(sec_raw)
    norm_no  = normalize_land_no(no_raw)
    lmk      = make_land_match_key(city, district, norm_sec, norm_no)

    records = []
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for owner_data in owners:
        if not isinstance(owner_data, dict):
            raise ParseError(f'所有人 項目格式錯誤：{owner_data!r}')

        name        = str(owner_data.get('姓名') or '').strip()
        masked      = str(owner_data.get('統一編號遮罩') or '').strip()
        full_id     = str(owner_data.get('統一編號完整') or '').strip()
        reg_seq     = str(owner_data.get('登記次序') or '').strip()
        reg_date    = str(owner_data.get('登記日期') or '').strip()
        reg_reason  = str(owner_data.get('登記原因') or '').strip()
        cause_date  = str(owner_data.get('原因發生日期') or '').strip()
        share_numer = str(owner_data.get('分子') or '').strip()
        share_denom = str(owner_data.get('分母') or '').strip()
        address     = str(owner_data.get('地址') or '').strip()
        postal      = str(owner_data.get('郵遞區號') or '').strip()
        phone       = str(owner_data.get('電話') or '').strip()
        note        = str(owner_data.get('備註') or '').strip()

        if not name:
            raise ParseError(f'所有人缺少姓名：{owner_data!r}')

        owner_key = make_owner_key(full_id or None, name, masked or None)
        actual    = calc_actual_area(total_area, share_numer, share_denom)

        rec = {
            # 地籍
            'city':              city,
            'district':          district,
            'section_raw':       sec_raw,
            'land_no_raw':       no_raw,
            'announced_value':   ann_val,
            'total_area_ping':   total_area,
            # 登記事件
            'reg_seq':           reg_seq or None,
            'reg_date':          reg_date or None,
            'cause_date':        cause_date or None,
            'reg_reason':        reg_reason or None,
            # 所有人
            'owner_name':        name,
            'owner_id_masked':   masked or None,
            'owner_id_full':     full_id or None,
            'postal_code':       postal or None,
            'address':           address or None,
            'phone':             phone or None,
            'share_numer':       share_numer or None,
            'share_denom':       share_denom or None,
            'note':              note or None,
            # Python 生成
            'normalized_section': norm_sec,
            'normalized_land_no': norm_no,
            'land_match_key':    lmk,
            'owner_key':         owner_key,
            'actual_owned_area': actual,
            # 系統
            'updated_at':        now_str,
            'imported_at':       now_str,
        }
        rec['event_key'] = make_event_key(rec)
        records.append(rec)

    return records


# ── SQLite DDL（log 表）────────────────────────────────────────────────────

DDL_LOG = """
CREATE TABLE IF NOT EXISTS transcript_import_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    filename    TEXT,
    event_key   TEXT,
    owner_name  TEXT,
    land_key    TEXT,
    reg_reason  TEXT,
    status      TEXT,       -- inserted / skipped / sold_marked / error
    error_msg   TEXT,
    created_at  TEXT DEFAULT (datetime('now'))
);
"""

INSERT_SQL = """
INSERT INTO land_master (
    updated_at, city, district, section_raw, land_no_raw,
    announced_value, total_area_ping,
    reg_seq, reg_date, cause_date, reg_reason,
    owner_name, owner_id_masked, owner_id_full,
    postal_code, address, phone,
    share_numer, share_denom, note,
    normalized_section, normalized_land_no,
    land_match_key, owner_key, event_key,
    actual_owned_area, imported_at,
    is_sold, realprice_match_status,
    sys_status, sys_note, sys_source, sys_updated_at, sys_batch_id
) VALUES (
    :updated_at, :city, :district, :section_raw, :land_no_raw,
    :announced_value, :total_area_ping,
    :reg_seq, :reg_date, :cause_date, :reg_reason,
    :owner_name, :owner_id_masked, :owner_id_full,
    :postal_code, :address, :phone,
    :share_numer, :share_denom, :note,
    :normalized_section, :normalized_land_no,
    :land_match_key, :owner_key, :event_key,
    :actual_owned_area, :imported_at,
    0, 'pending',
    :sys_status, :sys_note, :sys_source, :sys_updated_at, :sys_batch_id
)
"""


# ── Telegram 推播 ─────────────────────────────────────────────────────────────

def _send_telegram(text: str):
    import urllib.request, urllib.parse, json
    from pathlib import Path as _P
    env_path = _P(__file__).resolve().parent.parent / '.env'
    token = chat_id = ''
    for line in env_path.read_text().splitlines():
        if line.startswith('TELEGRAM_BOT_TOKEN='):
            token = line.split('=', 1)[1].strip()
        if line.startswith('TELEGRAM_CHAT_ID='):
            chat_id = line.split('=', 1)[1].strip()
    if not token or not chat_id:
        print('  [TG] .env 缺少 token/chat_id，略過推播')
        return
    url = f'https://api.telegram.org/bot{token}/sendMessage'
    payload = json.dumps({'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'}).encode()
    req = urllib.request.Request(url, data=payload,
                                  headers={'Content-Type': 'application/json'})
    try:
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        print(f'  [TG] 推播失敗：{e}')


def build_tg_message(filename: str, records: list[dict],
                     inserted: list[dict], sold_marked: int) -> str:
    lines = [f'✅ 電傳解析完成\n📁 {filename}']
    for rec in inserted:
        lines.append(
            f'\n📍 {rec["section_raw"]} {rec["land_no_raw"]}'
            f'\n👤 新地主：{rec["owner_name"]}'
            f'\n📄 登記原因：{rec.get("reg_reason") or "—"}'
            f'\n📅 登記日期：{rec.get("reg_date") or "—"}'
        )
    if sold_marked:
        lines.append(f'\n🏷 已標記 {sold_marked} 筆舊地主為已售出')
    return '\n'.join(lines)


# ── Excel 格式規則（顯示用，不改資料）────────────────────────────────────────
#
# 優先順序（高→低）：
#   1. 已售出 → 整列灰字 + 淡灰底（最高優先）
#   2. 近半年買賣 → 整列淡黃底
#   3. 其他 → 無特殊格式
#
# 已售出判斷：'已售出' 欄值為 '已售'/'是'/'y'/'1'/1/True 之一
# 近半年買賣：登記原因='買賣' AND 登記日期距今 ≤ 180 天 AND 未售出
#

_SOLD_VALS  = {'已售', '是', 'y', '1', '已售出'}   # lower-case 比對

from openpyxl.styles import Font, PatternFill, Color

MASTER_FONT      = Font(name='微軟正黑體', size=10)
SOLD_FONT_STYLE  = Font(name='微軟正黑體', size=10, color='999999')
SOLD_FILL_STYLE  = PatternFill('solid', fgColor='EBEBEB')   # 淡灰
RECENT_FILL_STYLE= PatternFill('solid', fgColor='FFFACD')   # 淡黃（lemon chiffon）
MARK_SOLD_FILL   = PatternFill('solid', fgColor='FFD7D7')   # 淡紅（標記已售瞬間保留，下次 reformat 轉灰）
NO_FILL          = PatternFill(fill_type=None)


def _sold_flag(v) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in _SOLD_VALS or v is True or v == 1


def _roc_date_to_ad(s) -> 'datetime.date | None':
    """將民國日期字串轉為西元 date，供半年比對用。"""
    from datetime import date
    import re
    if not s:
        return None
    s = str(s).strip()
    try:
        m = re.match(r'(\d+)年(\d+)月(\d+)日', s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y + 1911, mo, d)
        m = re.match(r'(\d+)/(\d+)/(\d+)', s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y + 1911, mo, d)
    except (ValueError, OverflowError):
        pass
    return None


def _row_fill_and_font(sold_val, reg_reason, reg_date_str, cutoff):
    """回傳 (fill, font) for a data row. cutoff = date 180 days ago."""
    if _sold_flag(sold_val):
        return SOLD_FILL_STYLE, SOLD_FONT_STYLE
    if str(reg_reason or '').strip() == '買賣':
        d = _roc_date_to_ad(reg_date_str)
        if d and d >= cutoff:
            return RECENT_FILL_STYLE, MASTER_FONT
    return NO_FILL, MASTER_FONT


def _apply_row_format(ws, row_num: int, fill, font, max_col: int):
    """對單列套用格式（只改 fill + font color，不改 border/alignment）。"""
    for c in range(1, max_col + 1):
        cell = ws.cell(row_num, c)
        cell.fill = fill
        cell.font = font


def reformat_and_sort_master(dry_run: bool = False) -> dict:
    """
    全表排序 + 重新套用格式。
    排序規則：地段 → 地號 → 登記日期（民國升序）→ 登記次序
    格式規則：已售出→灰；近半年買賣→淡黃；其他→正常
    回傳 {'sorted': True, 'sold_grey': N, 'recent_yellow': N, 'elapsed': secs}
    """
    import time
    from datetime import date, timedelta
    import openpyxl

    t0 = time.time()
    cutoff = date.today() - timedelta(days=180)

    EXCEL = str(EXCEL_MASTER)
    SEC_COL    = _XC['地段']        # 0-based
    NO_COL     = _XC['地號']
    DATE_COL   = _XC['登記日期']
    SEQ_COL    = _XC['次序']
    SOLD_COL   = _XC['已售出']
    REASON_COL = _XC['登記原因']

    # ── Step 1: 讀取所有列值（read_only 快速）──
    print('[reformat] Step 1/4: 讀取資料…')
    wb_r = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws_r = wb_r.active
    header = [ws_r.cell(1, c+1).value for c in range(ws_r.max_column)]
    all_rows = [list(row) for row in ws_r.iter_rows(min_row=2, values_only=True)]
    max_col = ws_r.max_column
    wb_r.close()
    print(f'  {len(all_rows):,} 列，{max_col} 欄，耗時 {time.time()-t0:.1f}s')

    # ── Step 2: 排序 ──
    print('[reformat] Step 2/4: 排序…')

    def sort_key(row):
        sec  = normalize_section(str(row[SEC_COL] or ''))
        no   = normalize_land_no(str(row[NO_COL] or ''))
        dt   = _roc_date_to_ad(row[DATE_COL]) or date(1900, 1, 1)
        try:
            seq = int(str(row[SEQ_COL] or '0').strip().lstrip('0') or '0')
        except ValueError:
            seq = 0
        return (sec, no, dt, seq)

    all_rows.sort(key=sort_key)
    print(f'  排序完成，耗時 {time.time()-t0:.1f}s')

    if dry_run:
        sold_c = sum(1 for r in all_rows if _sold_flag(r[SOLD_COL]))
        recent_c = sum(1 for r in all_rows
                       if not _sold_flag(r[SOLD_COL])
                       and str(r[REASON_COL] or '') == '買賣'
                       and _roc_date_to_ad(r[DATE_COL]) and _roc_date_to_ad(r[DATE_COL]) >= cutoff)
        return {'sorted': True, 'sold_grey': sold_c, 'recent_yellow': recent_c,
                'elapsed': time.time() - t0}

    # ── Step 3: 寫回 + 套用格式 ──
    print('[reformat] Step 3/4: 載入可寫 workbook…')
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    print(f'  載入完成，耗時 {time.time()-t0:.1f}s')

    print('[reformat] Step 4/4: 寫入排序後資料 + 格式…')
    sold_grey = 0
    recent_yellow = 0

    for i, row_data in enumerate(all_rows, start=2):
        sold_val   = row_data[SOLD_COL]
        reason_val = row_data[REASON_COL]
        date_val   = row_data[DATE_COL]

        fill, font = _row_fill_and_font(sold_val, reason_val, date_val, cutoff)
        if fill is SOLD_FILL_STYLE:
            sold_grey += 1
        elif fill is RECENT_FILL_STYLE:
            recent_yellow += 1

        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(i, c_idx)
            cell.value = val
            cell.fill  = fill
            cell.font  = font

        if i % 10000 == 0:
            pct = (i - 1) / len(all_rows) * 100
            print(f'  進度 {pct:.0f}%（{i-1:,}/{len(all_rows):,}）耗時 {time.time()-t0:.1f}s')

    # 清除多餘列（排序後列數不變，但以防萬一）
    if ws.max_row > len(all_rows) + 1:
        ws.delete_rows(len(all_rows) + 2, ws.max_row - len(all_rows) - 1)

    print(f'[reformat] 儲存中…')
    wb.save(EXCEL)
    wb.close()

    elapsed = time.time() - t0
    print(f'[reformat] 完成！已售灰={sold_grey} 近半年黃={recent_yellow} 耗時={elapsed:.1f}s')
    return {'sorted': True, 'sold_grey': sold_grey,
            'recent_yellow': recent_yellow, 'elapsed': elapsed}


# ── 實價提醒報表：標記已處理 ────────────────────────────────────────────────
#
# 實價提醒報表的正確定位：
#   = 買賣異動待調閱清單，來源為實價登錄比對結果
#   只追蹤「買賣交易」訊號，不追蹤增貸、抵押、他項權利設定等異動
#
# 增貸/抵押/他項異動 → 應歸屬「他項權利情報」或「地主金融壓力」模組（未來建置）
#
REALPRICE_REPORT = EXCEL_DIR / '最新完成版' / '實價提醒報表_最新完成版.xlsx'
_RP_COL_SECTION  = '地段'
_RP_COL_LAND_NO  = '地號'
_RP_COL_ACTION   = '建議動作'

# 只有這些登記原因的電傳才代表「買賣已調閱」，可標記實價提醒已處理
_RP_BUYSEL_REASONS = {'買賣'}


def mark_realprice_processed(all_inserted: list[dict], dry_run: bool = False) -> int:
    """
    電傳解析成功入庫後，將實價提醒報表中對應地號標記為「已調閱電傳」。

    觸發條件（三者同時成立）：
      1. reg_reason == '買賣'（買賣事件，對應實價登錄追蹤目的）
      2. 該地號成功寫入 SQLite land_master
      3. 實價提醒報表有該地號的未處理提醒（建議動作不含「已調閱」）

    不觸發條件：
      - 增貸、抵押權設定、他項設定、地目調整、繼承、分割等非買賣事件
      - 這些事件不代表買賣交易，不應從實價提醒清單移出

    回傳標記筆數。
    """
    import openpyxl

    # 只取買賣登記原因的地號
    buysel_keys = {
        (r['normalized_section'], r['normalized_land_no'])
        for r in all_inserted
        if (r.get('reg_reason') or '').strip() in _RP_BUYSEL_REASONS
    }
    if not REALPRICE_REPORT.exists() or not buysel_keys:
        return 0

    wb = openpyxl.load_workbook(str(REALPRICE_REPORT))
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    def col(name):
        return headers.index(name) + 1 if name in headers else None

    sec_col    = col(_RP_COL_SECTION)
    no_col     = col(_RP_COL_LAND_NO)
    action_col = col(_RP_COL_ACTION)
    if not (sec_col and no_col and action_col):
        return 0

    today = datetime.now().strftime('%Y-%m-%d')
    marked = 0
    for r in range(2, ws.max_row + 1):
        sec = normalize_section(str(ws.cell(r, sec_col).value or ''))
        no  = normalize_land_no(str(ws.cell(r, no_col).value or ''))
        if (sec, no) in buysel_keys:
            cur = str(ws.cell(r, action_col).value or '')
            if '已調閱' not in cur:
                ws.cell(r, action_col).value = f'已調閱電傳 {today}'
                marked += 1

    if marked and not dry_run:
        wb.save(str(REALPRICE_REPORT))
    wb.close()
    return marked


# ── Excel 主清冊同步 ─────────────────────────────────────────────────────────

def sync_excel(all_inserted: list[dict], dry_run: bool) -> dict:
    """
    同步更新主清冊 Excel。正確邏輯：

    新地主（買賣）：
      1. 找同地段地號的所有舊地主列
         → 已售出=1，備註追加「依新電傳確認移轉，已新增新地主列」，更新日期=今日
      2. 在「最後一筆舊地主列正下方」插入新地主列（insert_rows）
         → 已售出=0，更新日期=今日，填入電傳資料

    非買賣 / 新地主本人已在清冊：
      → 只更新次序/登記日期/登記原因/發生日期/更新日期，不新增列

    覆寫同一個正式主清冊檔案。
    回傳 {sold_rows, inserted_rows, updated_rows, preview, excel_path}
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    if not all_inserted:
        return {'sold_rows': 0, 'inserted_rows': 0, 'updated_rows': 0,
                'preview': [], 'excel_path': str(EXCEL_MASTER)}

    today = datetime.now().strftime('%Y/%m/%d')

    print(f'\n[Excel 同步] 讀取主清冊...')
    wb = openpyxl.load_workbook(str(EXCEL_MASTER), data_only=True)
    ws = wb.active

    # 建立 (norm_sec, norm_no) → [row_1based, ...] 排序後的列表
    # 以及 (norm_sec, norm_no, owner_name) → row_1based
    land_rows:  dict[tuple, list[int]] = {}
    owner_rows: dict[tuple, int]       = {}

    total_data_rows = ws.max_row - 1
    print(f'  {total_data_rows:,} 資料列，建立索引中...')

    for r in range(2, ws.max_row + 1):
        sec_v = ws.cell(r, _XC['地段'] + 1).value
        no_v  = ws.cell(r, _XC['地號'] + 1).value
        name  = str(ws.cell(r, _XC['所有權人'] + 1).value or '').strip()
        n_sec = normalize_section(str(sec_v or ''))
        n_no  = normalize_land_no(str(no_v or ''))
        lk = (n_sec, n_no)
        land_rows.setdefault(lk, []).append(r)
        if name:
            owner_rows[(n_sec, n_no, name)] = r

    DATA_FONT = Font(name='微軟正黑體', size=10)
    NEW_FILL  = PatternFill('solid', fgColor='E2EFDA')   # 淡綠：新增列（臨時，reformat 後依規則覆蓋）

    sold_rows     = 0
    inserted_rows = 0
    updated_rows  = 0
    share_chg_rows= 0
    preview       = []

    # 建立 land_match_key → perm info 的 lookup（供略過 Excel 時顯示原因）
    perm_info: dict[str, dict] = {}
    for rec in all_inserted:
        lmk = rec['land_match_key']
        if lmk not in perm_info:
            perm_info[lmk] = {
                'allow_sold_mark': rec.get('allow_sold_mark') is True,
                'reason': '',
            }

    # ── 按地號分組，收集同地號所有新地主，統一做 diff ──
    # all_inserted 可能含同一地號的多筆新地主（一份電傳可有多共有人）
    from collections import defaultdict
    by_land: dict[tuple, list[dict]] = defaultdict(list)
    for rec in all_inserted:
        lk = (rec['normalized_section'], rec['normalized_land_no'])
        by_land[lk].append(rec)

    # ops: list of (action, target_row_or_None, list[old_rows], rec)
    ops = []

    for lk, new_recs in by_land.items():
        n_sec, n_no = lk
        all_land_rows = land_rows.get(lk, [])
        land_label    = f'{new_recs[0]["section_raw"]} {new_recs[0]["land_no_raw"]}'

        # 同地號可能混有舊電傳(False)和新電傳(True)
        # 策略：
        #   舊電傳補歷史（sys_status='舊電傳補歷史'）→ 完全跳過 Excel（保護正式資料不被舊資料覆蓋）
        #   同日事件（sys_status='待人工確認'）       → 做 diff、可插入新地主，但不標已售
        #   新電傳（allow_sold_mark=True）            → 完整 diff + 標已售
        all_stale = [r for r in new_recs if r.get('sys_status') == '舊電傳補歷史']
        if all_stale and len(all_stale) == len(new_recs):
            # 全部都是舊電傳，完全跳過 Excel
            preview.append(f'  ⚠️  略過Excel更新  {land_label}  （全為舊電傳補歷史）')
            continue

        active_recs = [r for r in new_recs if r.get('allow_sold_mark') is True]
        # 同日事件：allow_sold_mark=False 但仍可插入 Excel（不標已售）
        same_day_recs = [r for r in new_recs if r.get('sys_status') == '待人工確認']
        # diff 優先用 active_recs（新電傳）；若只有同日事件，用 same_day_recs 做 diff（只插入不標售）
        diff_recs = active_recs if active_recs else same_day_recs
        mark_sold_allowed = bool(active_recs)   # 只有新電傳才允許標已售
        if not diff_recs:
            preview.append(f'  ⚠️  略過Excel更新  {land_label}  （無可用記錄）')
            continue
        new_recs = diff_recs

        # 把 Excel 舊地主整理成 dict 格式供 diff 使用
        old_recs_excel = []
        for r in all_land_rows:
            old_recs_excel.append({
                '_row':          r,
                'owner_name':    str(ws.cell(r, _XC['所有權人']       + 1).value or '').strip(),
                'owner_id_full': str(ws.cell(r, _XC['統一編號（完整）'] + 1).value or '').strip(),
                'owner_id_masked': str(ws.cell(r, _XC['統一編號（遮罩）'] + 1).value or '').strip(),
                'share_numer':   str(ws.cell(r, _XC['分子']           + 1).value or '').strip(),
                'share_denom':   str(ws.cell(r, _XC['分母']           + 1).value or '').strip(),
            })

        result = diff_owners(new_recs, old_recs_excel)

        # 判定插入位置：同地號 Excel 列中最大 row
        insert_after = max(all_land_rows) if all_land_rows else None

        # A. 消失的地主 → 標已售出（只有 allow_sold_mark=True 的電傳才執行）
        for old_r in result['disappeared']:
            row = old_r['_row']
            if mark_sold_allowed:
                ops.append(('mark_sold', row, old_r, None))
                sold_rows += 1
                preview.append(f'  ❌ 標已售出  {land_label}  row={row}  {old_r["owner_name"]}'
                               f'  （{_owner_id(old_r)}）')
            else:
                preview.append(f'  ⚠️  待人工確認  {land_label}  row={row}  {old_r["owner_name"]}'
                               f'  （同日事件，不標已售，待人工確認）')

        # B. 持分改變
        for old_r, new_r in result['share_changed']:
            row = old_r['_row']
            ops.append(('mark_share_changed', row, old_r, new_r))
            share_chg_rows += 1
            preview.append(f'  ⚠️  持分異動  {land_label}  row={row}  {old_r["owner_name"]}'
                           f'  {_share_str(old_r)} → {_share_str(new_r)}  備註待確認')

        # C. 未變動地主
        for old_r, new_r in result['unchanged']:
            preview.append(f'  ✅ 保持持有  {land_label}  row={old_r["_row"]}  {old_r["owner_name"]}'
                           f'  持分={_share_str(old_r)}')

        # D. 新增地主 → 插入在最後一筆下方
        for new_r in result['added']:
            ops.append(('insert_new', insert_after, None, new_r))
            inserted_rows += 1
            insert_pos = (insert_after + 1) if insert_after else '末尾'
            preview.append(f'  ➕ 新增地主  {land_label}  插入 row={insert_pos}  '
                           f'{new_r["owner_name"]}  持分={_share_str(new_r)}  '
                           f'event_key={new_r.get("event_key","")[:12]}...')
            # 確認下一個新增地主插入位置遞增
            if insert_after is not None:
                insert_after += 1

    # ── dry-run：只印不執行 ──
    if dry_run:
        print(f'\n  [dry-run] 共有人 diff 結果：')
        for line in preview:
            print(f'  {line}')
        print(f'\n  [dry-run] 摘要：標記已售出={sold_rows}  持分異動={share_chg_rows}'
              f'  插入新地主={inserted_rows}（未寫入）')
        wb.close()
        return {'sold_rows': sold_rows, 'inserted_rows': inserted_rows,
                'updated_rows': updated_rows, 'share_chg_rows': share_chg_rows,
                'preview': preview, 'excel_path': str(EXCEL_MASTER)}

    # ── 正式執行：先倒序 insert，再 mark_sold，最後 mark_share_changed ──
    def _build_new_row_data(rec: dict) -> dict:
        data = {
            _XC['更新日期']:          today,
            _XC['縣市']:              rec.get('city'),
            _XC['地區']:              rec.get('district'),
            _XC['地段']:              rec.get('section_raw'),
            _XC['小段']:              rec.get('sub_section') or '',
            _XC['地號']:              normalize_land_no(rec.get('land_no_raw') or ''),
            _XC['公告現值']:          rec.get('announced_value'),
            _XC['次序']:              rec.get('reg_seq'),
            _XC['登記日期']:          rec.get('reg_date'),
            _XC['登記原因']:          rec.get('reg_reason'),
            _XC['發生日期']:          rec.get('cause_date'),
            _XC['所有權人']:          rec.get('owner_name'),
            _XC['統一編號（遮罩）']:  rec.get('owner_id_masked'),
            _XC['統一編號（完整）']:  rec.get('owner_id_full'),
            _XC['郵遞區號']:          rec.get('postal_code'),
            _XC['住址']:              rec.get('address'),
            _XC['已售出']:            0,
            _XC['分母']:              rec.get('share_denom'),
            _XC['分子']:              rec.get('share_numer'),
            _XC['土地總坪數']:        rec.get('total_area_ping'),
            _XC['電話']:              rec.get('phone'),
            _XC['備註']:              rec.get('note'),
        }
        if EXCEL_HAS_SYS_COLS:
            data.update({
                _XC['系統處理狀態']: rec.get('sys_status') or '正常',
                _XC['系統處理備註']: rec.get('sys_note') or '',
                _XC['系統來源']:     rec.get('sys_source') or '電傳',
                _XC['系統更新時間']: rec.get('sys_updated_at') or today,
                _XC['系統批次ID']:   rec.get('sys_batch_id') or '',
            })
        return data

    # insert_new：倒序（避免行號偏移）
    insert_ops = [(op, tgt, old, rec) for op, tgt, old, rec in ops if op == 'insert_new']
    for op, insert_after, _, rec in sorted(insert_ops, key=lambda x: -(x[1] or 0)):
        nr = (insert_after + 1) if insert_after is not None else ws.max_row + 1
        if insert_after is not None:
            ws.insert_rows(nr)
        data = _build_new_row_data(rec)
        for ci, val in data.items():
            cell = ws.cell(nr, ci + 1)
            cell.value = val
            cell.font  = DATA_FONT
            cell.fill  = NEW_FILL
        for col_name in ('地號', '統一編號（遮罩）', '統一編號（完整）', '郵遞區號', '電話'):
            ws.cell(nr, _XC[col_name] + 1).number_format = '@'

    # mark_sold → 整列反灰
    for op, tgt, old_r, _ in ops:
        if op == 'mark_sold':
            old_note = str(ws.cell(tgt, _XC['備註'] + 1).value or '').strip()
            new_note = f'{old_note}｜依電傳確認移轉({today})已新增新地主'.lstrip('｜')
            ws.cell(tgt, _XC['已售出']   + 1).value = 1
            ws.cell(tgt, _XC['備註']     + 1).value = new_note
            ws.cell(tgt, _XC['更新日期'] + 1).value = today
            _apply_row_format(ws, tgt, SOLD_FILL_STYLE, SOLD_FONT_STYLE, ws.max_column)

    # mark_share_changed
    SHARE_CHG_FILL = PatternFill('solid', fgColor='FFF2CC')
    for op, tgt, old_r, new_r in ops:
        if op == 'mark_share_changed':
            old_note = str(ws.cell(tgt, _XC['備註'] + 1).value or '').strip()
            note_add = f'持分異動待確認({today})：{_share_str(old_r)}→{_share_str(new_r)}'
            ws.cell(tgt, _XC['備註']     + 1).value = f'{old_note}｜{note_add}'.lstrip('｜')
            ws.cell(tgt, _XC['更新日期'] + 1).value = today
            ws.cell(tgt, _XC['備註']     + 1).fill  = SHARE_CHG_FILL

    print(f'  儲存中...')
    wb.save(str(EXCEL_MASTER))
    wb.close()
    print(f'  已儲存：{EXCEL_MASTER}')

    return {'sold_rows': sold_rows, 'inserted_rows': inserted_rows,
            'updated_rows': updated_rows, 'share_chg_rows': share_chg_rows,
            'preview': preview, 'excel_path': str(EXCEL_MASTER)}


# ── 主流程 ──────────────────────────────────────────────────────────────────

def process_file(path: Path, con: sqlite3.Connection,
                 dry_run: bool, force: bool = False,
                 batch_id: str = '') -> dict:
    """
    處理單一電傳檔案。
    回傳 {status, records, inserted, sold_marked, stale_lands, error}
      stale_lands: list of {section_raw, land_no_raw, yaml_date, db_date}
                   — 因電傳日期舊於 DB 而跳過的地號
    force=True 時略過舊電傳保護，強制處理。
    """
    result = {'status': 'ok', 'records': [], 'inserted': [],
              'sold_marked': 0, 'stale_lands': [], 'error': ''}

    try:
        records = parse_transcript(path)
    except ParseError as e:
        result['status'] = 'error'
        result['error']  = str(e)
        return result

    result['records'] = records
    cur = con.cursor()

    # ── 舊電傳防覆蓋：計算每個地號的 allow_sold_mark ──
    perm = _compute_sold_mark_permission(records, cur)
    for lmk, info in perm.items():
        if not info['allow_sold_mark']:
            result['stale_lands'].append(info)
    # 將 allow_sold_mark 及系統判定欄位附加到每筆 record
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for rec in records:
        p = perm[rec['land_match_key']]
        rec['allow_sold_mark'] = p['allow_sold_mark']
        # force 模式覆蓋 sys_status
        if force and not p['allow_sold_mark']:
            rec['sys_status']     = 'force匯入'
            rec['sys_note']       = f'--force 旗標強制匯入（原狀態：{p["sys_status"]}）'
        else:
            rec['sys_status']     = p['sys_status']
            rec['sys_note']       = p['sys_note']
        rec['sys_source']     = '電傳'
        rec['sys_updated_at'] = now_str
        rec['sys_batch_id']   = batch_id

    for rec in records:
        ek = rec['event_key']
        if not ek:
            result['error'] += f'owner={rec["owner_name"]} event_key 為空（跳過）; '
            continue

        # 是否已存在
        exists = cur.execute('SELECT id FROM land_master WHERE event_key=?', (ek,)).fetchone()
        if exists:
            if not dry_run:
                cur.execute(
                    "INSERT INTO transcript_import_log (filename,event_key,owner_name,land_key,reg_reason,status) VALUES (?,?,?,?,?,?)",
                    (path.name, ek, rec['owner_name'], rec['land_match_key'], rec.get('reg_reason'), 'skipped')
                )
            continue

        # 新增
        if not dry_run:
            cur.execute(INSERT_SQL, rec)
            cur.execute(
                "INSERT INTO transcript_import_log (filename,event_key,owner_name,land_key,reg_reason,status) VALUES (?,?,?,?,?,?)",
                (path.name, ek, rec['owner_name'], rec['land_match_key'], rec.get('reg_reason'), 'inserted')
            )
        result['inserted'].append(rec)

    # ── SQLite 共有人 diff：依新電傳名單，標記消失的地主 is_sold=1 ──
    # 只有 allow_sold_mark=True 的地號才執行
    if not dry_run:
        from collections import defaultdict
        by_land_new: dict[str, list[dict]] = defaultdict(list)
        for rec in result['inserted']:
            if rec.get('allow_sold_mark') is True:   # 保守模式：缺失/False/None 一律不標已售
                by_land_new[rec['land_match_key']].append(rec)

        for lmk, new_recs in by_land_new.items():
            # 取 DB 現有地主（is_sold=0）
            old_db = cur.execute(
                """SELECT owner_key, owner_name, owner_id_masked, owner_id_full,
                          share_numer, share_denom
                   FROM land_master
                   WHERE land_match_key=? AND is_sold=0""",
                (lmk,)
            ).fetchall()
            old_recs_db = [
                {'owner_key': r[0], 'owner_name': r[1],
                 'owner_id_masked': r[2], 'owner_id_full': r[3],
                 'share_numer': str(r[4] or ''), 'share_denom': str(r[5] or '')}
                for r in old_db
            ]
            diff = diff_owners(new_recs, old_recs_db)

            # 消失的地主 → is_sold=1
            for old_r in diff['disappeared']:
                now_str = result['inserted'][0]['updated_at']
                updated = cur.execute(
                    """UPDATE land_master SET is_sold=1, updated_at=?
                       WHERE land_match_key=? AND owner_key=? AND is_sold=0""",
                    (now_str, lmk, old_r['owner_key'])
                ).rowcount
                result['sold_marked'] += updated
                if updated:
                    cur.execute(
                        "INSERT INTO transcript_import_log "
                        "(filename,event_key,owner_name,land_key,reg_reason,status) VALUES (?,?,?,?,?,?)",
                        (path.name, '', old_r['owner_name'], lmk, '共有人消失→標已售出', 'sold_marked')
                    )

    if not dry_run:
        con.commit()

    return result


def main():
    ap = argparse.ArgumentParser(description='電傳 YAML → SQLite')
    ap.add_argument('--dry-run', action='store_true', help='只解析，不寫入，不移動檔案')
    ap.add_argument('--force',   action='store_true', help='略過舊電傳保護，強制處理（危險）')
    ap.add_argument('--db', default=str(DB_PATH))
    args = ap.parse_args()

    dry_run = args.dry_run
    force   = args.force
    mode_label = '【DRY-RUN】' if dry_run else '【正式執行】'
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f'{mode_label} process_land_transcripts.py')
    print(f'時間：{ts}')
    print(f'DB  ：{args.db}')
    print(f'收件：{INBOX_DIR}')

    # 建立 log 表
    con = sqlite3.connect(args.db)
    con.execute(DDL_LOG)
    con.commit()

    # 掃描待解析資料夾
    files = sorted(
        p for p in INBOX_DIR.iterdir()
        if p.is_file() and (p.suffix in ('.yaml', '.txt') or p.name.endswith('.yaml.txt'))
    )
    print(f'\n掃描到 {len(files)} 個電傳檔案')
    if not files:
        print('  無待解析檔案，結束。')
        con.close()
        return

    # 產生本次執行的批次 ID
    batch_seq   = 1
    batch_id    = f'TXN_{ts[:10].replace("-","")}_{ts[11:16].replace(":",""):s}_{batch_seq:03d}'

    total_inserted   = 0
    total_sold       = 0
    total_skipped    = 0
    total_error      = 0
    total_stale      = 0
    all_stale_info   = []
    dry_run_samples  = []
    all_inserted_recs = []   # 收集所有成功插入的 rec，供 Excel 同步用

    for f in files:
        print(f'\n── {f.name} ──')
        result = process_file(f, con, dry_run=dry_run, force=force, batch_id=batch_id)

        n_rec     = len(result['records'])
        n_insert  = len(result['inserted'])
        n_sold    = result['sold_marked']
        # 跳過 = 解析成功但 event_key 已存在
        n_skipped = n_rec - n_insert if result['status'] == 'ok' else 0

        if result['status'] == 'error':
            total_error += 1
            print(f'  ❌ 解析失敗：{result["error"]}')
            if not dry_run:
                # 原始檔移到電傳錯誤/，內容與檔名完整保留
                dest = ERROR_DIR / f.name
                shutil.move(str(f), str(dest))
                # 錯誤說明另寫 .error.log，不覆蓋原始 YAML
                error_log = ERROR_DIR / (f.name + '.error.log')
                error_log.write_text(
                    f'時間：{ts}\n'
                    f'檔案：{f.name}\n'
                    f'錯誤原因：{result["error"]}\n',
                    encoding='utf-8'
                )
                con.execute(
                    "INSERT INTO transcript_import_log (filename,status,error_msg) VALUES (?,?,?)",
                    (f.name, 'error', result['error'])
                )
                con.commit()
                print(f'  → 移至 電傳錯誤/  （原始檔保留，錯誤說明寫入 {f.name}.error.log）')
        else:
            total_inserted += n_insert
            total_sold     += n_sold
            total_skipped  += n_skipped
            stale_lands = result.get('stale_lands', [])
            n_stale = len(stale_lands)
            total_stale    += n_stale
            if not force:
                all_stale_info.extend(stale_lands)
            print(f'  解析筆數：{n_rec}　新增：{n_insert}　SKIP：{n_skipped}　標記已售出：{n_sold}'
                  + (f'　⚠️ SQLite補歷史（不更新Excel）：{n_stale}地號' if n_stale and not force else '')
                  + (f'　[--force略過警告]' if n_stale and force else ''))
            if not force:
                for info in stale_lands:
                    print(f'    ⚠️  {info["section_raw"]} {info["land_no_raw"]}'
                          f'  {info["reason"]}')
            for rec in result['inserted'][:3]:
                participated = '（已售判定）' if rec.get('allow_sold_mark') is True else '（補歷史/不標已售）'
                print(f'    ✅ [{rec["reg_reason"] or "—"}] {rec["section_raw"]} {rec["land_no_raw"]} → {rec["owner_name"]}'
                      f'  {participated}  [{rec.get("sys_status","—")}]')

            if dry_run:
                dry_run_samples.extend(result['records'][:5])
            else:
                all_inserted_recs.extend(result['inserted'])
                # 移到已完成
                dest = DONE_DIR / f.name
                if dest.exists():
                    dest = DONE_DIR / f'{f.stem}_{ts.replace(":","").replace(" ","_")}{f.suffix}'
                shutil.move(str(f), str(dest))
                print(f'  → 移至 電傳已完成/')

                # Telegram
                if result['inserted']:
                    msg = build_tg_message(f.name, result['records'],
                                           result['inserted'], n_sold)
                    _send_telegram(msg)
                    print(f'  [TG] 摘要已推播')

    con.close()

    # ── Excel 同步（所有檔案處理完後一次執行）──
    excel_result = {'sold_rows': 0, 'updated_rows': 0, 'inserted_rows': 0,
                    'share_chg_rows': 0, 'excel_path': str(EXCEL_MASTER)}
    if all_inserted_recs or dry_run:
        recs_for_sync = all_inserted_recs if not dry_run else (dry_run_samples or [])
        if recs_for_sync or all_inserted_recs:
            excel_result = sync_excel(
                all_inserted_recs if not dry_run else dry_run_samples,
                dry_run=dry_run
            )

    # ── 全表排序 + 格式化 ──
    fmt_result = {'sorted': False, 'sold_grey': 0, 'recent_yellow': 0}
    if all_inserted_recs and not dry_run:
        print('\n[reformat] 重新套用格式與排序…')
        fmt_result = reformat_and_sort_master(dry_run=False)

    # ── 實價提醒報表：標記已處理（只處理買賣事件）──
    rp_marked = 0
    if all_inserted_recs and not dry_run:
        rp_marked = mark_realprice_processed(all_inserted_recs, dry_run=False)
        if rp_marked:
            print(f'\n[實價報表] 已標記 {rp_marked} 筆為「已調閱電傳」：{REALPRICE_REPORT.name}')

    print(f'\n{"─"*50}')
    if dry_run:
        print(f'{mode_label} 摘要')
        print(f'  可解析檔數          ：{len(files) - total_error}')
        print(f'  可新增 SQLite 事件  ：{total_inserted}')
        print(f'  SKIP（已存在）      ：{total_skipped}')
        print(f'  會標記已售出（SQLite）：（dry-run 不執行比對）')
        print(f'  解析失敗            ：{total_error}')
        print(f'\nExcel 同步預覽（dry-run）：')
        print(f'  消失地主→標已售出  ：{excel_result["sold_rows"]}')
        print(f'  插入新地主列        ：{excel_result["inserted_rows"]}')
        print(f'  持分異動→備註       ：{excel_result.get("share_chg_rows",0)}')
        print(f'  現有列保持不變      ：（見上方 diff 明細）')
        print(f'\n檔案移動規則（dry-run 不執行）：')
        print(f'  成功 → {DONE_DIR}')
        print(f'  失敗 → {ERROR_DIR}  + .error 說明檔')
        if dry_run_samples:
            print(f'\n前 {min(5,len(dry_run_samples))} 筆解析案例：')
            for i, rec in enumerate(dry_run_samples[:5], 1):
                print(f'  [{i}] {rec["city"]}{rec["district"]} {rec["section_raw"]} {rec["land_no_raw"]}'
                      f' / {rec["owner_name"]} / {rec.get("reg_reason","—")} / {rec.get("reg_date","—")}'
                      f' / event_key={rec["event_key"][:8]}...')
    else:
        print(f'完成  批次ID={batch_id}')
        if all_stale_info:
            print(f'\n⚠️  舊電傳防覆蓋 — 以下地號已補入 SQLite 歷史，但不更新 Excel（共 {total_stale} 個）：')
            for info in all_stale_info:
                print(f'  ⚠️  {info["section_raw"]} {info["land_no_raw"]}'
                      f'  {info["reason"]}')
            print()
        print(f'SQLite：')
        print(f'  實際新增事件        ：{total_inserted}')
        print(f'  SKIP（已存在）      ：{total_skipped}')
        print(f'  標記已售出          ：{total_sold}')
        print(f'  舊電傳跳過地號      ：{total_stale}')
        print(f'  解析失敗            ：{total_error}')
        print(f'Excel 主清冊同步：')
        print(f'  消失地主→標已售出  ：{excel_result["sold_rows"]}')
        print(f'  插入新地主列        ：{excel_result["inserted_rows"]}')
        print(f'  持分異動→備註       ：{excel_result.get("share_chg_rows",0)}')
        print(f'  主清冊路徑          ：{excel_result["excel_path"]}')
        print(f'格式化：')
        print(f'  已售出反灰列數      ：{fmt_result["sold_grey"]}')
        print(f'  近半年買賣反黃列數  ：{fmt_result["recent_yellow"]}')
        print(f'  全表排序            ：{"✅" if fmt_result["sorted"] else "跳過"}')
        print(f'實價提醒報表：')
        print(f'  標記已調閱電傳      ：{rp_marked}')
        print(f'系統：')
        print(f'  批次ID              ：{batch_id}')


if __name__ == '__main__':
    main()
